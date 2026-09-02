# src/callbacks/screener_callbacks.py
from dash import Input, Output, State, callback_context, no_update, html, dcc, ALL
from dash.exceptions import PreventUpdate
import dash_bootstrap_components as dbc
import plotly.graph_objects as go
import gc
import dash
from src.app_instance import app
from src.backend.data_loader import load_market_data, load_financial_data, get_latest_snapshot, get_snapshot_df
from src.backend.quant_engine import calculate_all_scores
from src.backend.quant_engine_strategies import run_strategy
from src.backend.data_loader import load_financial_data
from src.constants.gics_translation import translate_gics_industry, translate_gics_sector
import logging
import numpy as np
import pandas as pd
from src.backend.quant_engine_strategies import (
    apply_value_filter, apply_turnaround_filter, apply_quality_filter,
    apply_garp_filter, apply_dividend_filter, apply_piotroski_filter,
    apply_canslim_filter, apply_garp_filter
)

# Import column definitions để gộp vào callback chính (tránh double-render)
from src.callbacks.column_callbacks import (
    FIXED_COLS, FILTER_TO_COLDEF, STRATEGY_FILTER_IDS, STRATEGY_DIRECT_COLS
)

import pandas as pd
import os

# =====================================================================
# 1. ĐỌC FILE CSV 1 LẦN DUY NHẤT Ở NGOÀI CALLBACK (Global Scope)
# =====================================================================
import os as _os

def _load_comp_info():
    """Load COMP INFO với multiple path fallback cho cả local và HF."""
    # Danh sách path thử theo thứ tự ưu tiên
    _base = _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__)))
    candidates = [
        "data/raw/COMP INFO.csv", # Ưu tiên path tương đối (local dev)
        _os.path.join(_base, "data", "raw", "COMP INFO.csv"),
        _os.path.join(_base, "data", "raw", "COMP_INFO.csv"),
        "/data/raw/COMP INFO.csv",          # HF persistent storage
        "/app/data/raw/COMP INFO.csv",      # HF Docker mount
    ]
    for path in candidates:
        if _os.path.exists(path):
            try:
                df = pd.read_csv(path)
                if 'symbol' in df.columns:
                    df.rename(columns={'symbol': 'Ticker'}, inplace=True)
                # Chuẩn hóa Ticker: bỏ đuôi .HM/.HN/.HNO nếu có
                if 'Ticker' in df.columns:
                    df['Ticker'] = df['Ticker'].str.replace(
                        r'\.(HNO|HN|HM)$', '', regex=True
                    ).str.strip()
                print(f"✅ Loaded COMP INFO từ: {path} ({len(df)} rows)")
                return df
            except Exception as e:
                print(f"⚠️ Lỗi đọc {path}: {e}")
    print("⚠️ Không tìm thấy COMP INFO.csv ở bất kỳ path nào — dùng snapshot fallback")
    return pd.DataFrame()

df_comp_info = _load_comp_info()

# 🟢 IMPORT TOÀN BỘ CÁC THÔNG SỐ TỪ QUANT ENGINE
from src.backend.quant_engine_strategies import (
    # 1. VALUE
    VALUE_THRESHOLDS, VALUE_IDX_CURRENT_RATIO_MIN, VALUE_IDX_EPS_GROWTH_5Y_MIN,
    VALUE_IDX_PE_MAX, VALUE_IDX_PB_MAX, VALUE_IDX_DEBT_TO_WC_MAX, VALUE_IDX_NET_INCOME_MIN,

    # 2. TURNAROUND
    TURNAROUND_THRESHOLDS, TURNAROUND_IDX_PE_HIST_NORM_MAX, TURNAROUND_IDX_OPERATING_MARGIN_MIN,
    TURNAROUND_IDX_PEG_MIN, TURNAROUND_IDX_PEG_MAX,

    # 3. QUALITY
    QUALITY_THRESHOLDS, QUALITY_IDX_ROE_MIN, QUALITY_IDX_GROSS_MARGIN_MIN,
    QUALITY_IDX_RE_GROWTH_MIN, QUALITY_IDX_FCF_MARGIN_MIN,

    # 4. GARP
    GARP_THRESHOLDS, GARP_IDX_EPS_GROWTH_MIN, GARP_IDX_EPS_GROWTH_MAX, GARP_IDX_PE_MAX,
    GARP_IDX_PEG_MIN, GARP_IDX_PEG_MAX, GARP_IDX_D_E_MAX, GARP_IDX_SGR_MIN_PCT, GARP_IDX_MC_QUANTILE,

    # 5. DIVIDEND
    DIVIDEND_THRESHOLDS, DIV_IDX_MC_QUANTILE, DIV_IDX_YIELD_MIN, DIV_IDX_PAYOUT_MAX,

    # 6. PIOTROSKI
    PIOTROSKI_THRESHOLDS, PIOTROSKI_IDX_F_MIN,

    # 7. CANSLIM
    CANSLIM_THRESHOLDS, CANSLIM_IDX_EPS_GROWTH_Q_MIN, CANSLIM_IDX_REV_GROWTH_Q_MIN,
    CANSLIM_IDX_EPS_GROWTH_Y_MIN, CANSLIM_IDX_ROE_MIN, CANSLIM_IDX_RS_MIN, CANSLIM_IDX_VOL_MULT,
    CANSLIM_IDX_AVG_VOL_MIN, CANSLIM_IDX_QUICK_RATIO_MIN, CANSLIM_IDX_DEBT_EQUITY_MAX,

    # 8. FISHER (GROWTH)
    FISHER_THRESHOLDS, FISHER_IDX_REV_GROWTH_5Y_MIN, FISHER_IDX_DILUTION_RATE_MAX,
    FISHER_IDX_ROE_MIN, FISHER_IDX_OPEX_EFF_MAX, FISHER_IDX_TURNOVER_MIN, FISHER_IDX_REINVEST_MIN
)

logger = logging.getLogger(__name__)

# ============================================================================
# TỪ ĐIỂN MAPPING VÀ GROUPING (TÓM TẮT) - ĐƠN VỊ TRIỆU VND
# ============================================================================
FINANCIAL_UI_MAP = {
    # ------------------ KẾT QUẢ KINH DOANH (IS) ------------------
    "Revenue from Business Activities - Total_x": {"name": "Doanh thu thuần", "group": "1. Kết quả kinh doanh"},
    "Cost of Revenues - Total": {"name": "Giá vốn hàng bán", "group": "1. Kết quả kinh doanh"},
    "Gross Profit - Industrials/Property - Total": {"name": "Lợi nhuận gộp", "group": "1. Kết quả kinh doanh"},
    "Operating Expenses - Total": {"name": "Tổng chi phí hoạt động", "group": "1. Kết quả kinh doanh"},
    "Earnings before Interest Taxes Depreciation & Amortization": {"name": "EBITDA", "group": "1. Kết quả kinh doanh"},
    "Earnings before Interest & Taxes (EBIT)": {"name": "EBIT", "group": "1. Kết quả kinh doanh"},
    "Income before Taxes": {"name": "Lợi nhuận trước thuế", "group": "1. Kết quả kinh doanh"},
    "Income Taxes": {"name": "Thuế TNDN", "group": "1. Kết quả kinh doanh"},
    "Net Income after Minority Interest": {"name": "LNST của cổ đông công ty mẹ", "group": "1. Kết quả kinh doanh"},
    "EPS - Basic - excl Extraordinary Items, Common - Total": {"name": "EPS Cơ bản", "group": "1. Kết quả kinh doanh"},
    "DPS - Common - Net - Issue - By Announcement Date": {"name": "Cổ tức mỗi cổ phiếu (DPS)",
                                                          "group": "1. Kết quả kinh doanh"},

    # ------------------ BẢNG CÂN ĐỐI KẾ TOÁN (BS) ------------------
    "Cash & Cash Equivalents - Total_x": {"name": "Tiền & Tương đương tiền", "group": "2. Bảng cân đối kế toán"},
    "Short-Term Investments - Total": {"name": "Đầu tư tài chính ngắn hạn", "group": "2. Bảng cân đối kế toán"},
    "Trade Accounts & Trade Notes Receivable - Net": {"name": "Phải thu khách hàng",
                                                      "group": "2. Bảng cân đối kế toán"},
    "Inventories - Total": {"name": "Hàng tồn kho", "group": "2. Bảng cân đối kế toán"},
    "Total Current Assets": {"name": "TỔNG TÀI SẢN NGẮN HẠN", "group": "2. Bảng cân đối kế toán"},
    "Property Plant & Equipment - Net - Total": {"name": "Tài sản cố định (Net)", "group": "2. Bảng cân đối kế toán"},
    "Investments - Long-Term": {"name": "Đầu tư dài hạn", "group": "2. Bảng cân đối kế toán"},
    "Total Assets": {"name": "TỔNG TÀI SẢN", "group": "2. Bảng cân đối kế toán"},

    "Trade Accounts & Trade Notes Payable - Short-Term": {"name": "Phải trả người bán",
                                                          "group": "2. Bảng cân đối kế toán"},
    "Short-Term Debt & Current Portion of Long-Term Debt": {"name": "Nợ vay ngắn hạn",
                                                            "group": "2. Bảng cân đối kế toán"},
    "Total Current Liabilities": {"name": "TỔNG NỢ NGẮN HẠN", "group": "2. Bảng cân đối kế toán"},
    "Debt - Long-Term - Total": {"name": "Nợ vay dài hạn", "group": "2. Bảng cân đối kế toán"},
    "Total Liabilities": {"name": "TỔNG NỢ PHẢI TRẢ", "group": "2. Bảng cân đối kế toán"},
    "Common Equity - Total": {"name": "Vốn góp", "group": "2. Bảng cân đối kế toán"},
    "Retained Earnings - Total": {"name": "Lợi nhuận giữ lại", "group": "2. Bảng cân đối kế toán"},
    "Total Shareholders' Equity incl Minority Intr & Hybrid Debt": {"name": "TỔNG VỐN CHỦ SỞ HỮU",
                                                                    "group": "2. Bảng cân đối kế toán"},

    # ------------------ LƯU CHUYỂN TIỀN TỆ (CF) ------------------
    "Net Cash Flow from Operating Activities": {"name": "Dòng tiền từ HĐKD (CFO)", "group": "3. Lưu chuyển tiền tệ"},
    "Capital Expenditures - Total_x": {"name": "Chi phí vốn (CAPEX)", "group": "3. Lưu chuyển tiền tệ"},
    "Net Cash Flow from Investing Activities": {"name": "Dòng tiền từ HĐ Đầu tư (CFI)",
                                                "group": "3. Lưu chuyển tiền tệ"},
    "Dividends Paid - Cash - Total - Cash Flow_x": {"name": "Cổ tức đã trả bằng tiền",
                                                    "group": "3. Lưu chuyển tiền tệ"},
    "Free Cash Flow": {"name": "DÒNG TIỀN TỰ DO (FCF)", "group": "3. Lưu chuyển tiền tệ"},
    "Net Cash - Ending Balance": {"name": "Tiền & TĐ tiền cuối kỳ", "group": "3. Lưu chuyển tiền tệ"}
}

# ============================================================================
# TỪ ĐIỂN CHỈ SỐ TÀI CHÍNH (METRICS MAPPING)
# ============================================================================
METRICS_UI_MAP = {
    # 1. Per Share (Dữ liệu gốc và tự tính)
    "EPS": {"name": "EPS Cơ bản (VND)", "group": "1"},
    "BVPS": {"name": "Giá trị sổ sách - BVPS (VND)", "group": "1"},
    "DPS - Common - Net - Issue - By Announcement Date": {"name": "Cổ tức mỗi CP - DPS (VND)", "group": "1"},

    # 2. Sinh lời (Profitability)
    "ROE": {"name": "ROE (%)", "group": "2"},
    "ROA": {"name": "ROA (%)", "group": "2"},
    "Gross Margin": {"name": "Biên Lợi nhuận gộp (%)", "group": "2"},
    "Net Margin": {"name": "Biên Lợi nhuận ròng (%)", "group": "2"},
    "EBIT Margin": {"name": "Biên EBIT (%)", "group": "2"},

    # 3. Thanh khoản (Liquidity)
    "Current Ratio": {"name": "Thanh toán hiện hành (Lần)", "group": "3"},
    "Quick Ratio": {"name": "Thanh toán nhanh (Lần)", "group": "3"},
    "Cash Ratio": {"name": "Thanh toán tiền mặt (Lần)", "group": "3"},

    # 4. Đòn bẩy (Leverage)
    "Debt to Equity": {"name": "Nợ vay / Vốn CSH (Lần)", "group": "4"},
    "Debt to Assets": {"name": "Nợ vay / Tổng tài sản (Lần)", "group": "4"},
    "Equity Multiplier": {"name": "Đòn bẩy tài chính (Lần)", "group": "4"},

    # 5. Hiệu quả (Efficiency)
    "Asset Turnover": {"name": "Vòng quay Tổng tài sản (Vòng)", "group": "5"},
    "Inventory Turnover": {"name": "Vòng quay Hàng tồn kho (Vòng)", "group": "5"},

    # 6. Tăng trưởng (Growth)
    "Revenue Growth": {"name": "Tăng trưởng Doanh thu (%)", "group": "6"},
    "Net Income Growth": {"name": "Tăng trưởng Lợi nhuận ròng (%)", "group": "6"}
}


# ============================================================================
# HELPER FUNCTIONS (Moved from detail_tabs_callbacks.py)
# ============================================================================

def fmt_number(val, prefix="", suffix=""):
    if val is None or val == "" or (isinstance(val, float) and (val != val)):  # Check for NaN
        return "---"
    try:
        return f"{prefix}{val:,.0f}{suffix}"
    except Exception as _e:  # noqa: audit-fix bare-except
        logger.debug(f"Suppressed non-critical error at src/callbacks/screener_callbacks.py:200: {_e}")
        return "---"


def fmt_decimal(val, decimals=2, suffix=""):
    if val is None or val == "" or (isinstance(val, float) and (val != val)):
        return "---"
    try:
        return f"{val:.{decimals}f}{suffix}"
    except Exception as _e:  # noqa: audit-fix bare-except
        logger.debug(f"Suppressed non-critical error in {__name__} near line 210: {_e}")
        return "---"


def fmt_percent(val):
    if val is None or val == "" or (isinstance(val, float) and (val != val)):
        return "---"
    try:
        return f"{val:.2f}%"
    except Exception as _e:  # noqa: audit-fix bare-except
        logger.debug(f"Suppressed non-critical error in {__name__} near line 220: {_e}")
        return "---"


def get_percent_style(val):
    """Return style based on percentage value"""
    if val is None or val == "" or (isinstance(val, float) and (val != val)):
        return {"color": "#c9d1d9"}  # Grey
    try:
        if val > 0:
            return {"color": "#3fb950", "fontWeight": "bold"}  # Green
        elif val < 0:
            return {"color": "#f85149", "fontWeight": "bold"}  # Red
        else:
            return {"color": "#e6edf3"}  # White
    except Exception as _e:  # noqa: audit-fix bare-except
        logger.debug(f"Suppressed non-critical error in {__name__} near line 236: {_e}")
        return {"color": "#c9d1d9"}


def get_trend_style(current_price, sma_value):
    """Return style for SMA comparison"""
    if sma_value is None or sma_value == "---" or current_price is None:
        return {"color": "#c9d1d9"}, "---"

    try:
        if current_price > sma_value:
            return {"color": "#3fb950", "fontWeight": "bold"}, "Tăng (Giá > SMA)"
        elif current_price < sma_value:
            return {"color": "#f85149", "fontWeight": "bold"}, "Giảm (Giá < SMA)"
        else:
            return {"color": "#e6edf3"}, "Đi ngang"
    except Exception as _e:  # noqa: audit-fix bare-except
        logger.debug(f"Suppressed non-critical error in {__name__} near line 253: {_e}")
        return {"color": "#c9d1d9"}, "---"

def _build_col_defs(active_filters, strategy_id, trading_mode="investing"):
    """Xây dựng columnDefs từ active_filters + strategy + trading_mode."""
    from src.callbacks.column_callbacks import INVESTING_MODE_COLS, TRADING_MODE_COLS
    
    seen_fields = {c["field"] for c in FIXED_COLS}
    dynamic_cols = []
    af = active_filters or {}

    # Sửa đoạn if not af and not strategy_id: thành thế này:
    if not af and not strategy_id:
        from src.callbacks.column_callbacks import INVESTING_MODE_COLS, TRADING_MODE_COLS, ALL_MARKET_COLS # ← Thêm ALL_MARKET_COLS
        
        if trading_mode == "trading":
            default_mode_cols = TRADING_MODE_COLS
        elif trading_mode == "all_market":
            default_mode_cols = ALL_MARKET_COLS
        else:
            default_mode_cols = INVESTING_MODE_COLS
            
        for col in default_mode_cols:
            if col["field"] not in seen_fields:
                dynamic_cols.append(col)
                seen_fields.add(col["field"])
        return FIXED_COLS + dynamic_cols

    # Nếu có filter/strategy: giữ nguyên logic cũ
    for filter_id in af:
        if filter_id not in FILTER_TO_COLDEF:
            continue
        col = FILTER_TO_COLDEF[filter_id]
        if col["field"] not in seen_fields:
            dynamic_cols.append(col)
            seen_fields.add(col["field"])

    if strategy_id and strategy_id in STRATEGY_FILTER_IDS:
        for filter_id in STRATEGY_FILTER_IDS[strategy_id]:
            if filter_id not in FILTER_TO_COLDEF:
                continue
            col = FILTER_TO_COLDEF[filter_id]
            if col["field"] not in seen_fields:
                dynamic_cols.append(col)
                seen_fields.add(col["field"])
        for col in STRATEGY_DIRECT_COLS.get(strategy_id, []):
            if col["field"] not in seen_fields:
                dynamic_cols.append(col)
                seen_fields.add(col["field"])

    return FIXED_COLS + dynamic_cols


def _add_forward_pe(df):
    """Tính Forward P/E inline — tránh callback riêng gây double-render."""
    try:
        if 'Forward P/E *' in df.columns:
            return df
        if not all(c in df.columns for c in ['EPS', 'EPS Growth YoY (%)', 'Price Close']):
            return df
        eps    = pd.to_numeric(df['EPS'], errors='coerce')
        growth = pd.to_numeric(df['EPS Growth YoY (%)'], errors='coerce').clip(-90, 500)
        price  = pd.to_numeric(df['Price Close'], errors='coerce')
        fwd_eps = eps * (1 + growth / 100)
        df['Forward P/E *'] = np.where(
            (fwd_eps > 0) & (price > 0),
            (price / fwd_eps).round(2),
            np.nan
        )
    except Exception:
        pass
    return df
def _add_profile_match_col(df: pd.DataFrame, profile: dict) -> pd.DataFrame:
    risk       = profile.get("risk", 3)
    strategies = profile.get("strategy", [])
    score      = pd.Series(0, index=df.index)

    if "ROE (%)" in df.columns:
        roe    = pd.to_numeric(df["ROE (%)"], errors="coerce").fillna(0)
        score += (roe >= 15).astype(int) * 2
        score -= (roe < 0).astype(int) * 3

    if "D/E" in df.columns:
        de     = pd.to_numeric(df["D/E"], errors="coerce").fillna(999)
        score += (de <= 1.5).astype(int)
        score -= (de > 3).astype(int) * 2

    if "VGM Score" in df.columns:
        score += df["VGM Score"].map(
            {"A": 3, "B": 2, "C": 1, "D": 0, "F": -1}
        ).fillna(0)

    if risk <= 2 and "RSI_14" in df.columns:
        rsi    = pd.to_numeric(df["RSI_14"], errors="coerce").fillna(50)
        score -= (rsi > 70).astype(int)

    if "trading" in strategies and "Perf_1M" in df.columns:
        p1m    = pd.to_numeric(df["Perf_1M"], errors="coerce").fillna(0)
        score += (p1m > 0).astype(int)

    df["_profile_match"] = pd.cut(
        score,
        bins=[-999, 1, 4, 999],
        labels=["✗", "✓", "✓✓"]
    ).astype(str)
    df.loc[df["_profile_match"] == "nan", "_profile_match"] = "–"
    return df



# ============================================================================
# CALLBACK: MAIN SCREENER TABLE UPDATE (rowData + columnDefs trong 1 lần)
# ============================================================================
@app.callback(
    [Output("screener-table", "rowData",    allow_duplicate=True),
     Output("screener-table", "columnDefs", allow_duplicate=True),
     Output("result-count",   "children"),
     Output("filter-stats",   "children"),
     # >>> THÊM 2 OUTPUT NÀY CHO TOAST CẢNH BÁO <<<
     Output("api-error-toast", "is_open"),
     Output("api-error-toast", "children"),
     Output("filter-null-alert", "children")],
    [
        # ── TRIGGERS chính (thay đổi những thứ này → chạy filter) ──
        Input("btn-reset",                  "n_clicks"),
        Input("search-ticker-input",        "value"),
        Input("strategy-preset-dropdown",   "value"),
        Input("filter-all-industry",        "value"),
        Input("active-filters-store",       "data"),   # ← nguồn sự thật duy nhất
        Input("filter-sub-industry",        "value"),
        Input("filter-exchange",             "value"),   # ← lọc theo sàn
        Input("filter-year-store",          "data"),   # ← lọc theo năm
        Input("include-null-data-store",    "data"),   # ← chế độ incl. null
        Input("trading-mode-store", "data"),   # ← thêm sau filter-year-store
        Input("investor-profile-store", "data"),

        Input("filter-index", "value"),# >>> THÊM INPUT CỦA DROPDOWN CHỈ SỐ <<<
        Input("nav-input", "value"), # 🟢 BẠN THÊM DÒNG NÀY VÀO
        Input("auth-store", "data"),   # ← THÊM DÒNG NÀY: login/logout → build lại rowData ngay
    ],
    [
        # ── STATE: đọc giá trị hiện tại của từng store khi callback chạy ──
        # Tổng quan
        State("filter-price",               "data"),
        State("filter-volume",              "data"),
        State("filter-market-cap",          "data"),
        State("filter-eps",                 "data"),
        State("filter-perf-1w",             "data"),
        State("filter-perf-1m",             "data"),
        # Định giá
        State("filter-pe",                  "data"),
        State("filter-pb",                  "data"),
        State("filter-ps",                  "data"),
        State("filter-ev-ebitda",           "data"),
        State("filter-div-yield",           "data"),
        # Sinh lời
        State("filter-roe",                 "data"),
        State("filter-roa",                 "data"),
        State("filter-gross-margin",        "data"),
        State("filter-net-margin",          "data"),
        State("filter-ebit-margin",         "data"),
        # Tăng trưởng
        State("filter-rev-growth-yoy",      "data"),
        State("filter-rev-cagr-5y",         "data"),
        State("filter-eps-growth-yoy",      "data"),
        State("filter-eps-cagr-5y",         "data"),
        # Sức khỏe
        State("filter-de",                  "data"),
        State("filter-current-ratio",       "data"),
        State("filter-net-cash-cap",        "data"),
        State("filter-net-cash-assets",     "data"),
        # Scores
        State("filter-value-score",         "data"),
        State("filter-growth-score",        "data"),
        State("filter-momentum-score",      "data"),
        State("filter-vgm-score",           "data"),
        State("filter-canslim",             "data"),
        # Kỹ thuật – Giá vs SMA
        State("filter-price-vs-sma5",       "data"),
        State("filter-price-vs-sma10",      "data"),
        State("filter-price-vs-sma20",      "data"),
        State("filter-price-vs-sma50",      "data"),
        State("filter-price-vs-sma100",     "data"),
        State("filter-price-vs-sma200",     "data"),
        # Kỹ thuật – Đỉnh/Đáy
        State("filter-pct-from-high-1y",    "data"),
        State("filter-pct-from-low-1y",     "data"),
        State("filter-pct-from-high-all",   "data"),
        State("filter-pct-from-low-all",    "data"),
        State("filter-break-high-52w",      "data"),
        State("filter-break-low-52w",       "data"),
        # Kỹ thuật – Oscillators
        State("filter-rsi14",               "data"),
        State("filter-macd-hist",           "data"),
        State("filter-bb-width",            "data"),
        State("filter-adx14",               "data"),
        State("filter-plus-di14",           "data"),
        State("filter-minus-di14",          "data"),
        State("filter-consec-up",           "data"),
        State("filter-consec-down",         "data"),
        # Kỹ thuật – Momentum/RS
        State("filter-beta",                "data"),
        State("filter-alpha",               "data"),
        State("filter-rs-3d",               "data"),
        State("filter-rs-1m",               "data"),
        State("filter-rs-3m",               "data"),
        State("filter-rs-1y",               "data"),
        State("filter-rs-avg",              "data"),
        # Kỹ thuật – Volume
        State("filter-vol-vs-sma5",         "data"),
        State("filter-vol-vs-sma10",        "data"),
        State("filter-vol-vs-sma20",        "data"),
        State("filter-vol-vs-sma50",        "data"),
        State("filter-avg-vol-5d",          "data"),
        State("filter-avg-vol-10d",         "data"),
        State("filter-avg-vol-50d",         "data"),
        # GTGD
        State("filter-gtgd-1w",             "data"),
        State("filter-gtgd-10d",            "data"),
        State("filter-gtgd-1m",             "data"),
        # >>> THÊM STATE CHO DROPDOWN CHỈ SỐ <<<
        State("filter-fib-position",       "data"),
        State("filter-wave-momentum",      "data"),
        State("filter-elliott-corrective", "data"),
    ],
    prevent_initial_call='initial_duplicate'
)
def update_screener_table(
        btn_reset, search_text, current_strategy, selected_sectors, active_filters, selected_subs,
        selected_exchange, filter_year, include_null, trading_mode, investor_profile,
        filter_index, nav_value,
        auth_data,          # ← DỜI LÊN ĐÂY, khớp vị trí Input("auth-store","data") trong decorator
        # Tổng quan (State)
        price_range, volume_range, market_cap_range, eps_range, perf_1w_range, perf_1m_range,
        # Định giá
        pe_range, pb_range, ps_range, ev_ebitda_range, div_yield_range,
        # Sinh lời
        roe_range, roa_range, gross_margin_range, net_margin_range, ebit_margin_range,
        # Tăng trưởng
        rev_growth_yoy_range, rev_cagr_5y_range, eps_growth_yoy_range, eps_cagr_5y_range,
        # Sức khỏe
        de_range, current_ratio_range, net_cash_cap_range, net_cash_assets_range,
        # Scores
        value_scores, growth_scores, momentum_scores, vgm_scores, canslim_range,
        # Kỹ thuật – Giá vs SMA
        pvsma5, pvsma10, pvsma20, pvsma50, pvsma100, pvsma200,
        # Kỹ thuật – Đỉnh/Đáy
        pct_high_1y, pct_low_1y, pct_high_all, pct_low_all,
        break_high_52w, break_low_52w,
        # Kỹ thuật – Oscillators
        rsi14_range, macd_range, bb_range, adx_range, plus_di_range, minus_di_range, consec_up_range, consec_down_range,
        # Kỹ thuật – Momentum/RS
        beta_range, alpha_range, rs3d, rs1m, rs3m, rs1y, rs_avg,
        # Kỹ thuật – Volume
        vvsma5, vvsma10, vvsma20, vvsma50, avg5d, avg10d, avg50d,
        # GTGD
        gtgd_1w, gtgd_10d, gtgd_1m,
        fib_pos, wave_mom, elliott_corr,
):
    try:
        # ── DEBUG BLOCK — BẮT ĐẦU ──────────────────────────────────────
        ctx = callback_context
        triggered_id = ctx.triggered[0]['prop_id'].split('.')[0] if ctx.triggered else "NONE"
        triggered_val = ctx.triggered[0].get('value', '?') if ctx.triggered else '?'
        import time as _time
        logger.warning(
            f"\n{'='*70}\n"
            f"[SCREENER TRIGGER] @ {_time.strftime('%H:%M:%S')}\n"
            f"  triggered_id  = {triggered_id}\n"
            f"  strategy      = {current_strategy}\n"
            f"  active_filters = {list((active_filters or {}).keys())}\n"
            f"  val_preview   = {str(triggered_val)[:100]}\n"
            f"{'='*70}"
        )
        # ── DEBUG BLOCK — KẾT THÚC ──────────────────────────────────────
        # 1. LƯỚI AN TOÀN: Bắt buộc giá trị mặc định là "investing" nếu có lỗi
        if not trading_mode:
            trading_mode = "investing"
        ctx = callback_context
        triggered_id = ctx.triggered[0]['prop_id'].split('.')[0] if ctx.triggered else None

        # Load Snapshot trực tiếp dưới dạng DataFrame (không qua list[dict] roundtrip)
        df = get_snapshot_df()
        if df is None or df.empty:
            return [], FIXED_COLS, "⚠️ Không có dữ liệu", "", False, "", None

        df = df.copy()  # tránh modify in-place trên cache
        total_stocks = len(df)

        # ── LỌC THEO NĂM (qua BCTC — snapshot chỉ có 1 ngày nên lọc qua df_fin) ──
        if filter_year and filter_year != "all":
            yr = int(filter_year)
            try:
                df_fin_yr = load_financial_data('yearly')
                if df_fin_yr is not None and not df_fin_yr.empty and 'Date' in df_fin_yr.columns:
                    df_fin_yr = df_fin_yr.copy()
                    df_fin_yr['_yr'] = pd.to_datetime(df_fin_yr['Date'], errors='coerce').dt.year
                    tickers_in_year = set(df_fin_yr[df_fin_yr['_yr'] == yr]['Ticker'].dropna().unique())
                    df = df[df['Ticker'].isin(tickers_in_year)]
                    logger.info(f"[YEAR FILTER] Năm {yr}: {len(tickers_in_year)} ticker có BCTC → còn {len(df)} mã")
            except Exception as e:
                logger.warning(f"[YEAR FILTER] Lỗi lọc năm: {e}")
            total_stocks = len(df)
        # ─────────────────────────────────────────────────────────────────────

        if 'FSS_Smart_Rank' in df.columns:
            df = df.sort_values('FSS_Smart_Rank', ascending=False)
        elif 'Star_Rating' in df.columns:
            df = df.sort_values('Star_Rating', ascending=False)
        elif 'VGM Score' in df.columns:
            grade_order = {'A': 1, 'B': 2, 'C': 3, 'D': 4, 'F': 5}
            df['_sort'] = df['VGM Score'].map(grade_order).fillna(6)
            df = df.sort_values('_sort').drop('_sort', axis=1)

        if triggered_id == 'btn-reset' or triggered_id == 'btn-reset.n_clicks':
            df = _add_forward_pe(df)
            col_defs = _build_col_defs(active_filters, current_strategy, trading_mode)
            return df.to_dict('records'), col_defs, f"📊 Hiển thị tất cả: {total_stocks} mã", "", False, "", None

        df_filtered = df.copy()
        df_null_excluded = pd.DataFrame()  # Accumulate mã bị loại vì null

        # ── HARD FILTER theo chế độ đầu tư ──────────────────────────────────────────
        # ── HARD FILTER theo hồ sơ nhà đầu tư ──────────
        _has_user_filters = bool(active_filters)  # True nếu còn thẻ nào đó

        if trading_mode != "all_market" and investor_profile and investor_profile.get("auto_filters") and _has_user_filters:   # ← THÊM điều kiện này
            af = investor_profile["auto_filters"]
            min_vol   = af.get("min_vol",   30_000)
            min_cap   = af.get("min_cap",   200_000_000_000)
            min_price = af.get("min_price", 3_000)

            if "Avg_Vol_20D" in df_filtered.columns:
                df_filtered = df_filtered[
                    pd.to_numeric(df_filtered["Avg_Vol_20D"],
                                errors="coerce").fillna(0) >= min_vol
                ]
            if "Market Cap" in df_filtered.columns:
                df_filtered = df_filtered[
                    pd.to_numeric(df_filtered["Market Cap"],
                                errors="coerce").fillna(0) >= min_cap
                ]
            if "Price Close" in df_filtered.columns:
                df_filtered = df_filtered[
                    pd.to_numeric(df_filtered["Price Close"],
                                errors="coerce").fillna(0) >= min_price
                ]

            logger.info(f"[Profile Filter] Vốn={investor_profile.get('capital')} "
                        f"→ vol≥{min_vol:,}, cap≥{min_cap/1e9:.0f}tỷ, "
                        f"price≥{min_price:,} → còn {len(df_filtered)} mã")
            
        # ── Hard filter: Loại ngành Ngân hàng / BĐS khi tick avoid_bank_re ──
        if trading_mode != "all_market" and investor_profile and investor_profile.get("avoid_bank_re"):
            # Tên ngành trong data sau khi dịch qua GICS_SECTOR_TRANSLATION
            _EXCLUDE_SECTORS = {
                # Tiếng Việt (sau khi translate)
                "Tài chính", "Bất động sản",
                # Tiếng Anh gốc (fallback nếu chưa dịch)
                "Financials", "Real Estate",
            }
            _EXCLUDE_INDUSTRY_KEYWORDS = [
                "Ngân hàng", "Bank", "Bất động sản", "Real Estate",
                "Bảo hiểm", "Insurance", "Chứng khoán",
            ]

            before_excl = len(df_filtered)

            # Bước 1: Thử lọc qua cột Sector (chuẩn nhất)
            sec_col = next(
                (c for c in ["Sector", "GICS Sector Name"] if c in df_filtered.columns),
                None,
            )
            if sec_col:
                df_filtered = df_filtered[
                    ~df_filtered[sec_col]
                    .astype(str)
                    .str.strip()
                    .isin(_EXCLUDE_SECTORS)
                ]

            # Bước 2: Fallback — lọc thêm qua GICS Industry Name nếu còn sót
            ind_col = next(
                (c for c in ["GICS Industry Name", "GICS Sub-Industry Name"]
                 if c in df_filtered.columns),
                None,
            )
            if ind_col:
                _kw_pattern = "|".join(_EXCLUDE_INDUSTRY_KEYWORDS)
                df_filtered = df_filtered[
                    ~df_filtered[ind_col]
                    .astype(str)
                    .str.contains(_kw_pattern, case=False, na=False, regex=True)
                ]

            logger.info(
                f"[avoid_bank_re] Đã loại ngành Tài chính/BĐS: "
                f"{before_excl} → {len(df_filtered)} mã"
            )

            # Đưa thông tin hard filter vào result-count để hiển thị
            _hard_filter_note = (
                f"  ·  🔒 Lọc tự động theo hồ sơ: KL≥{min_vol//1000}K, "
                f"Vốn hóa≥{min_cap//1_000_000_000:.0f}tỷ, Giá≥{min_price:,}đ"
            )

        if trading_mode == "trading":
            # === CHẾ ĐỘ LƯỚT SÓNG T+ ===
            # Mục tiêu: Cổ phiếu có thanh khoản đủ để vào/ra nhanh, có momentum

            # 1. Thanh khoản tối thiểu — cần đủ để mua/bán không bị trượt giá
            if "Avg_Vol_20D" in df_filtered.columns:
                avg_vol = pd.to_numeric(df_filtered["Avg_Vol_20D"], errors="coerce").fillna(0)
                df_filtered = df_filtered[avg_vol >= 200_000]  # 200K CP/ngày

            # 2. Giá tối thiểu — loại penny stock, dễ bị làm giá
            if "Price Close" in df_filtered.columns:
                price = pd.to_numeric(df_filtered["Price Close"], errors="coerce").fillna(0)
                df_filtered = df_filtered[price >= 5_000]  # >= 5,000 VNĐ

            # 3. Vốn hóa tối thiểu — loại cổ siêu nhỏ dễ bị thao túng
            if "Market Cap" in df_filtered.columns:
                mc = pd.to_numeric(df_filtered["Market Cap"], errors="coerce").fillna(0)
                df_filtered = df_filtered[mc >= 500_000_000_000]  # >= 500 tỷ VNĐ

            # 4. Không bị giảm sàn liên tục (RSI không quá thấp — không bắt dao rơi)
            if "RSI_14" in df_filtered.columns:
                rsi = pd.to_numeric(df_filtered["RSI_14"], errors="coerce")
                df_filtered = df_filtered[rsi.isna() | (rsi >= 25)]  # RSI >= 25

            # 5. Momentum dương — giá không thấp hơn 30% so với đỉnh 1 năm
            if "Pct_From_High_1Y" in df_filtered.columns:
                pct_h = pd.to_numeric(df_filtered["Pct_From_High_1Y"], errors="coerce")
                df_filtered = df_filtered[pct_h.isna() | (pct_h >= -40)]

            logger.info(f"[Trading Mode] Sau hard filter: {len(df_filtered)} mã")

        elif trading_mode == "investing":
            # === CHẾ ĐỘ TÍCH SẢN ===
            # Mục tiêu: Doanh nghiệp thật sự hoạt động, có lợi nhuận, không rủi ro mất vốn

            # 1. Thanh khoản tối thiểu — đủ để thoát khi cần
            if "Avg_Vol_20D" in df_filtered.columns:
                avg_vol = pd.to_numeric(df_filtered["Avg_Vol_20D"], errors="coerce").fillna(0)
                df_filtered = df_filtered[avg_vol >= 30_000]  # 30K CP/ngày

            # 2. Giá tối thiểu — loại cổ dưới mệnh giá (nguy cơ hủy niêm yết)
            if "Price Close" in df_filtered.columns:
                price = pd.to_numeric(df_filtered["Price Close"], errors="coerce").fillna(0)
                df_filtered = df_filtered[price >= 3_000]  # >= 3,000 VNĐ (mệnh giá 10,000đ)

            # 3. Doanh nghiệp có lợi nhuận — loại công ty thua lỗ
            if "P/E" in df_filtered.columns:
                pe = pd.to_numeric(df_filtered["P/E"], errors="coerce")
                # P/E dương = có lãi; P/E > 150 = bất thường
                df_filtered = df_filtered[pe.isna() | ((pe > 0) & (pe <= 150))]

            # 4. Vốn hóa tối thiểu — không đầu tư vào cổ siêu nhỏ (< 200 tỷ)
            if "Market Cap" in df_filtered.columns:
                mc = pd.to_numeric(df_filtered["Market Cap"], errors="coerce").fillna(0)
                df_filtered = df_filtered[mc >= 200_000_000_000]  # >= 200 tỷ

            # 5. Không đang trong tình trạng thua lỗ nặng (ROE không quá âm)
            if "ROE (%)" in df_filtered.columns:
                roe = pd.to_numeric(df_filtered["ROE (%)"], errors="coerce")
                df_filtered = df_filtered[roe.isna() | (roe >= -20)]  # ROE >= -20%

            logger.info(f"[Investing Mode] Sau hard filter: {len(df_filtered)} mã")

        elif trading_mode == "all_market":
            # Không lọc gì — hiển thị toàn bộ ~1,500 mã
            pass
        # ================================================================
        # 🟢 TẦNG 0: LỌC THEO TỪ KHÓA TÌM KIẾM (SEARCH BAR)
        # ================================================================
        # ================================================================
        # 🟢 TẦNG 0: LỌC THEO TỪ KHÓA TÌM KIẾM (Bug #4 fix: tìm cả tên công ty)
        # ================================================================
        if search_text:
            # search_text có thể là ticker thuần (value từ Dropdown) hoặc chuỗi tên
            # Dropdown.value luôn là ticker (vd "FPT"), nhưng giữ logic search tên phòng khi
            # người dùng xoá rồi gõ tay
            search_val = str(search_text).strip()
            search_upper = search_val.upper()

            # Exact match theo ticker trước
            exact_ticker = df_filtered['Ticker'].astype(str).str.upper() == search_upper
            if exact_ticker.any():
                df_filtered = df_filtered[exact_ticker]
            else:
                # Tìm theo ticker startswith + tên tiếng Anh + tên tiếng Việt (từ COMP INFO)
                ticker_match = df_filtered['Ticker'].astype(str).str.upper().str.startswith(search_upper, na=False)
                name_en_match = pd.Series([False] * len(df_filtered), index=df_filtered.index)
                name_vi_match = pd.Series([False] * len(df_filtered), index=df_filtered.index)

                if 'Company Common Name' in df_filtered.columns:
                    name_en_match = df_filtered['Company Common Name'].astype(str).str.upper().str.contains(
                        search_upper, na=False, regex=False)

                # Load tên tiếng Việt từ COMP INFO và match
                try:
                    if not df_comp_info.empty:
                        # df_comp_info đã được load global ở đầu screener_callbacks.py
                        vn_col = next((c for c in ['organ_name','company_name_vi','ten_cong_ty','name_vi']
                                    if c in df_comp_info.columns), None)
                        if vn_col:
                            vn_map = df_comp_info.set_index('Ticker')[vn_col].to_dict()
                            df_filtered['_vn_name'] = df_filtered['Ticker'].map(vn_map).fillna('')
                            name_vi_match = df_filtered['_vn_name'].str.upper().str.contains(
                                search_upper, na=False, regex=False)
                            df_filtered = df_filtered.drop(columns=['_vn_name'])
                except Exception:
                    pass

                df_filtered = df_filtered[ticker_match | name_en_match | name_vi_match]
                
        from src.backend.data_loader import fetch_index_constituents
    
        # Khởi tạo trạng thái của Toast
        toast_is_open = False
        toast_msg = ""

        # ── 1. LỌC THEO CHỈ SỐ THỊ TRƯỜNG (VN30, HNX30...) ──
        if filter_index and filter_index != "all":
            tickers_list, api_error = fetch_index_constituents(filter_index)
            
            if api_error:
                # Nếu gọi API thất bại -> Bật Toast cảnh báo
                toast_is_open = True
                toast_msg = api_error
            elif tickers_list is not None:
                # Nếu gọi API thành công -> Lọc data chồng lên df_filtered hiện tại
                df_filtered = df_filtered[df_filtered["Ticker"].isin(tickers_list)]

        # ================================================================
        # TẦNG 1: LỌC THEO TRƯỜNG PHÁI (STRATEGY)
        # ================================================================
        if current_strategy:
            logger.info(f"Áp dụng Tầng 1 (Trường phái): {current_strategy}")
            try:
                df_fin = load_financial_data('yearly')
            except Exception as e:
                logger.warning(f"Lỗi load df_fin: {e}")
                df_fin = None
            df_filtered = run_strategy(df_filtered, current_strategy, df_fin=df_fin)
            df_filtered = df_filtered.replace([float('inf'), float('-inf')], None)

        # ================================================================
        # TẦNG 2: LỌC THEO NGÀNH (Bug #1 fix)
        # ================================================================
        if selected_sectors and isinstance(selected_sectors, list):
            clean = [s for s in selected_sectors if s != "all"]
            if clean:
                # Xác định cột sector tồn tại trong df
                sec_col = next((c for c in ['Sector', 'GICS Sector Name'] if c in df_filtered.columns), None)
                if sec_col:
                    # Lấy tất cả giá trị sector thực tế trong data
                    actual_vals = set(df_filtered[sec_col].dropna().unique())
                    # Lọc chỉ giữ những giá trị clean thực sự tồn tại trong data
                    # (loại bỏ các giá trị cũ từ localStorage như "Chemicals" - là GICS Industry chứ không phải Sector)
                    valid_clean = [s for s in clean if s in actual_vals]
                    if valid_clean:
                        df_filtered = df_filtered[df_filtered[sec_col].isin(valid_clean)]
                    # Nếu valid_clean rỗng (toàn giá trị lạ từ localStorage) → không lọc, tránh mất sạch data

        # ================================================================
        # TẦNG 3: LỌC CHỈ TIÊU — đọc range từ active_filters["value"]
        # (active-filters-store là nguồn sự thật duy nhất, cập nhật bởi
        #  cả manage_filter_ui lẫn activate_readonly_filter_on_drag)
        # Fallback về State params nếu active_filters không có "value"
        # ================================================================
        if not active_filters:
            active_filters = {}

        def apply_range(col_name, rng):
            nonlocal df_filtered, df_null_excluded
            if col_name in df_filtered.columns and rng and isinstance(rng, (list, tuple)) and len(rng) == 2:
                numeric = pd.to_numeric(df_filtered[col_name], errors='coerce')

                mask_in_range = (numeric >= rng[0]) & (numeric <= rng[1])
                mask_is_null  = numeric.isna()

                # Tách riêng các mã null để có thể tái sử dụng nếu toggle bật
                null_rows = df_filtered[mask_is_null & ~mask_in_range]
                df_null_excluded = pd.concat(
                    [df_null_excluded, null_rows]
                ).drop_duplicates(subset=["Ticker"])

                if include_null:
                    # Chế độ rủi ro: giữ lại null, đánh dấu cảnh báo
                    df_filtered = df_filtered[mask_in_range | mask_is_null]
                    # Thêm cột cảnh báo nếu chưa có
                    if "_null_warning" not in df_filtered.columns:
                        df_filtered["_null_warning"] = ""
                    df_filtered.loc[mask_is_null, "_null_warning"] = "⚠️ Thiếu dữ liệu"
                else:
                    # Chế độ mặc định (chuẩn): loại bỏ null
                    df_filtered = df_filtered[mask_in_range]

        def apply_grade(col_name, grades):
            nonlocal df_filtered
            if col_name in df_filtered.columns and grades:
                # BẢO VỆ: Nếu cột này rỗng toàn bộ (NaN) thì bỏ qua để tránh rớt 0 mã
                if not df_filtered[col_name].isna().all():
                    df_filtered = df_filtered[df_filtered[col_name].isin(grades)]

        # Map filter_id → (col_name, fallback_state_value, is_grade)
        FILTER_MAP = [
            # Tổng quan
            ("filter-price",            "Price Close",              price_range,            False),
            ("filter-volume",           "Volume",                   volume_range,           False),
            ("filter-market-cap",       "Market Cap",               market_cap_range,       False),
            ("filter-eps",              "EPS",                      eps_range,              False),
            ("filter-perf-1w",          "Perf_1W",                  perf_1w_range,          False),
            ("filter-perf-1m",          "Perf_1M",                  perf_1m_range,          False),
            # Định giá
            ("filter-pe",               "P/E",                      pe_range,               False),
            ("filter-pb",               "P/B",                      pb_range,               False),
            ("filter-ps",               "P/S",                      ps_range,               False),
            ("filter-ev-ebitda",        "EV/EBITDA",                ev_ebitda_range,        False),
            ("filter-div-yield",        "Dividend Yield (%)",       div_yield_range,        False),
            # Sinh lời
            ("filter-roe",              "ROE (%)",                  roe_range,              False),
            ("filter-roa",              "ROA (%)",                  roa_range,              False),
            ("filter-gross-margin",     "Gross Margin (%)",         gross_margin_range,     False),
            ("filter-net-margin",       "Net Margin (%)",           net_margin_range,       False),
            ("filter-ebit-margin",      "EBIT Margin (%)",          ebit_margin_range,      False),
            # Tăng trưởng
            ("filter-rev-growth-yoy",   "Revenue Growth YoY (%)",   rev_growth_yoy_range,   False),
            ("filter-rev-cagr-5y",      "Revenue CAGR 5Y (%)",      rev_cagr_5y_range,      False),
            ("filter-eps-growth-yoy",   "EPS Growth YoY (%)",       eps_growth_yoy_range,   False),
            ("filter-eps-cagr-5y",      "EPS CAGR 5Y (%)",          eps_cagr_5y_range,      False),
            # Sức khỏe
            ("filter-de",               "D/E",                      de_range,               False),
            ("filter-current-ratio",    "Current Ratio",            current_ratio_range,    False),
            ("filter-net-cash-cap",     "Net Cash / Market Cap (%)",net_cash_cap_range,     False),
            ("filter-net-cash-assets",  "Net Cash / Assets (%)",    net_cash_assets_range,  False),
            # Scores (grade)
            ("filter-value-score",      "Value Score",              value_scores,           True),
            ("filter-growth-score",     "Growth Score",             growth_scores,          True),
            ("filter-momentum-score",   "Momentum Score",           momentum_scores,        True),
            ("filter-vgm-score",        "VGM Score",                vgm_scores,             True),
            ("filter-canslim",          "CANSLIM Score",            canslim_range,          False),
            # Kỹ thuật – Giá vs SMA
            ("filter-price-vs-sma5",    "Price_vs_SMA5",            pvsma5,                 False),
            ("filter-price-vs-sma10",   "Price_vs_SMA10",           pvsma10,                False),
            ("filter-price-vs-sma20",   "Price_vs_SMA20",           pvsma20,                False),
            ("filter-price-vs-sma50",   "Price_vs_SMA50",           pvsma50,                False),
            ("filter-price-vs-sma100",  "Price_vs_SMA100",          pvsma100,               False),
            ("filter-price-vs-sma200",  "Price_vs_SMA200",          pvsma200,               False),
            # Kỹ thuật – Đỉnh/Đáy
            ("filter-pct-from-high-1y", "Pct_From_High_1Y",         pct_high_1y,            False),
            ("filter-pct-from-low-1y",  "Pct_From_Low_1Y",          pct_low_1y,             False),
            ("filter-pct-from-high-all","Pct_From_High_All",         pct_high_all,           False),
            ("filter-pct-from-low-all", "Pct_From_Low_All",         pct_low_all,            False),
            # Kỹ thuật – Oscillators
            ("filter-rsi14",            "RSI_14",                   rsi14_range,            False),
            ("filter-macd-hist",        "MACD_Histogram",           macd_range,             False),
            ("filter-bb-width",         "BB_Width",                 bb_range,               False),
            ("filter-adx14",            "ADX_14",                   adx_range,              False),
            ("filter-plus-di14",        "Plus_DI_14",                plus_di_range,         False),
            ("filter-minus-di14",       "Minus_DI_14",               minus_di_range,        False),
            ("filter-consec-up",        "Consec_Up",                consec_up_range,        False),
            ("filter-consec-down",      "Consec_Down",              consec_down_range,      False),
            # Kỹ thuật – Momentum/RS
            ("filter-beta",             "Beta",                     beta_range,             False),
            ("filter-alpha",            "Alpha",                    alpha_range,            False),
            ("filter-rs-3d",            "RS_3D",                    rs3d,                   False),
            ("filter-rs-1m",            "RS_1M",                    rs1m,                   False),
            ("filter-rs-3m",            "RS_3M",                    rs3m,                   False),
            ("filter-rs-1y",            "RS_1Y",                    rs1y,                   False),
            ("filter-rs-avg",           "RS_Avg",                   rs_avg,                 False),
            # Kỹ thuật – Volume
            ("filter-vol-vs-sma5",      "Vol_vs_SMA5",              vvsma5,                 False),
            ("filter-vol-vs-sma10",     "Vol_vs_SMA10",             vvsma10,                False),
            ("filter-vol-vs-sma20",     "Vol_vs_SMA20",             vvsma20,                False),
            ("filter-vol-vs-sma50",     "Vol_vs_SMA50",             vvsma50,                False),
            ("filter-avg-vol-5d",       "Avg_Vol_5D",               avg5d,                  False),
            ("filter-avg-vol-10d",      "Avg_Vol_10D",              avg10d,                 False),
            ("filter-avg-vol-50d",      "Avg_Vol_50D",              avg50d,                 False),
            # GTGD
            ("filter-gtgd-1w",          "GTGD_1W",                  gtgd_1w,                False),
            ("filter-gtgd-10d",         "GTGD_10D",                 gtgd_10d,               False),
            ("filter-gtgd-1m",          "GTGD_1M",                  gtgd_1m,                False),
            # Elliott Wave Proxy
            ("filter-fib-position",       "Fib_Position_%",        fib_pos,       False),
            ("filter-wave-momentum",      "Wave_Momentum_Score",    wave_mom,      False),
        ]

        for (filter_id, col_name, fallback_val, is_grade) in FILTER_MAP:
            if filter_id not in active_filters:
                continue  # Chỉ áp dụng khi filter đang active
            # Ưu tiên lấy value từ active_filters (được cập nhật khi kéo slider)
            # Fallback về State param nếu active_filters chưa có "value"
            af_entry = active_filters[filter_id]
            rng_or_grades = af_entry.get("value", fallback_val) if isinstance(af_entry, dict) else fallback_val
            if is_grade:
                apply_grade(col_name, rng_or_grades)
            else:
                apply_range(col_name, rng_or_grades)

        # ── Boolean filters: Break_High_52W / Break_Low_52W ──
        _BOOL_MAP = [
            ("filter-break-high-52w",     "Break_High_52W",   break_high_52w),
            ("filter-break-low-52w",      "Break_Low_52W",    break_low_52w),
            # Elliott: cùng 1 cột, khác value
            # filter-elliott-impulse    → Elliott_Corrective == 0  (KHÔNG hồi = đang đẩy)
            # filter-elliott-corrective → Elliott_Corrective == 1  (đang hồi)
            ("filter-elliott-impulse",    "Elliott_Corrective", None),
            ("filter-elliott-corrective", "Elliott_Corrective", None),
        ]

        for (fid, col, bval) in _BOOL_MAP:
            if fid not in active_filters:
                continue
            if col not in df_filtered.columns:
                logger.warning(f"[BOOL_MAP] Cột '{col}' không tồn tại trong snapshot — bỏ qua")
                continue
            af_entry = active_filters[fid]
            bool_val = af_entry.get("value") if isinstance(af_entry, dict) else None
            if bool_val is None:
                bool_val = bval
            if bool_val is None:
                continue

            # ── FIX ELLIOTT: map filter_id → giá trị cột thực tế ──────────────────
            # Cả 2 filter đều dùng cột Elliott_Corrective
            # impulse   → người dùng bấm "Có"(1) → lọc Elliott_Corrective == 0
            # corrective → người dùng bấm "Có"(1) → lọc Elliott_Corrective == 1
            # SAU (đúng — dùng bool_val để đảo chiều):
            if fid == "filter-elliott-corrective":
                # bool_val=1 (bấm "Có") → muốn xem hồi sóng → Elliott_Corrective == 1
                # bool_val=0 (bấm "Không") → muốn xem sóng đẩy → Elliott_Corrective == 0
                target_val = 1 if int(bool_val) == 1 else 0

            elif fid == "filter-elliott-impulse":
                # bool_val=1 (bấm "Có") → muốn xem sóng đẩy → Elliott_Corrective == 0
                # bool_val=0 (bấm "Không") → muốn xem hồi sóng → Elliott_Corrective == 1
                target_val = 0 if int(bool_val) == 1 else 1

            else:
                target_val = int(bool_val)
            # ──────────────────────────────────────────────────────────────────────

            logger.info(f"[BOOL_MAP] Lọc {fid}: {col} == {target_val}")
            df_filtered[col] = pd.to_numeric(df_filtered[col], errors='coerce').fillna(-1).astype(int)
            df_filtered = df_filtered[df_filtered[col] == target_val]
            logger.info(f"[BOOL_MAP] Còn {len(df_filtered)} mã sau khi lọc {col}")

        # ── Categorical filters: ADX_State (multi-select dropdown) ──────────
        _CATEGORICAL_MAP = [
            ("filter-adx-state", "ADX_State"),
        ]
        for (fid, col) in _CATEGORICAL_MAP:
            if fid not in active_filters:
                continue
            if col not in df_filtered.columns:
                logger.warning(f"[CAT_MAP] Cột '{col}' không tồn tại trong snapshot — bỏ qua")
                continue
            af_entry = active_filters[fid]
            selected_vals = af_entry.get("value") if isinstance(af_entry, dict) else None
            if not selected_vals:
                continue  # rỗng = không lọc, hiển thị tất cả trạng thái
            before = len(df_filtered)
            df_filtered = df_filtered[df_filtered[col].isin(selected_vals)]
            logger.info(f"[CAT_MAP] Lọc {fid}: {col} in {selected_vals} → {before} → {len(df_filtered)} mã")

        # [ĐÃ GỠ] Khối lọc theo "adx-strategy-checklist" (3 checkbox cố định:
        # uptrend/super_stock/not_sideway) — UI checklist đã bị gỡ khỏi sidebar.
        # Logic Is_Steady_Uptrend / Is_Super_Stock_ADX giờ được áp dụng qua
        # preset chính thức "STRAT_ADX_MOMENTUM" trong run_strategy() (Tầng 1),
        # xem quant_engine_strategies.apply_adx_strategy_filter.

        # ── Sub-industry filter (Bug #2 fix: xử lý trong callback chính,
        #    không dùng callback riêng nữa để tránh bị overwrite) ──
        if selected_subs and isinstance(selected_subs, list):
            clean_subs = [s for s in selected_subs if s != "all"]
            if clean_subs:
                sub_col = next((c for c in ['GICS Industry Name', 'GICS Sub-Industry Name']
                                if c in df_filtered.columns), None)
                if sub_col:
                    actual_subs = set(df_filtered[sub_col].dropna().unique())
                    valid_subs = [s for s in clean_subs if s in actual_subs]
                    if valid_subs:
                        df_filtered = df_filtered[df_filtered[sub_col].isin(valid_subs)]

        # ── Filter theo sàn giao dịch ──
        # FIX: Xử lý cả string (multi=False) lẫn list (multi=True), normalize giá trị
        if selected_exchange:
            if isinstance(selected_exchange, str):
                clean_ex = [selected_exchange] if selected_exchange not in ("all", "") else []
            else:
                clean_ex = [e for e in selected_exchange if e and e != "all"]

            if clean_ex:
                if 'Exchange' in df_filtered.columns:
                    df_filtered['Exchange'] = df_filtered['Exchange'].astype(str).str.strip()
                    before = len(df_filtered)
                    df_filtered = df_filtered[df_filtered['Exchange'].isin(clean_ex)]
                    logger.info(f"[Exchange Filter] {clean_ex} → {before} → {len(df_filtered)} mã")
                else:
                    logger.warning("[Exchange Filter] Cột 'Exchange' KHÔNG tồn tại trong snapshot! "
                                   "Hãy xóa data/processed/snapshot_cache.parquet và restart.")

        # 🟢 🟢 🟢 THÊM ĐOẠN LOGIC NÀY VÀO TRƯỚC KHI RETURN BẢNG
        if nav_value:
            try:
                # Xóa dấu phẩy để biến chuỗi "50,000,000" thành số nguyên 50000000
                clean_nav = int(str(nav_value).replace(',', ''))
                if clean_nav > 0:
                    # Điều kiện: Tiền của user phải lớn hơn giá trị 1 lô (100 cổ phiếu)
                    # (Giả định 'Price Close' của bạn lưu ở mức VND. Nếu lưu dạng 25.5 (nghìn đồng) thì lấy Price * 1000 * 100)
                    df_filtered = df_filtered[(df_filtered['Price Close'] * 100) <= clean_nav]
            except Exception as e:
                logger.error(f"Lỗi filter NAV: {e}")
                pass
        # 🟢 🟢 🟢 KẾT THÚC ĐOẠN THÊM (NAV filter)

        # ── [NULL ALERT] Tạo thông báo số mã bị loại vì thiếu dữ liệu ──────────
        n_null = len(df_null_excluded["Ticker"].unique()) if not df_null_excluded.empty else 0
        if n_null > 0 and not include_null:
            null_alert = html.Span(
                [
                    html.I(className="fas fa-info-circle",
                           style={"marginRight": "6px", "color": "#f59e0b"}),
                    f"Đã loại trừ các mã không đủ dữ liệu BCTC cho tiêu chí đang lọc "
                    f"(chủ yếu ở sàn UPCoM). Bật toggle 'Incl. N/A' bên trong tab 'Chiến lược' để xem kèm cảnh báo ⚠️",
                ],
                style={
                    "fontSize": "11px", "color": "#9ca3af",
                    "backgroundColor": "rgba(245,158,11,0.08)",
                    "border": "1px solid rgba(245,158,11,0.2)",
                    "borderRadius": "6px", "padding": "6px 6px",
                    "display": "block", "marginTop": "6px",
                },
            )
        else:
            null_alert = None
        # ── [NULL ALERT] KẾT THÚC ───────────────────────────────────────────────

        filtered_count = len(df_filtered)

        # Tính Forward P/E và build columnDefs trong cùng 1 lần → AG Grid nhận 1 batch update
        df_filtered = _add_forward_pe(df_filtered)
        if investor_profile and not df_filtered.empty:
            df_filtered = _add_profile_match_col(df_filtered, investor_profile)
        col_defs = _build_col_defs(active_filters, current_strategy, trading_mode)

        # ── VIP GATE: chỉ hiện 3 mã đầu với user chưa VIP ───────────────────
        # auth_data được truyền vào qua State (thêm ở bước D bên dưới)
        # AUDIT FIX (mục 4 - Major Issue): dùng chung require_entitlement()
        # (src/callbacks/auth_callbacks.py) thay vì tự viết lại logic
        # re-check server-side ở đây — tránh 2 bản sao có thể lệch nhau.
        try:
            from src.callbacks.auth_callbacks import require_entitlement
            is_vip = require_entitlement(auth_data, allowed_tiers=["pro", "b2b"])
        except Exception:
            is_vip = False
        total_rows = len(df_filtered)

        if not is_vip and total_rows > 5:
            visible   = df_filtered.head(5).to_dict("records")
            locked_template = {col: None for col in df_filtered.columns}
            locked_template.update({
                "Ticker": "🔒 VIP",
                "Sector": "Đăng nhập để xem",
                "_locked": True,
            })
            n_locked = min(total_rows - 5, 17)   # hiện tối đa 17 dòng mờ
            locked_rows = [dict(locked_template) for _ in range(n_locked)]
            final_rows = visible + locked_rows
        else:
            final_rows = df_filtered.to_dict("records")

        # [CẬP NHẬT] Trả về thêm 2 tham số của Toast cảnh báo ở cuối
        return (
            final_rows,
            col_defs,
            f"Tìm thấy {filtered_count} / {total_stocks} mã phù hợp",
            f"Lọc: {filtered_count} mã | Tổng: {total_stocks} mã",
            toast_is_open,
            toast_msg,
            null_alert,      # ← Output mới: "filter-null-alert" "children"
        )

    except Exception as e:
        logger.error(f"Error in update_screener_table: {e}")
        import traceback;
        traceback.print_exc()
        # [CẬP NHẬT] Xử lý lỗi cũng phải trả đủ số lượng return (6 Outputs)
        return [], FIXED_COLS, f"❌ Lỗi: {str(e)}", "Vui lòng thử lại", True, "Lỗi hệ thống khi tải dữ liệu.", None



# ============================================================================
# HELPERS: METHODOLOGY MODAL UI COMPONENTS
# ============================================================================

def _meth_section(icon_cls, color, title):
    return html.Div([
        html.I(className=icon_cls, style={"color": color, "marginRight": "8px", "fontSize": "12px"}),
        html.Span(title, style={"fontSize": "12px", "fontWeight": "700", "color": color,
                                "fontFamily": "JetBrains Mono,monospace", "letterSpacing": "0.05em"}),
    ], style={"display": "flex", "alignItems": "center", "marginBottom": "10px"})


def _meth_step(num, color, title, desc):
    return html.Div([
        html.Div(num, style={
            "width": "22px", "height": "22px", "borderRadius": "50%", "flexShrink": "0",
            "backgroundColor": f"rgba{tuple(int(color.lstrip('#')[i:i+2], 16) for i in (0, 2, 4)) + (0.15,)}",
            "border": f"1px solid {color}50",
            "color": color, "fontSize": "11px", "fontWeight": "700",
            "display": "flex", "alignItems": "center", "justifyContent": "center",
            "marginRight": "10px", "marginTop": "1px",
        }),
        html.Div([
            html.Span(title, style={"fontSize": "12px", "fontWeight": "700", "color": "#c9d1d9",
                                    "display": "block", "marginBottom": "2px"}),
            html.Span(desc, style={"fontSize": "11px", "color": "#8b949e", "lineHeight": "1.5"}),
        ]),
    ], style={"display": "flex", "alignItems": "flex-start", "marginBottom": "10px"})


def _meth_metric(title, scoring, rationale):
    return html.Div([
        html.Div([
            html.I(className="fas fa-chart-line",
                   style={"color": "#00e676", "marginRight": "8px", "fontSize": "11px", "marginTop": "2px", "flexShrink": "0"}),
            html.Span(title, style={"fontSize": "12px", "fontWeight": "700", "color": "#c9d1d9"}),
        ], style={"display": "flex", "alignItems": "flex-start", "marginBottom": "5px"}),
        html.P(scoring, style={"fontSize": "11px", "color": "#8b949e", "lineHeight": "1.5",
                                "marginBottom": "4px", "marginLeft": "19px"}),
        html.P([html.I(className="fas fa-lightbulb", style={"color": "#f59e0b", "marginRight": "5px", "fontSize": "10px"}),
                rationale],
               style={"fontSize": "10px", "color": "#6e7681", "lineHeight": "1.5",
                      "marginBottom": "0", "marginLeft": "19px", "fontStyle": "italic"}),
    ], style={
        "marginBottom": "14px", "paddingBottom": "14px",
        "borderBottom": "1px solid rgba(33,38,45,0.8)",
    })

from dash import State # Đảm bảo bạn đã import State ở đầu file
from dash import State

# ============================================================================
# CALLBACK 2A: MỞ MODAL NGAY (< 50ms) — chỉ set title + lưu stock vào store
# ============================================================================
@app.callback(
    Output("detail-modal",          "is_open"),
    Output("modal-title",           "children"),
    Output("selected-stock-store",  "data"),
    # 🟢 THÊM OUTPUT NÀY ĐỂ BƠM MÃ CỔ PHIẾU CHO 2 TAB CÒN LẠI:
    Output("selected-ticker-store", "data"), 
    
    Input("screener-table", "cellDoubleClicked"), 
    State("screener-table", "rowData"), 
    prevent_initial_call=True,
)
def open_detail_modal_fast(double_clicked_cell, grid_data):
    if not double_clicked_cell or not grid_data:
        return dash.no_update, dash.no_update, dash.no_update, dash.no_update

    row_id_str = double_clicked_cell.get("rowId")
    
    if row_id_str is not None and str(row_id_str).isdigit():
        real_index = int(row_id_str)
        stock = grid_data[real_index]
    else:
        return dash.no_update, dash.no_update, dash.no_update, dash.no_update

    ticker       = stock.get('Ticker', 'N/A')
    company_name = stock.get('Company Common Name', '')

    company_name_vn = company_name   # default = tên tiếng Anh từ snapshot

    # Ưu tiên 1: đọc từ COMP INFO.csv (có tên tiếng Việt)
    if not df_comp_info.empty:
        match = df_comp_info[df_comp_info['Ticker'] == ticker]
        if not match.empty:
            # Thử nhiều tên cột có thể có trong file
            for col in ['organ_name', 'company_name_vi', 'ten_cong_ty', 'name_vi']:
                if col in match.columns:
                    val = str(match[col].values[0]).strip()
                    if val and val.lower() not in ('nan', 'none', ''):
                        company_name_vn = val
                        break

    # Ưu tiên 2 (fallback): snapshot đã có Company Common Name
    if company_name_vn == company_name or not company_name_vn:
        company_name_vn = company_name or ticker

    title_text = f"Cổ phiếu {ticker} – {company_name_vn}"
    
    # 🟢 TRẢ VỀ THÊM CÁI `ticker` Ở CUỐI CÙNG CHO VỪA VỚI 4 OUTPUT
    return True, title_text, stock, ticker

# ============================================================================
# CALLBACK 2B: LOAD NỘI DUNG SAU KHI MODAL ĐÃ HIỆN — trigger từ store
# ============================================================================
@app.callback(
    Output("tab-overview-content", "children"),
    Input("selected-stock-store", "data"),
    State("theme-store", "data"),
    prevent_initial_call=True,
)
def load_detail_content(stock, theme="dark"):
    """Load data nặng SAU khi modal đã hiện — user thấy UI ngay."""
    theme = theme or "dark"
    from src.utils.kpi_theme import get_kpi_theme, kpi_card as kpi_card_pastel, \
        plotly_base_layout, plotly_axis_style
    T = get_kpi_theme(theme)
    # ── DEBUG: log khi callback trigger ──────────────────────────────────
    import logging
    _log = logging.getLogger(__name__)
    _log.warning(f"[DEBUG 2B] triggered | stock={bool(stock)} | ticker={stock.get('Ticker') if stock else None}")
    # ─────────────────────────────────────────────────────────────────────
    if not stock:
        return "", "", [], [], [], [], [], []

    ticker       = stock.get('Ticker', 'N/A')
    company_name = stock.get('Company Common Name', '')
    price_close  = stock.get('Price Close', 0)

    company_name_vn = company_name
    if not df_comp_info.empty:
        match = df_comp_info[df_comp_info['Ticker'] == ticker]
        if not match.empty:
            for col in ['organ_name', 'company_name_vi', 'ten_cong_ty', 'name_vi']:
                if col in match.columns:
                    val = str(match[col].values[0]).strip()
                    if val and val.lower() not in ('nan', 'none', ''):
                        company_name_vn = val
                        break
    if not company_name_vn or company_name_vn == company_name:
        company_name_vn = company_name or ticker

    # =========================================================================
    # === TAB 1: OVERVIEW (HỒ SƠ, KPI VÀ SỨC KHỎE TÀI CHÍNH) ===
    # =========================================================================

    # --- 1. TRÍCH XUẤT THÔNG TIN HỒ SƠ DOANH NGHIỆP ---
    # GIỮ NGUYÊN DATA GỐC (không dịch sang tiếng Việt)
    # --- 1. TRÍCH XUẤT THÔNG TIN HỒ SƠ DOANH NGHIỆP ---
    # Đã bọc thêm hàm dịch tiếng Việt
    sector = translate_gics_sector(stock.get('GICS Sector Name', stock.get('Sector', 'Đang cập nhật')))
    industry = translate_gics_industry(stock.get('GICS Industry Name', 'Đang cập nhật'))
    sub_industry = translate_gics_industry(
        stock.get('GICS Sub-Industry Name', stock.get('GICS Industry Name', 'Đang cập nhật'))
    )

    founded_year = stock.get('Organization Founded Year', '---')
    ipo_date = stock.get('Date Became Public', '---')
    auditor = stock.get('Auditor Details', 'Đang cập nhật')
    exchange = stock.get('Exchange', '---')
    exchange_label_map = {'HOSE': 'HOSE (HoSE)', 'HNX': 'HNX', 'UPCOM': 'UPCoM'}
    exchange_display = exchange_label_map.get(str(exchange).strip(), exchange if exchange != '---' else '---')
    exchange_color = {'HOSE': '#3fb950', 'HNX': '#58a6ff', 'UPCOM': '#f59e0b'}.get(str(exchange).strip(), '#8b949e')

    try:
        if ipo_date != '---' and pd.notna(ipo_date):
            ipo_date_str = ipo_date.strftime("%d/%m/%Y") if isinstance(ipo_date, pd.Timestamp) else str(ipo_date)[:10]
        else:
            ipo_date_str = '---'
    except Exception as _e:  # noqa: audit-fix bare-except
        logger.debug(f"Suppressed non-critical error in {__name__} near line 1323: {_e}")
        ipo_date_str = '---'

    try:
        founded_year_str = str(int(float(founded_year))) if founded_year != '---' and pd.notna(founded_year) else '---'
    except Exception as _e:  # noqa: audit-fix bare-except
        logger.debug(f"Suppressed non-critical error in {__name__} near line 1329: {_e}")
        founded_year_str = '---'

    # --- 2. XỬ LÝ DỮ LIỆU ĐỂ TÍNH TOÁN KPI & BIỂU ĐỒ SỨC KHỎE ---
    # Tải dữ liệu BCTC Quý để vẽ biểu đồ lịch sử
    df_history = pd.DataFrame()
    try:
        df_fin_q = load_financial_data('quarterly')  # ✅ có cache in-memory
        df_history = df_fin_q[df_fin_q['Ticker'] == ticker].sort_values("Date", ascending=False).head(
            8)  # Lấy 8 quý gần nhất
        df_history = df_history.sort_values("Date", ascending=True)  # Đảo lại để vẽ từ cũ tới mới
    except Exception as e:
        logger.warning(f"Không thể tải BCTC quý để vẽ biểu đồ sức khỏe: {e}")

    # Lấy các giá trị tính toán — ưu tiên lấy từ snapshot (đã tính sẵn)
    market_cap_raw = stock.get('Market Cap', None)
    shares_out_raw = stock.get('Shares Outstanding', stock.get('Common Shares Outstanding', None))

    # Fallback: tính từ df_history nếu snapshot không có
    if shares_out_raw is None and not df_history.empty and 'Common Shares - Outstanding - Total_x' in df_history.columns:
        shares_out_raw = df_history['Common Shares - Outstanding - Total_x'].iloc[-1]

    shares_out = float(shares_out_raw) if shares_out_raw is not None and pd.notna(shares_out_raw) else np.nan

    if market_cap_raw is not None and pd.notna(market_cap_raw) and float(market_cap_raw) > 0:
        market_cap = float(market_cap_raw) / 1_000_000  # Chuyển sang Triệu VND
    elif not pd.isna(shares_out):
        market_cap = price_close * shares_out / 1_000_000
    else:
        market_cap = 0

    eps = stock.get('EPS', 0)
    pe = stock.get('P/E', 0)
    pb = stock.get('P/B', 0)
    roe = stock.get('ROE (%)', 0)
    # Beta THẬT — đã tính sẵn trong technical_indicators.py bằng
    # cov(return cổ phiếu, return VN-Index) / var(return VN-Index), có lọc
    # mã giao dịch thưa. Trước đây bị hardcode "1.15" cho MỌI mã dù giá trị
    # thật đã có sẵn trong pipeline, chỉ là chưa được đọc ra ở đây.
    beta = stock.get('Beta', None)

    # Tính Cổ tức & Tỷ suất (Dividend Yield)
    dps = df_history['DPS - Common - Net - Issue - By Announcement Date'].iloc[
        -1] if not df_history.empty and 'DPS - Common - Net - Issue - By Announcement Date' in df_history.columns else 0
    div_yield = (dps / price_close * 100) if price_close > 0 and not pd.isna(dps) else 0

    # Tính các chỉ số cho thẻ Phân tích chi tiết (từ dòng dữ liệu mới nhất)
    gross_margin = 0;
    debt_equity = 0;
    ocf_net = 0;
    inv_days = 0;
    ev_ebitda = 0
    if not df_history.empty:
        latest = df_history.iloc[-1]
        gross_margin = (latest.get('Gross Profit - Industrials/Property - Total', 0) / latest.get(
            'Revenue from Business Activities - Total_x', 1)) * 100
        debt_equity = (latest.get('Short-Term Debt & Current Portion of Long-Term Debt', 0) + latest.get(
            'Debt - Long-Term - Total', 0)) / latest.get('Common Equity - Total', 1)
        ocf_net = latest.get('Net Cash Flow from Operating Activities', 0) / latest.get(
            'Net Income after Minority Interest', 1)

        cogs = abs(latest.get('Cost of Revenues - Total', 1))
        inv_turnover = cogs / latest.get('Inventories - Total', 1) if cogs != 0 else 0
        inv_days = 365 / inv_turnover if inv_turnover > 0 else 0

        ebitda = latest.get('Earnings before Interest Taxes Depreciation & Amortization', 1)
        ev = (market_cap * 1_000_000) + (
                    latest.get('Short-Term Debt & Current Portion of Long-Term Debt', 0) + latest.get(
                'Debt - Long-Term - Total', 0)) - latest.get('Cash & Cash Equivalents - Total_x', 0)
        ev_ebitda = ev / ebitda if ebitda > 0 else 0

    # Hàm đánh giá điểm (giả lập logic 0-100 dựa trên giá trị)
    def calc_score(val, thresholds, inverse=False):
        # thresholds: [bad, ok, good]
        if pd.isna(val) or val == np.inf or val == -np.inf: return 50, "Trung Bình", "warning"
        if not inverse:
            if val >= thresholds[2]:
                return 90, "Rất Tốt", "success"
            elif val >= thresholds[1]:
                return 70, "Tốt", "success"
            elif val >= thresholds[0]:
                return 50, "Trung Bình", "warning"
            else:
                return 30, "Yếu", "danger"
        else:  # Các chỉ số như nợ, tỷ số càng nhỏ càng tốt
            if val <= thresholds[0]:
                return 90, "Rất Tốt", "success"
            elif val <= thresholds[1]:
                return 70, "Tốt", "success"
            elif val <= thresholds[2]:
                return 50, "Trung Bình", "warning"
            else:
                return 30, "Yếu", "danger"

    score_gm, label_gm, color_gm = calc_score(gross_margin, [10, 20, 30])
    score_de, label_de, color_de = calc_score(debt_equity, [0.5, 1.0, 1.5], inverse=True)
    score_ocf, label_ocf, color_ocf = calc_score(ocf_net, [0.5, 1.0, 1.5])
    score_inv, label_inv, color_inv = calc_score(inv_days, [30, 60, 90], inverse=True)
    score_ev, label_ev, color_ev = calc_score(ev_ebitda, [5, 10, 15], inverse=True)

    total_health_score = int((score_gm + score_de + score_ocf + score_inv) / 4)

    # --- 3. VẼ BIỂU ĐỒ SỨC KHỎE LỊCH SỬ --- (Premium Redesign)
    fig_health = go.Figure()
    if not df_history.empty:
        periods = df_history['Date'].dt.year.astype(str) + "-Q" + df_history['Date'].dt.quarter.astype(str)

        # TÍNH ĐIỂM SỨC KHỎE THẬT CHO TỪNG QUÝ (không dùng số ngẫu nhiên).
        # Trước đây: np.random.normal(total_health_score, 10, ...) → biểu đồ
        # "lịch sử" thực chất là nhiễu giả lập quanh điểm hiện tại, không phải
        # dữ liệu lịch sử thật. Giờ tính lại calc_score() cho từng dòng quý
        # trong df_history bằng đúng công thức dùng cho quý mới nhất.
        def _row_health_score(row):
            gm = (row.get('Gross Profit - Industrials/Property - Total', 0) /
                  row.get('Revenue from Business Activities - Total_x', 1) or 0) * 100
            de = ((row.get('Short-Term Debt & Current Portion of Long-Term Debt', 0) +
                   row.get('Debt - Long-Term - Total', 0)) /
                  (row.get('Common Equity - Total', 1) or 1))
            ocf = (row.get('Net Cash Flow from Operating Activities', 0) /
                   (row.get('Net Income after Minority Interest', 1) or 1))
            cogs_r = abs(row.get('Cost of Revenues - Total', 1) or 1)
            inv_turn = cogs_r / (row.get('Inventories - Total', 1) or 1) if cogs_r != 0 else 0
            inv_d = 365 / inv_turn if inv_turn > 0 else 0

            s_gm, _, _ = calc_score(gm, [10, 20, 30])
            s_de, _, _ = calc_score(de, [0.5, 1.0, 1.5], inverse=True)
            s_ocf, _, _ = calc_score(ocf, [0.5, 1.0, 1.5])
            s_inv, _, _ = calc_score(inv_d, [30, 60, 90], inverse=True)
            return int((s_gm + s_de + s_ocf + s_inv) / 4)

        historical_scores = np.array([_row_health_score(r) for _, r in df_history.iterrows()])
        # Đảm bảo điểm quý mới nhất khớp chính xác với total_health_score hiển thị ở trên
        historical_scores[-1] = total_health_score
        y_min_dynamic = max(0, int(np.min(historical_scores)) - 15)

        # Màu gradient theo điểm — theo theme (xanh tốt / vàng trung bình / đỏ yếu)
        good_rgba = "rgba(63,185,80,0.85)" if theme == "light" else "rgba(0,230,118,0.85)"
        ok_rgba   = "rgba(217,119,6,0.85)" if theme == "light" else "rgba(255,183,3,0.85)"
        bad_rgba  = "rgba(220,38,38,0.75)" if theme == "light" else "rgba(255,61,87,0.75)"
        good_line = "#15803d" if theme == "light" else "#00e676"
        ok_line   = "#b45309" if theme == "light" else "#ffb703"
        bad_line  = "#dc2626" if theme == "light" else "#ff3d57"
        trend_color = T["line_accent"]

        bar_colors = [
            good_rgba if s >= 70 else
            ok_rgba if s >= 50 else
            bad_rgba
            for s in historical_scores
        ]
        border_colors = [
            good_line if s >= 70 else
            ok_line if s >= 50 else
            bad_line
            for s in historical_scores
        ]

        # Bars với border neon
        fig_health.add_trace(go.Bar(
            x=list(periods), y=list(historical_scores),
            name="Điểm Sức Khỏe",
            marker=dict(
                color=bar_colors,
                line=dict(color=border_colors, width=1.5),
            ),
            hovertemplate='<b>%{x}</b><br>Điểm: <b>%{y}</b>/100<extra></extra>',
            showlegend=False,
        ))

        # Đường trend mượt
        fig_health.add_trace(go.Scatter(
            x=list(periods), y=list(historical_scores),
            mode='lines+markers', name="Xu hướng",
            line=dict(color=trend_color, width=2.5, shape='spline', smoothing=0.8),
            marker=dict(
                size=9, color=trend_color,
                line=dict(color=T["chart_paper"] if theme == "light" else '#020810', width=2),
                symbol='circle'
            ),
            hovertemplate='<b>%{x}</b><br>%{y}/100<extra></extra>',
            showlegend=False,
        ))

        # Vùng tô dưới đường trend
        fig_health.add_trace(go.Scatter(
            x=list(periods), y=list(historical_scores),
            fill='tozeroy',
            fillcolor="rgba(15,118,110,0.06)" if theme == "light" else 'rgba(30, 136, 229,0.06)',
            line=dict(color='rgba(0,0,0,0)', width=0),
            showlegend=False, hoverinfo='skip',
        ))

        # Đường tham chiếu 70 (tốt) và 50 (trung bình)
        fig_health.add_hline(y=70, line=dict(color=good_rgba.replace('0.85', '0.3').replace('0.75','0.3'), width=1, dash='dot'))
        fig_health.add_hline(y=50, line=dict(color=ok_rgba.replace('0.85', '0.3'), width=1, dash='dot'))

        fig_health.update_layout(
            **plotly_base_layout(theme, height=260),
            yaxis=dict(
                **plotly_axis_style(theme),
                range=[y_min_dynamic, 105],
                ticksuffix=' ',
            ),
            xaxis={
                **plotly_axis_style(theme),
                "tickangle": -30,
                "gridcolor": 'rgba(0,0,0,0)',
            },
            bargap=0.3,
            showlegend=False,
        )

    # --- 4. RENDER GIAO DIỆN --- (Premium Redesign)
    # Bảng tone/icon xoay vòng theo 5 màu pastel (Global Data 365 style),
    # khớp với thứ tự 8 ô KPI được gọi ở dưới.
    _KPI_TONE_CYCLE = ["sky", "green", "purple", "amber", "rose", "sky", "green", "purple"]
    _KPI_ICON_CYCLE = ["fas fa-building-columns", "fas fa-layer-group", "fas fa-percent",
                       "fas fa-chart-line", "fas fa-coins", "fas fa-scale-balanced",
                       "fas fa-dollar", "fas fa-arrow-trend-up"]
    _kpi_counter = {"i": 0}

    def kpi_card(title, value):
        idx = _kpi_counter["i"] % len(_KPI_TONE_CYCLE)
        _kpi_counter["i"] += 1
        return kpi_card_pastel(
            theme, title, value,
            tone=_KPI_TONE_CYCLE[idx],
            icon_class=_KPI_ICON_CYCLE[idx],
        )

    def make_progress_bar(label, value_str, score, label_text, color, desc):
        # Map color to premium palette
        accent = {"success": T["positive"], "warning": ("#b45309" if theme == "light" else "#ffb703"),
                  "danger": T["negative"]}.get(color, T["line_accent"])
        bg_glow = {"success": "rgba(21,128,61,0.08)" if theme == "light" else "rgba(0,230,118,0.08)",
                   "warning": "rgba(180,83,9,0.08)" if theme == "light" else "rgba(255,183,3,0.08)",
                   "danger": "rgba(220,38,38,0.08)" if theme == "light" else "rgba(255,61,87,0.08)"}.get(
                       color, "rgba(15,118,110,0.06)" if theme == "light" else "rgba(30, 136, 229,0.06)")
        badge_bg = {"success": "rgba(21,128,61,0.15)" if theme == "light" else "rgba(0,230,118,0.15)",
                    "warning": "rgba(180,83,9,0.15)" if theme == "light" else "rgba(255,183,3,0.15)",
                    "danger": "rgba(220,38,38,0.15)" if theme == "light" else "rgba(255,61,87,0.15)"}.get(
                        color, "rgba(15,118,110,0.12)" if theme == "light" else "rgba(30, 136, 229,0.12)")

        return html.Div([
            html.Div([
                html.Span(label, style={
                    "color": T["page_text"], "fontSize": "0.88rem", "fontWeight": "600",
                }),
                html.Span([
                    html.Span(value_str, style={"fontWeight": "700", "marginRight": "6px", "color": accent}),
                    html.Span(label_text, style={
                        "fontSize": "0.72rem", "padding": "2px 7px", "borderRadius": "4px",
                        "backgroundColor": badge_bg, "color": accent, "fontWeight": "600",
                        "border": f"1px solid {accent}22"
                    })
                ], style={"float": "right", "fontSize": "0.83rem", "fontFamily": "JetBrains Mono, monospace"})
            ], style={"marginBottom": "10px", "overflow": "hidden"}),
            # Progress bar custom
            html.Div([
                html.Div(style={
                    "width": f"{score}%",
                    "height": "100%",
                    "background": f"linear-gradient(90deg, {accent}88, {accent})",
                    "borderRadius": "4px",
                    "boxShadow": f"0 0 8px {accent}55",
                    "transition": "width 0.6s ease",
                })
            ], style={
                "height": "8px", "backgroundColor": "rgba(15,118,110,0.10)" if theme == "light" else "rgba(255,255,255,0.05)",
                "borderRadius": "4px", "marginBottom": "10px",
                "overflow": "hidden", "border": f"1px solid {T['card_border']}"
            }),
            html.Div(desc, style={
                "fontSize": "0.78rem", "color": T["page_text_dim"],
                "lineHeight": "1.5", "fontStyle": "italic"
            })
        ], style={
            "padding": "14px 16px", "marginBottom": "10px",
            "background": bg_glow,
            "borderRadius": "8px",
            "border": f"1px solid {T['card_border']}",
            "borderLeft": f"2px solid {accent}44",
        })

    # Helper to recalculate component scores (same logic as quant_engine.py)
    def _calc_fss_components(stock_row, all_df):
        """
        Percentile 4 thành phần Size/Liquidity/Valuation/Quality.

        ĐÃ SỬA (đối chiếu trực tiếp với quant_engine.calculate_fss_smart_rank):
        bản cũ dùng "(giá trị <= x).count()/len(...)" và loại hẳn cổ phiếu P/E<=0
        ra khỏi mẫu số — khác cách backend tính thật (dùng pandas .rank(pct=True),
        KHÔNG loại P/E<=0 mà xếp chúng ở đáy bằng na_option='bottom'). Đây chính là
        nguyên nhân Level 2 (95.6) lệch với Level 1/FSS_Smart_Rank thật (87). Giờ
        tính lại Y HỆT công thức backend nên 2 số sẽ khớp nhau (chỉ lệch do làm tròn).
        """
        try:
            if all_df is None or len(all_df) == 0 or 'Ticker' not in all_df.columns:
                return None, None, None, None

            ticker = str(stock_row.get('Ticker', '')).strip().upper()
            mask = all_df['Ticker'].astype(str).str.strip().str.upper() == ticker
            if not mask.any():
                return None, None, None, None
            row_idx = all_df.index[mask][0]

            # Size — Market Cap, fillna(0), rank(pct=True) — giống hệt quant_engine
            size_rank = pd.to_numeric(all_df.get('Market Cap', 0), errors='coerce').fillna(0).rank(pct=True)

            # Liquidity — ưu tiên GTGD_20D, fallback Avg_Vol_20D, fillna(0), rank(pct=True)
            liq_col = 'GTGD_20D' if 'GTGD_20D' in all_df.columns else 'Avg_Vol_20D'
            liq_rank = pd.to_numeric(all_df.get(liq_col, 0), errors='coerce').fillna(0).rank(pct=True)

            # Valuation — 1/P/E (P/E>0), na_option='bottom': P/E âm/lỗ KHÔNG bị loại
            # khỏi mẫu số như bản cũ, mà bị xếp hạng thấp nhất — đúng cách backend làm
            pe_series = pd.to_numeric(all_df.get('P/E', np.nan), errors='coerce')
            val_rank = (1 / pe_series.where(pe_series > 0)).rank(pct=True, na_option='bottom')

            # Quality — Star_Rating/5, fillna(1) — giống hệt quant_engine
            qual_rank = pd.to_numeric(all_df.get('Star_Rating', 1), errors='coerce').fillna(1) / 5

            size_pct      = round(size_rank.loc[row_idx] * 100, 1)
            liquidity_pct = round(liq_rank.loc[row_idx] * 100, 1)
            valuation_pct = round(val_rank.loc[row_idx] * 100, 1)
            quality_pct   = round(qual_rank.loc[row_idx] * 100, 1)

            return size_pct, liquidity_pct, valuation_pct, quality_pct
        except Exception as e:
            logger.warning(f"Failed to calc FSS components: {e}")
            return None, None, None, None

    # Add these helper functions before the overview_content assignment:

    def _fss_grade(score_pct):
        """Grade mapping for FSS Smart Rank (0-100 scale)."""
        if score_pct >= 90: return "A+"
        if score_pct >= 80: return "A"
        if score_pct >= 65: return "B"
        if score_pct >= 50: return "C"
        if score_pct >= 35: return "D"
        return "F"

    def _fss_component_bar(label, pct, weight_pct, desc):
        """Render a progress bar + label for one FSS component."""
        if pct is None or (isinstance(pct, float) and pd.isna(pct)):
            # Fallback for missing data
            return make_progress_bar(label, "N/A", 0, f"{weight_pct}%", "warning",
                                    desc + " (chưa có dữ liệu)")
        color = "success" if pct >= 65 else ("warning" if pct >= 40 else "danger")
        return make_progress_bar(label, f"{int(pct)}/100", int(pct),
                                f"{weight_pct}% · {_fss_grade(pct)}", color, desc)

    def _fss_signal_chip(label, value):
        """Chip cho Level 3 (VGM/Value/Growth/Momentum) — tăng cỡ đáng kể + thêm khung
        nhẹ để có trọng lượng thị giác tương xứng với Level 1/2 (trước quá nhỏ)."""
        return html.Div([
            html.Div(label, style={"fontSize": "0.8rem", "color": T["page_text_dim"],
                                    "letterSpacing": "0.08em", "fontWeight": "700"}),
            html.Div(str(value), style={"fontSize": "1.5rem", "color": T["page_text"],
                                        "fontWeight": "800", "marginTop": "4px",
                                        "fontFamily": "JetBrains Mono, monospace"}),
        ], style={"textAlign": "center", "padding": "12px 8px", "borderRadius": "8px",
                  "background": "rgba(255,255,255,0.025)" if theme != "light" else "rgba(0,0,0,0.02)",
                  "border": f"1px solid {T['card_border']}"})

    # Style dùng chung cho toàn bộ tooltip trong section FSS — trước dùng mặc định
    # của dbc.Tooltip (bong bóng đen nhỏ, không khớp theme) nên nhìn "cụt" và xấu.
    _fss_tooltip_style = {
        "backgroundColor": T["pastel"]["sky"]["bg"],
        "color": T["page_text"],
        "border": f"1px solid {T['card_border']}",
        "borderRadius": "8px",
        "padding": "10px 12px",
        "fontSize": "0.78rem",
        "lineHeight": "1.5",
        "maxWidth": "260px",
        "textAlign": "left",
        "boxShadow": T["card_shadow"],
        "opacity": "1",
    }

    def _fss_factor_compact(label, pct, weight_pct, tooltip_desc, comp_id):
        """Compact factor card cho lưới 2x2 (theo góp ý UI/UX):
        - chỉ giữ name + percentile + weight + bar trên mặt card
        - "97/100" đổi thành "Phân vị 97 · Top X%" — đây LÀ percentile rank trong
          universe (backend/on-the-fly đều tính theo percentile), không phải "điểm
          chất lượng tuyệt đối", tránh người xem hiểu nhầm 97 = 97 điểm chất lượng.
        - card nền trung tính, màu chỉ dùng cho badge/bar theo status (không xanh toàn bộ)
        - mô tả dài chuyển vào tooltip (ⓘ hover), không hiện thường trực trên card
        """
        if pct is None or (isinstance(pct, float) and pd.isna(pct)):
            pct_display, top_display, bar_pct, color = "N/A", "Chưa có dữ liệu", 0, "warning"
        else:
            pct_int = int(pct)
            pct_display, bar_pct = f"Phân vị {pct_int}", pct_int
            top_display = f"Top {max(1, 100 - pct_int)}% toàn universe"
            color = "success" if pct >= 65 else ("warning" if pct >= 40 else "danger")
        accent = {"success": T["positive"], "warning": ("#b45309" if theme == "light" else "#ffb703"),
                  "danger": T["negative"]}.get(color, T["line_accent"])
        neutral_bg = "rgba(255,255,255,0.025)" if theme != "light" else "rgba(0,0,0,0.02)"

        return html.Div([
            html.Div([
                html.Span([
                    html.Span(label, style={"color": T["page_text"], "fontSize": "0.85rem",
                                             "fontWeight": "700", "letterSpacing": "0.03em"}),
                    html.I(className="fas fa-circle-info", id=comp_id,
                           style={"marginLeft": "6px", "fontSize": "0.72rem",
                                  "color": T["page_text_dim"], "cursor": "help"}),
                ]),
                html.Span([
                    html.Span(pct_display, style={"fontWeight": "700", "marginRight": "6px",
                              "color": accent, "fontFamily": "JetBrains Mono, monospace", "fontSize": "0.95rem"}),
                    html.Span(f"{weight_pct}%", style={
                        "fontSize": "0.72rem", "padding": "1px 6px", "borderRadius": "4px",
                        "backgroundColor": T["card_border"], "color": T["page_text_dim"], "fontWeight": "600",
                    }),
                ]),
            ], style={"display": "flex", "alignItems": "center", "justifyContent": "space-between",
                       "marginBottom": "3px"}),
            html.Div(top_display, style={
                "fontSize": "0.7rem", "color": T["page_text_dim"], "marginBottom": "7px",
            }),
            html.Div([
                html.Div(style={
                    "width": f"{bar_pct}%", "height": "100%",
                    "background": f"linear-gradient(90deg, {accent}88, {accent})",
                    "borderRadius": "3px",
                })
            ], style={
                "height": "7px",
                "backgroundColor": "rgba(255,255,255,0.06)" if theme != "light" else "rgba(15,118,110,0.10)",
                "borderRadius": "3px", "overflow": "hidden", "border": f"1px solid {T['card_border']}",
            }),
            dbc.Tooltip(tooltip_desc, target=comp_id, placement="top", style=_fss_tooltip_style),
        ], style={
            "padding": "12px 14px", "borderRadius": "8px", "background": neutral_bg,
            "border": f"1px solid {T['card_border']}", "borderLeft": f"2px solid {accent}55",
            "height": "100%",
        })



    # --- Dữ liệu & helper cho HỆ THỐNG CHẤM ĐIỂM FSS SMART RANK (8 KPI CARDS) ---
    # Lưu ý: các def/gán biến phải nằm TRƯỚC statement "overview_content = html.Div([...])"
    # vì bên trong đó là MỘT list literal duy nhất (một biểu thức) — không thể chèn
    # statement (gán biến, def hàm) vào giữa list literal, chỉ có thể chèn expression.
    fss_smart_rank = stock.get('FSS_Smart_Rank', 0)
    fss_smart_rank_pct = int(fss_smart_rank * 100) if pd.notna(fss_smart_rank) else 0
    vgm_score_pct = stock.get('VGM_Score_Pct', 0)
    vgm_score = stock.get('VGM Score', 'N/A')
    value_score = stock.get('Value Score', 'N/A')
    growth_score = stock.get('Growth Score', 'N/A')
    momentum_score = stock.get('Momentum Score', 'N/A')
    star_rating = stock.get('Star_Rating', 0)

    _FSS_TONE_CYCLE = ["sky", "green", "amber", "rose"]
    _FSS_ICON_CYCLE = ["fas fa-chart-pie", "fas fa-arrow-trend-up", "fas fa-coins", "fas fa-star"]
    _fss_counter = {"i": 0}

    def kpi_fss(title, value):
        idx = _fss_counter["i"] % len(_FSS_TONE_CYCLE)
        _fss_counter["i"] += 1
        return kpi_card_pastel(
            theme, title, value,
            tone=_FSS_TONE_CYCLE[idx],
            icon_class=_FSS_ICON_CYCLE[idx],
        )

    # Danh sách 8 card được build TRƯỚC (là list các expression đã hoàn chỉnh),
    # để bên trong list literal chỉ cần tham chiếu lại qua list comprehension.
    _fss_kpi_cards = [
        kpi_fss("ĐIỂM XẾP HẠNG FSS", f"{fss_smart_rank_pct}/100"),
        kpi_fss("ĐÁNH GIÁ CHẤT LƯỢNG", f"{star_rating}/5 ⭐"),
        kpi_fss("VGM PHÂN VỊ", f"{vgm_score_pct}/100"),
        kpi_fss("VGM XẾP HẠNG", f"{vgm_score}"),
        kpi_fss("ĐỊNH GIÁ", f"{value_score}"),
        kpi_fss("TĂNG TRƯỞNG", f"{growth_score}"),
        kpi_fss("ĐỘNG LỰC", f"{momentum_score}"),
        kpi_fss("ĐỊNH GIÁ P/E", f"{pe:,.1f}x" if pd.notna(pe) else "N/A"),
    ]

    # Get all data to calculate percentiles
    df_all = get_snapshot_df()  # Your existing data loader
    size_pct, liquidity_pct, valuation_pct, quality_pct = _calc_fss_components(stock, df_all)

    # Weighted total để nối 87 (Level 1) với 4 factor (Level 2) — theo góp ý:
    # BGK cần thấy rõ 87 đến từ đâu thay vì phải tự suy luận.
    _fss_parts = [(size_pct, 30), (liquidity_pct, 20), (valuation_pct, 20), (quality_pct, 30)]
    if all(p is not None and not (isinstance(p, float) and pd.isna(p)) for p, _ in _fss_parts):
        _fss_weighted_total = sum(p * w for p, w in _fss_parts) / 100
    else:
        _fss_weighted_total = None

    # Arc gauge cho Level 1 (thay số thuần "87/100") — dùng đúng bảng màu FSS hiện
    # tại (T["pastel"]["sky"]["fg"]), KHÔNG dùng skin neon/glow như bản tham khảo,
    # để giữ tông "institutional/terminal" đã có thay vì "AI app" màu mè.
    _fss_gauge_fig = go.Figure(go.Indicator(
        mode="gauge",
        value=fss_smart_rank_pct,
        gauge={
            "axis": {"range": [0, 100], "visible": False},
            "bar": {"color": T["pastel"]["sky"]["fg"], "thickness": 0.30},
            "bgcolor": "rgba(0,0,0,0)",
            "borderwidth": 0,
            "steps": [{"range": [0, 100], "color": T["card_border"]}],
            "shape": "angular",
        },
        domain={"x": [0, 1], "y": [0, 1]},
    ))
    _fss_gauge_fig.update_layout(
        height=150,
        margin=dict(l=16, r=16, t=8, b=0),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
    )

    overview_content = html.Div([

        # --- HEADER ---
        dbc.Row([
            dbc.Col([
                html.Div([
                    html.H3(f"{ticker}", style={"color": T["pastel"]["sky"]["fg"], "display": "inline-block", "marginRight": "15px",
                                                "fontWeight": "bold"}),
                    html.Span(f"{company_name_vn}",
                              style={"color": T["page_text"], "fontSize": "1.2rem", "fontWeight": "normal"}),
                ]),
                html.Div([
                    html.Span("Ngành: ", style={"color": T["page_text_dim"], "fontSize": "0.9rem"}),
                    html.Span(f"{sector}", style={"color": T["positive"], "fontWeight": "bold", "fontSize": "0.9rem",
                                                  "marginRight": "15px"}),
                    # MỚI — đổi thành sub_industry (ngành con thực sự):
                    html.Span("Ngành con: ", style={"color": T["page_text_dim"], "fontSize": "0.9rem"}),
                    html.Span(f"{sub_industry}", style={"color": T["pastel"]["sky"]["fg"], "fontWeight": "bold", "fontSize": "0.9rem"}),
                ], className="mb-3")
            ], width=8),
            dbc.Col([
                html.Div([
                    html.Div("Giá Hiện Tại", style={"color": T["page_text_dim"], "textAlign": "right", "fontSize": "0.9rem"}),
                    html.Div(
                        id="realtime-overview-price",          # ← thêm id này
                        children=f"{price_close:,.0f} VND",
                        style={
                            "textAlign": "right", "fontSize": "28px",
                            "color": "#e6edf3", "fontWeight": "bold",
                            "fontFamily": "'JetBrains Mono', monospace",
                            "transition": "color 0.3s ease",   # ← thêm transition
                        }
                    ),
                ])
            ], width=4)
        ], className="mb-4", style={"borderBottom": f"1px solid {T['card_border']}", "paddingBottom": "15px"}),
        # --- HỒ SƠ DOANH NGHIỆP (PHẲNG, KHÔNG KHUNG) ---
        html.H6([
            html.I(className="fas fa-building", style={"marginRight": "8px", "color": T["pastel"]["sky"]["fg"]}),
            "Hồ sơ Doanh nghiệp"
        ], className="mb-3", style={"fontWeight": "bold", "color": T["page_text"]}),

        dbc.Row([
            # Cột 1: Ngành con
            dbc.Col([
                html.Div([
                    html.Span("Ngành con:", style={
                        "color": T["page_text_dim"], "display": "block",
                        "fontSize": "0.85rem", "marginBottom": "5px"
                    }),
                    html.Span(f"{sub_industry}", style={
                        "color": T["pastel"]["sky"]["fg"], "fontWeight": "600", "fontSize": "0.95rem"
                    })
                ]),
            ], width=3),

            # Cột 2: Sàn giao dịch  ← THÊM MỚI
            dbc.Col([
                html.Div([
                    html.Span("Sàn GD:", style={
                        "color": T["page_text_dim"], "display": "block",
                        "fontSize": "0.85rem", "marginBottom": "5px"
                    }),
                    html.Span(f"{exchange_display}", style={
                        "color": exchange_color, "fontWeight": "700", "fontSize": "0.95rem"
                    })
                ]),
            ], width=2),

            # Cột 3: Năm thành lập  ← đổi width từ 3 → 2
            dbc.Col([
                html.Div([
                    html.Span("Năm thành lập:", style={
                        "color": T["page_text_dim"], "display": "block",
                        "fontSize": "0.85rem", "marginBottom": "5px"
                    }),
                    html.Span(f"{founded_year_str}", style={
                        "color": T["page_text"], "fontWeight": "500", "fontSize": "0.95rem"
                    })
                ]),
            ], width=2),

            # Cột 4: Ngày IPO  ← đổi width từ 3 → 2
            dbc.Col([
                html.Div([
                    html.Span("Ngày IPO:", style={
                        "color": T["page_text_dim"], "display": "block",
                        "fontSize": "0.85rem", "marginBottom": "5px"
                    }),
                    html.Span(f"{ipo_date_str}", style={
                        "color": T["page_text"], "fontWeight": "500", "fontSize": "0.95rem"
                    })
                ]),
            ], width=2),

            # Cột 5: Kiểm toán  ← đổi width từ 3 → 3 (giữ nguyên cho text dài)
            dbc.Col([
                html.Div([
                    html.Span("Kiểm toán:", style={
                        "color": T["page_text_dim"], "display": "block",
                        "fontSize": "0.85rem", "marginBottom": "5px"
                    }),
                    html.Span(f"{auditor}", style={
                        "color": T["page_text"], "fontWeight": "500", "fontSize": "0.9rem"
                    })
                ]),
            ], width=3),
        ], className="mb-5"),

        # --- LƯỚI 8 KPI TIÊU BIỂU ---
        html.H6([html.I(className="fas fa-th", style={"marginRight": "8px", "color": T["pastel"]["sky"]["fg"]}), "Chỉ số nổi bật"],
                className="mb-3", style={"fontWeight": "bold", "color": T["page_text"]}),
        dbc.Row([
            dbc.Col(kpi_card("Vốn hóa TT (Tr. VND)", f"{market_cap:,.0f}" if market_cap > 0 else "N/A"), width=3,
                    className="mb-3"),
            dbc.Col(kpi_card("Số CP lưu hành", f"{shares_out:,.0f}" if pd.notna(shares_out) else "N/A"), width=3,
                    className="mb-3"),
            dbc.Col(kpi_card("Tỷ suất Cổ tức", f"{div_yield:,.1f}%" if div_yield > 0 else "N/A"), width=3,
                    className="mb-3"),
            dbc.Col(kpi_card("Beta", f"{beta:,.2f}" if pd.notna(beta) else "N/A"), width=3, className="mb-3"),
            # Beta = None/NaN nghĩa là mã giao dịch quá thưa (<60% số ngày có
            # biến động giá trong 252 phiên) nên chưa đủ tin cậy để tính —
            # hiển thị "N/A" trung thực thay vì một con số không có căn cứ.
            # LƯU Ý: đổi nhãn từ "P/E (TTM)" → "P/E (Năm gần nhất)" vì EPS
            # dùng ở đây tính từ BCTC năm gần nhất, KHÔNG PHẢI TTM thật (tổng
            # 4 quý gần nhất). Gọi là TTM khi chưa đúng bản chất gây hiểu lầm.
            dbc.Col(kpi_card("P/E (Năm gần nhất)", f"{pe:,.1f}x" if pd.notna(pe) else "N/A"), width=3),
            dbc.Col(kpi_card("P/B", f"{pb:,.2f}x" if pd.notna(pb) else "N/A"), width=3),
            dbc.Col(kpi_card("EPS", f"{eps:,.0f} VND"), width=3),
            dbc.Col(kpi_card("ROE", f"{roe:,.1f}%" if pd.notna(roe) else "N/A"), width=3),
        ], className="mb-5"),

        # --- HỆ THỐNG CHẤM ĐIỂM XẾP HẠNG (FSS SMART RANK) — 3 tầng ---
        html.Div([
            html.Div([
                html.I(className="fas fa-ranking-star", style={"marginRight": "10px", "color": T["pastel"]["sky"]["fg"], "fontSize": "14px"}),
                html.Span("HỆ THỐNG CHẤM ĐIỂM XẾP HẠNG (FSS SMART RANK)", style={
                    "fontSize": "0.78rem", "letterSpacing": "0.12em", "color": T["pastel"]["sky"]["fg"],
                    "fontWeight": "700",
                }),
            ], style={"display": "flex", "alignItems": "center"}),
            html.Button(
                html.I(className="fas fa-circle-info"),
                id="btn-fss-methodology", n_clicks=0,
                title="Xem phương pháp luận chấm điểm FSS Smart Rank",
                style={"background": "none", "border": "none", "cursor": "pointer",
                       "color": T["page_text_dim"], "fontSize": "14px", "padding": "0"},
            ),
        ], className="mb-3", style={"display": "flex", "alignItems": "center", "justifyContent": "space-between"}),

        # LEVEL 1 — Overall score, dạng arc gauge (tham khảo layout bạn gửi, reskin
        # theo màu FSS hiện tại — bỏ glow/neon/animation để giữ tông institutional)
        html.Div([
            html.Div("FSS SMART RANK", style={
                "fontSize": "0.78rem", "letterSpacing": "0.14em", "color": T["page_text_dim"],
                "fontWeight": "700", "textTransform": "uppercase", "textAlign": "center",
            }),
            dcc.Graph(figure=_fss_gauge_fig, config={"displayModeBar": False},
                      style={"height": "150px", "marginTop": "2px", "marginBottom": "-36px"}),
            html.Div([
                html.Span(f"{fss_smart_rank_pct}", style={"fontSize": "2.6rem", "fontWeight": "800", "color": T["page_text"]}),
                html.Span("/100", style={"fontSize": "1.15rem", "color": T["page_text_dim"], "marginLeft": "4px"}),
                html.Span(f" — {_fss_grade(fss_smart_rank_pct)}", style={"fontSize": "1.15rem", "fontWeight": "700",
                          "color": T["pastel"]["sky"]["fg"], "marginLeft": "10px"}),
            ], style={"textAlign": "center"}),
            html.Div(
                "Xếp hạng tương đối trong universe hiện tại — không phải dự báo lợi nhuận hoặc khuyến nghị mua/bán.",
                style={"fontSize": "0.85rem", "color": T["page_text_dim"], "marginTop": "6px",
                       "fontStyle": "italic", "textAlign": "center"}
            ),
        ], style={
            "padding": "18px 20px 20px", "borderRadius": "10px",
            "background": "rgba(30,136,229,0.06)" if theme != "light" else "rgba(15,118,110,0.05)",
            "border": f"1px solid {T['card_border']}", "marginBottom": "16px",
        }),

        # LEVEL 2 — Component scores, lưới 2 cột x 2 hàng (gọn hơn, card trung tính,
        # mô tả dài chuyển sang tooltip ⓘ thay vì luôn hiện trên card)
        dbc.Row([
            dbc.Col(_fss_factor_compact(
                "SIZE (Market Presence)", size_pct, 30,
                "Quy mô/vốn hóa — khả năng tiếp cận vốn & mức phù hợp với nhà đầu tư; KHÔNG đồng nghĩa tốt hơn.",
                f"fss-tip-size-{ticker}"), md=6, className="mb-2"),
            dbc.Col(_fss_factor_compact(
                "LIQUIDITY", liquidity_pct, 20,
                "Khả năng giao dịch/khối lượng — hỗ trợ execution, không phản ánh kỳ vọng lợi nhuận.",
                f"fss-tip-liquidity-{ticker}"), md=6, className="mb-2"),
            dbc.Col(_fss_factor_compact(
                "VALUATION", valuation_pct, 20,
                "Hiện tính từ P/E percentile. Roadmap: kết hợp P/B, EV/EBITDA, P/S theo ngành.",
                f"fss-tip-valuation-{ticker}"), md=6, className="mb-2"),
            dbc.Col(_fss_factor_compact(
                "QUALITY", quality_pct, 30,
                "Hiện tạm dùng Star_Rating × 20. Roadmap: composite độc lập ROE/ROIC, biên lợi nhuận, CFO/NI, đòn bẩy.",
                f"fss-tip-quality-{ticker}"), md=6, className="mb-2"),
        ], className="g-2"),

        # Weighted total — giờ TÍNH Y HỆT công thức backend (xem _calc_fss_components),
        # nên số này sẽ khớp sát FSS Smart Rank ở Level 1 (chỉ lệch nhẹ do làm tròn
        # từng thành phần trước khi cộng), không còn là "minh họa lệch số" như trước.
        html.Div([
            html.Span("Kiểm tra công thức (30/20/20/30): ", style={
                "fontSize": "0.72rem", "color": T["page_text_dim"]}),
            html.Span(
                (f"{_fss_weighted_total:.1f}/100"
                 if _fss_weighted_total is not None else "— (thiếu dữ liệu một số factor)"),
                style={"fontSize": "0.72rem", "color": T["page_text_dim"], "fontWeight": "700",
                       "fontFamily": "JetBrains Mono, monospace"}
            ),
            html.I(className="fas fa-circle-info", id=f"fss-tip-weighted-{ticker}", style={
                "marginLeft": "5px", "fontSize": "0.68rem", "color": T["page_text_dim"], "cursor": "help"}),
            dbc.Tooltip(
                "Cộng lại 4 phân vị bên trên theo đúng trọng số 30% Size + 20% Liquidity + 20% "
                "Valuation + 30% Quality — dùng công thức percentile giống hệt FSS Smart Rank ở "
                "trên, nên hai số sẽ khớp sát nhau (chênh lệch nhỏ nếu có là do làm tròn).",
                target=f"fss-tip-weighted-{ticker}", placement="top", style=_fss_tooltip_style),
        ], style={"textAlign": "right", "marginTop": "6px", "marginBottom": "12px"}),

        # LEVEL 3 — Additional signals (CHIPS, not overall scores)
        html.Div([
            html.Div("TÍN HIỆU PHÂN TÍCH BỔ SUNG — KHÔNG PHẢI ĐIỂM TỔNG", style={
                "fontSize": "0.7rem", "letterSpacing": "0.1em", "color": T["page_text_dim"],
                "fontWeight": "700", "marginBottom": "10px", "textTransform": "uppercase",
            }),
            dbc.Row([
                dbc.Col(_fss_signal_chip("VGM", f"{vgm_score_pct}/100 · {vgm_score}"), width=3),
                dbc.Col(_fss_signal_chip("VALUE", value_score), width=3),
                dbc.Col(_fss_signal_chip("GROWTH", growth_score), width=3),
                dbc.Col(_fss_signal_chip("MOMENTUM", momentum_score), width=3),
            ], className="g-2"),
        ], style={
            "padding": "12px 14px", "borderRadius": "8px",
            "backgroundColor": "rgba(255,255,255,0.02)" if theme != "light" else "rgba(0,0,0,0.02)",
            "border": f"1px dashed {T['card_border']}", "marginBottom": "16px",
        }),

        # --- MODAL: Phương pháp luận FSS Smart Rank ─────────────────────
        dbc.Modal([
            dbc.ModalHeader(
                html.Span([
                    html.I(className="fas fa-ranking-star", style={"color": T["pastel"]["sky"]["fg"], "marginRight": "10px"}),
                    "Phương pháp luận — Hệ Thống Chấm Điểm FSS SMART RANK",
                ], style={"fontFamily": "JetBrains Mono, monospace", "fontSize": "13px",
                           "color": T["page_text"], "fontWeight": "700"}),
                close_button=True,
                style={"backgroundColor": "var(--bg-secondary)", "borderBottom": f"1px solid {T['card_border']}"},
            ),
            dbc.ModalBody([
                html.P(
                    "FSS Smart Rank là điểm xếp hạng tổng hợp kết hợp bốn thành phần quan trọng nhất: "
                    "quy mô doanh nghiệp (Size), khả năng thanh khoản (Liquidity), mức định giá hấp dẫn (Valuation), "
                    "và chất lượng cơ bản (Quality). Mỗi thành phần được tính toán độc lập dựa trên dữ liệu thị trường "
                    "và báo cáo tài chính, sau đó kết hợp theo trọng số đã định để tạo ra một điểm duy nhất từ 0.0 đến 1.0 "
                    "(hoặc 0-100 khi hiển thị dưới dạng phần trăm).",
                    style={"fontSize": "12px", "color": T["page_text_dim"], "lineHeight": "1.7",
                           "marginBottom": "20px"}
                ),
                html.Div([
                    html.H6("Bốn Thành Phần Chính", style={"color": T["pastel"]["sky"]["fg"],
                            "fontWeight": "700", "marginBottom": "12px", "fontSize": "13px", "letterSpacing": "0.05em"}),
                    html.Div([
                        html.Div([
                            html.Span("🏢 Size (Quy Mô) — Trọng Số 30%", style={
                                "fontWeight": "700", "color": T["pastel"]["sky"]["fg"], "fontSize": "11px",
                                "textTransform": "uppercase", "letterSpacing": "0.05em"
                            }),
                        ], style={"marginBottom": "6px"}),
                        html.P(
                            "Đo lường quy mô doanh nghiệp qua vốn hóa thị trường (Market Cap). "
                            "Doanh nghiệp lớn hơn thường có ưu thế cạnh tranh, sức mạnh tài chính, "
                            "và khả năng tồn tại lâu dài cao hơn. Được xếp hạng theo phần vị so với toàn bộ mẫu.",
                            style={"fontSize": "11px", "color": T["page_text_dim"], "lineHeight": "1.6", "marginBottom": "10px"}
                        ),
                    ], style={"backgroundColor": "rgba(88,166,255,0.06)", "padding": "12px", "borderRadius": "6px",
                             "borderLeft": f"3px solid {T['pastel']['sky']['fg']}", "marginBottom": "12px"}),
                    html.Div([
                        html.Div([
                            html.Span("💧 Liquidity (Thanh Khoản) — Trọng Số 20%", style={
                                "fontWeight": "700", "color": "#10b981", "fontSize": "11px",
                                "textTransform": "uppercase", "letterSpacing": "0.05em"
                            }),
                        ], style={"marginBottom": "6px"}),
                        html.P(
                            "Đo lường khả năng giao dịch của cổ phiếu thông qua khối lượng giao dịch trung bình "
                            "(Avg Volume 20D hoặc GTGD_20D). Thanh khoản cao cho phép nhà đầu tư vào/ra dễ dàng "
                            "mà không gây tác động lớn đến giá.",
                            style={"fontSize": "11px", "color": T["page_text_dim"], "lineHeight": "1.6", "marginBottom": "10px"}
                        ),
                    ], style={"backgroundColor": "rgba(16,185,129,0.06)", "padding": "12px", "borderRadius": "6px",
                             "borderLeft": "3px solid #10b981", "marginBottom": "12px"}),
                    html.Div([
                        html.Div([
                            html.Span("💰 Valuation (Định Giá) — Trọng Số 20%", style={
                                "fontWeight": "700", "color": "#f59e0b", "fontSize": "11px",
                                "textTransform": "uppercase", "letterSpacing": "0.05em"
                            }),
                        ], style={"marginBottom": "6px"}),
                        html.P(
                            "Đo lường mức định giá hấp dẫn của cổ phiếu qua tỷ số P/E (Price-to-Earnings). "
                            "P/E thấp hơn là một dấu hiệu tích cực cho nhà đầu tư Value, cho thấy cổ phiếu "
                            "có thể bị định giá thấp so với lợi nhuận thực tế.",
                            style={"fontSize": "11px", "color": T["page_text_dim"], "lineHeight": "1.6", "marginBottom": "10px"}
                        ),
                    ], style={"backgroundColor": "rgba(245,158,11,0.06)", "padding": "12px", "borderRadius": "6px",
                             "borderLeft": "3px solid #f59e0b", "marginBottom": "12px"}),
                    html.Div([
                        html.Div([
                            html.Span("⭐ Quality (Chất Lượng) — Trọng Số 30%", style={
                                "fontWeight": "700", "color": "#ef4444", "fontSize": "11px",
                                "textTransform": "uppercase", "letterSpacing": "0.05em"
                            }),
                        ], style={"marginBottom": "6px"}),
                        html.P(
                            "Đo lường chất lượng cơ bản của doanh nghiệp qua Star_Rating (1-5 sao). "
                            "Star_Rating là tổng hợp của VGM Score (Value + Growth + Momentum) "
                            "với các điều chỉnh phạt lỗi (CFO âm, thanh khoản thấp). Trọng số cao nhất (30%) "
                            "nhấn mạnh tầm quan trọng của chất lượng cơ bản.",
                            style={"fontSize": "11px", "color": T["page_text_dim"], "lineHeight": "1.6", "marginBottom": "10px"}
                        ),
                    ], style={"backgroundColor": "rgba(239,68,68,0.06)", "padding": "12px", "borderRadius": "6px",
                             "borderLeft": "3px solid #ef4444", "marginBottom": "12px"}),
                ], style={"marginBottom": "20px"}),

                html.Hr(style={"borderColor": T["card_border"], "margin": "16px 0"}),

                html.Div([
                    html.H6("Công Thức Tính", style={"color": T["pastel"]["sky"]["fg"],
                            "fontWeight": "700", "marginBottom": "10px", "fontSize": "13px", "letterSpacing": "0.05em"}),
                    html.P(
                        "FSS Smart Rank = Rank(Size) × 0.30 + Rank(Liquidity) × 0.20 + Rank(Valuation) × 0.20 + Rank(Quality) × 0.30",
                        style={"fontFamily": "JetBrains Mono, monospace", "fontSize": "11px", "color": T["positive"],
                               "backgroundColor": "rgba(16,185,129,0.08)", "padding": "10px", "borderRadius": "4px",
                               "lineHeight": "1.6", "marginBottom": "10px"}
                    ),
                    html.P(
                        "Kết quả là một con số từ 0.0 đến 1.0 (được hiển thị dưới dạng 0-100 để dễ hiểu). "
                        "Điểm càng cao, mã cổ phiếu càng đạt tiêu chí toàn diện.",
                        style={"fontSize": "11px", "color": T["page_text_dim"], "lineHeight": "1.6"}
                    ),
                ], style={"marginBottom": "16px"}),

                html.Hr(style={"borderColor": T["card_border"], "margin": "16px 0"}),

                html.Div([
                    html.H6("Tại Sao FSS Smart Rank Quan Trọng?", style={"color": T["pastel"]["sky"]["fg"],
                            "fontWeight": "700", "marginBottom": "10px", "fontSize": "13px", "letterSpacing": "0.05em"}),
                    html.Ul([
                        html.Li("📊 Tổng hợp đa chiều: Không chỉ nhìn một chỉ số, mà kết hợp 4 khía cạnh quan trọng",
                               style={"fontSize": "11px", "color": T["page_text_dim"], "marginBottom": "6px"}),
                        html.Li("⚖️ Cân bằng rủi ro: Trọng số được thiết kế để cân bằng tầm quan trọng của từng yếu tố",
                               style={"fontSize": "11px", "color": T["page_text_dim"], "marginBottom": "6px"}),
                        html.Li("🎯 Giải quyết vấn đề phân biệt: Tránh được tình trạng \"quá nhiều mã 5 sao mất khả năng phân biệt\"",
                               style={"fontSize": "11px", "color": T["page_text_dim"], "marginBottom": "6px"}),
                        html.Li("📈 Dễ sắp xếp: Được dùng làm cơ sở xếp hạng chính trong bảng dữ liệu",
                               style={"fontSize": "11px", "color": T["page_text_dim"]}),
                    ], style={"marginLeft": "20px", "marginBottom": "10px"}),
                ], style={"marginBottom": "16px"}),

            ], style={"maxHeight": "600px", "overflowY": "auto"}),
        ], id="modal-fss-methodology", size="lg", is_open=False, centered=True, backdrop=True),

        # --- KHỐI ĐÁNH GIÁ SỨC KHỎE CHI TIẾT VÀ BIỂU ĐỒ ---
        html.Div([
            html.Div([
                html.I(className="fas fa-heartbeat", style={"marginRight": "10px", "color": T["positive"], "fontSize": "14px"}),
                html.Span("BÁO CÁO SỨC KHỎE TÀI CHÍNH", style={
                    "fontSize": "0.72rem", "letterSpacing": "0.12em", "color": T["positive"],
                    "fontWeight": "700",
                }),
            ], style={"display": "flex", "alignItems": "center"}),
            # Nút ⓘ — mở modal giải thích phương pháp luận
            html.Button(
                html.I(className="fas fa-circle-info"),
                id="btn-health-methodology",
                n_clicks=0,
                title="Xem phương pháp luận chấm điểm",
                style={
                    "background": "none", "border": "none", "cursor": "pointer",
                    "color": T["page_text_dim"], "fontSize": "14px", "padding": "0",
                    "transition": "color 0.2s",
                },
            ),
        ], className="mb-3", style={"display": "flex", "alignItems": "center", "justifyContent": "space-between"}),

        # ── Modal: Phương pháp luận chấm điểm ──────────────────────────────
        dbc.Modal([
            dbc.ModalHeader(
                html.Span([
                    html.I(className="fas fa-heartbeat", style={"color": "#00e676", "marginRight": "10px"}),
                    "Phương pháp luận — Báo cáo Sức khỏe Tài chính",
                ], style={"fontFamily": "JetBrains Mono,monospace", "fontSize": "13px",
                           "color": "#c9d1d9", "fontWeight": "700"}),
                close_button=True,
                style={"backgroundColor": "#0d1117", "borderBottom": "1px solid #21262d"},
            ),
            dbc.ModalBody([
                # Intro
                html.P(
                    "Hệ thống không phải là công cụ khuyến nghị đầu tư — đây là lăng kính định lượng "
                    "đo lường sức khỏe tài chính và lợi thế cạnh tranh của doanh nghiệp, "
                    "giúp bạn đầu tư bài bản và bền vững hơn.",
                    style={"fontSize": "12px", "color": "#8b949e", "lineHeight": "1.7",
                           "marginBottom": "16px", "fontStyle": "italic"}
                ),

                # Quy trình 3 bước
                _meth_section("fas fa-gears", "#58a6ff", "Quy trình Phân tích 3 Bước"),
                _meth_step("1", "#58a6ff",
                    "Thu thập & Xử lý Dữ liệu",
                    "Dữ liệu Báo cáo Tài chính từ các nguồn công khai, đáng tin cậy."),
                _meth_step("2", "#58a6ff",
                    "Phân tích Chuyên sâu theo Ngành",
                    "Dựa trên đặc thù của từng ngành, hệ thống áp dụng mô hình phân tích riêng "
                    "với thang chấm điểm riêng. Các chỉ số được đánh giá trong bối cảnh ngành "
                    "đó để đảm bảo tính khách quan."),
                _meth_step("3", "#58a6ff",
                    "Chấm điểm & Tổng hợp",
                    "Kết quả được tổng hợp bằng hệ thống tính điểm có trọng số, đưa ra điểm "
                    "\"Sức Khỏe Tài Chính\" tổng thể trên thang điểm 100 và các báo cáo trực quan."),

                html.Hr(style={"borderColor": "#21262d", "margin": "16px 0"}),

                # 5 chỉ số hiện tại
                _meth_section("fas fa-chart-bar", "#00e676", "Các Chỉ số Đánh giá Hiện tại"),
                _meth_metric("Biên Lợi nhuận Gộp (Gross Margin)",
                    "Phản ánh lợi thế cạnh tranh và hiệu quả chi phí bền vững. "
                    "≥ 30% = Rất Tốt | 20–30% = Tốt | 10–20% = Trung Bình | < 10% = Yếu.",
                    "Nền tảng Philip Fisher: ưu tiên doanh nghiệp có \"con hào kinh tế\" thể hiện "
                    "qua biên lợi nhuận gộp cao và ổn định nhiều năm."),
                _meth_metric("Nợ vay / Vốn chủ sở hữu (D/E)",
                    "Cấu trúc vốn an toàn là yếu tố sống còn để vượt qua giai đoạn khó khăn. "
                    "≤ 0.5x = Rất Tốt | 0.5–1.0x = Tốt | 1.0–1.5x = Trung Bình | > 1.5x = Yếu.",
                    "Chỉ số này đặc biệt quan trọng với ngành sản xuất và bất động sản nơi chu kỳ "
                    "lãi suất tác động mạnh đến khả năng trả nợ."),
                _meth_metric("OCF / Lợi nhuận ròng (Chất lượng Lợi nhuận)",
                    "Lợi nhuận có thực sự chuyển hóa thành tiền mặt? "
                    "≥ 1.5x = Rất Tốt | 1.0–1.5x = Tốt | 0.5–1.0x = Trung Bình | < 0.5x = Yếu.",
                    "Lợi nhuận kế toán có thể bị \"thổi phồng\" bởi doanh thu ghi nhận nhưng chưa "
                    "thu tiền. OCF/Net Income > 1 xác nhận lợi nhuận là thực chất."),
                _meth_metric("Số ngày Tồn kho (Inventory Days)",
                    "Quản trị tồn kho hiệu quả giảm thiểu rủi ro khi giá hàng hóa biến động. "
                    "≤ 30 ngày = Rất Tốt | 30–60 ngày = Tốt | 60–90 ngày = Trung Bình | > 90 ngày = Yếu.",
                    "Đặc biệt quan trọng với ngành sản xuất (thép, thủy sản, dệt may) nơi hàng "
                    "tồn kho lớn có thể dẫn đến lỗ nặng khi giá nguyên liệu giảm."),
                _meth_metric("Định giá EV/EBITDA",
                    "Chỉ số định giá tiêu chuẩn so sánh công bằng giữa các cấu trúc vốn khác nhau. "
                    "≤ 5x = Rất Tốt | 5–10x = Tốt | 10–15x = Trung Bình | > 15x = Yếu.",
                    "EV/EBITDA loại bỏ ảnh hưởng của cấu trúc nợ và khấu hao, cho phép so sánh "
                    "định giá công bằng giữa doanh nghiệp thâm dụng vốn và doanh nghiệp nhẹ vốn."),

                html.Hr(style={"borderColor": "#21262d", "margin": "16px 0"}),

                # Nguồn cảm hứng
                _meth_section("fas fa-book-open", "#d2a8ff", "Nguồn Cảm Hứng Phân Tích"),
                html.Div([
                    html.Div([
                        html.Span("Philip Fisher — Common Stocks and Uncommon Profits",
                                  style={"fontSize": "12px", "fontWeight": "700", "color": "#c9d1d9"}),
                        html.P(
                            "Ưu tiên doanh nghiệp có sức khỏe tài chính vững chắc, lợi thế cạnh tranh "
                            "bền vững (\"con hào kinh tế\") và dòng tiền mạnh. Biên lợi nhuận gộp cao "
                            "và ổn định là dấu hiệu của sản phẩm/dịch vụ khó thay thế.",
                            style={"fontSize": "11px", "color": "#8b949e", "lineHeight": "1.6",
                                   "marginBottom": "10px", "marginTop": "4px"}
                        ),
                    ]),
                    html.Div([
                        html.Span("William O'Neil — How to Make Money in Stocks (CANSLIM)",
                                  style={"fontSize": "12px", "fontWeight": "700", "color": "#c9d1d9"}),
                        html.P(
                            "Chú trọng vào yếu tố tăng trưởng bùng nổ về doanh thu và lợi nhuận "
                            "trong các quý gần nhất — là chất xúc tác cho giá cổ phiếu. "
                            "Động lượng tăng trưởng là yếu tố phân biệt \"siêu cổ phiếu\" với phần còn lại.",
                            style={"fontSize": "11px", "color": "#8b949e", "lineHeight": "1.6",
                                   "marginBottom": "0", "marginTop": "4px"}
                        ),
                    ]),
                ], style={"padding": "12px", "backgroundColor": "rgba(210,168,255,0.06)",
                           "borderRadius": "8px", "border": "1px solid rgba(210,168,255,0.15)"}),

                html.Hr(style={"borderColor": "#21262d", "margin": "16px 0"}),

                # Disclaimer
                html.Div([
                    html.I(className="fas fa-triangle-exclamation",
                           style={"color": "#e3b341", "marginRight": "8px", "fontSize": "11px"}),
                    html.Span(
                        "TUYÊN BỐ MIỄN TRỪ TRÁCH NHIỆM: Mọi phân tích và điểm số chỉ mang tính "
                        "chất tham khảo, không được xem là lời khuyên đầu tư. Nhà đầu tư cần thực "
                        "hiện các phân tích sâu hơn và chịu hoàn toàn trách nhiệm cho quyết định của mình.",
                        style={"fontSize": "10px", "color": "#7d6608", "lineHeight": "1.5"}
                    ),
                ], style={
                    "backgroundColor": "rgba(227,179,65,0.08)",
                    "border": "1px solid rgba(227,179,65,0.2)",
                    "borderRadius": "6px", "padding": "10px 12px",
                    "display": "flex", "alignItems": "flex-start",
                }),
            ], style={"backgroundColor": "#0d1117", "padding": "20px"}),
        ],
            id="health-methodology-modal",
            is_open=False,
            centered=True,
            size="lg",
            scrollable=True,
            style={"fontFamily": "'Sora', sans-serif"},
            contentClassName="border-0",
        ),
        dbc.Row([
            # Bên trái: Biểu đồ
            dbc.Col([
                html.Div([
                    html.Div("LỊCH SỬ SỨC KHỎE TÀI CHÍNH (8 QUÝ)", style={
                        "fontSize": "0.68rem", "letterSpacing": "0.1em", "color": T["page_text_dim"],
                        "fontWeight": "600", "textTransform": "uppercase", "marginBottom": "4px",
                        "paddingLeft": "4px"
                    }),
                    dcc.Graph(figure=fig_health, config={"displayModeBar": False},
                              style={"height": "640px"})
                ], style={
                    "background": T["pastel"]["sky"]["bg"],
                    "borderRadius": "12px", "border": f"1px solid {T['card_border']}",
                    "padding": "14px 14px 6px",
                    "boxShadow": T["card_shadow"]
                })
            ], width=6),

            # Bên phải: Progress Bars
            dbc.Col([
                html.Div([
                    html.Div([
                        html.Span(f"{total_health_score}", style={
                            "fontSize": "52px", "fontWeight": "900", "letterSpacing": "-0.04em",
                            "color": T["positive"] if total_health_score >= 70 else (
                                ("#b45309" if theme == "light" else "#ffb703") if total_health_score >= 50 else T["negative"]),
                        }),
                        html.Span("/100", style={"fontSize": "20px", "color": T["page_text_dim"],
                                                 "marginLeft": "4px"}),
                    ], style={"marginBottom": "4px", "display": "flex", "alignItems": "baseline"}),
                    html.Div("ĐIỂM SỨC KHỎE TỔNG HỢP",
                             style={"fontSize": "0.9rem", "color": T["page_text_dim"], "marginBottom": "20px"}),

                    make_progress_bar("Biên LNG", f"{gross_margin:.1f}%", score_gm, label_gm, color_gm,
                                      "Phản ánh lợi thế cạnh tranh và hiệu quả chi phí một cách bền vững."),
                    make_progress_bar("Nợ vay / VCSH", f"{debt_equity:.2f}x", score_de, label_de, color_de,
                                      "Cấu trúc vốn an toàn là yếu tố sống còn để vượt qua giai đoạn khó khăn của chu kỳ."),
                    make_progress_bar("OCF / Lợi nhuận ròng", f"{ocf_net:.2f}x", score_ocf, label_ocf, color_ocf,
                                      "Cho thấy chất lượng của lợi nhuận, lợi nhuận có thực sự chuyển hóa thành tiền mặt hay không."),
                    make_progress_bar("Số ngày tồn kho", f"{inv_days:.0f} ngày", score_inv, label_inv, color_inv,
                                      "Quản trị hàng tồn kho hiệu quả giúp giảm thiểu rủi ro khi giá hàng hóa biến động."),
                    make_progress_bar("Định giá EV/EBITDA", f"{ev_ebitda:.1f}x", score_ev, label_ev, color_ev,
                                      "Chỉ số định giá tiêu chuẩn giúp so sánh công bằng cấu trúc vốn giữa các công ty.")
                ], style={
                    "background": T["pastel"]["green"]["bg"],
                    "padding": "20px", "borderRadius": "12px",
                    "border": f"1px solid {T['card_border']}",
                    "boxShadow": T["card_shadow"]
                })
            ], width=6)
        ])

    ], style={"padding": "20px"})

    
    # Callback 2B: chỉ trả overview
    return overview_content

# ============================================================================
# CALLBACK 2C: LOAD TAB KỸ THUẬT — CHỈ KHI CLICK TAB NÀY
# ============================================================================
@app.callback(
    Output("tab-technical-content", "children"),
    Input("detail-tabs",            "active_tab"),
    State("selected-stock-store",   "data"),
    State("theme-store",            "data"),
    prevent_initial_call=True,
)
def load_technical_tab(active_tab, stock, theme="dark"):
    if active_tab != "tab-technical" or not stock:
        return no_update
    theme = theme or "dark"
    from src.utils.kpi_theme import get_kpi_theme
    T = get_kpi_theme(theme)

    ticker = stock.get('Ticker', 'N/A')

    try:
        df_price = load_market_data()
        df_tech  = df_price[df_price['Ticker'] == ticker].copy()
        del df_price
        gc.collect()

        if len(df_tech) < 50:
            return html.Div("Không đủ dữ liệu giá (cần ít nhất 50 phiên) để tính toán kỹ thuật.",
                            className="text-muted text-center p-5")

        df_tech     = df_tech.sort_values('Date')
        close_price = df_tech['Price Close'].iloc[-1]

        sma10  = df_tech['Price Close'].rolling(10).mean().iloc[-1]
        sma20  = df_tech['Price Close'].rolling(20).mean().iloc[-1]
        sma50  = df_tech['Price Close'].rolling(50).mean().iloc[-1]
        sma200 = df_tech['Price Close'].rolling(200).mean().iloc[-1] if len(df_tech) >= 200 else np.nan
        ema10  = df_tech['Price Close'].ewm(span=10, adjust=False).mean().iloc[-1]
        ema20  = df_tech['Price Close'].ewm(span=20, adjust=False).mean().iloc[-1]
        ema50  = df_tech['Price Close'].ewm(span=50, adjust=False).mean().iloc[-1]

        def eval_ma(val):
            if pd.isna(val): return "N/A", "#8b949e"
            return ("MUA", "#3fb950") if close_price > val else ("BÁN", "#f85149")

        delta = df_tech['Price Close'].diff()
        gain  = (delta.where(delta > 0, 0)).rolling(14).mean()
        loss  = (-delta.where(delta < 0, 0)).rolling(14).mean()
        rsi   = 100 - (100 / (1 + gain / loss))
        rsi_val = rsi.iloc[-1]

        if rsi_val < 30:   sig_rsi, col_rsi = "MUA MẠNH (Quá bán)", "#3fb950"
        elif rsi_val > 70: sig_rsi, col_rsi = "BÁN MẠNH (Quá mua)", "#f85149"
        else:              sig_rsi, col_rsi = "TRUNG TÍNH",          "#8b949e"

        ema12     = df_tech['Price Close'].ewm(span=12, adjust=False).mean()
        ema26     = df_tech['Price Close'].ewm(span=26, adjust=False).mean()
        macd_line = ema12 - ema26
        sig_line  = macd_line.ewm(span=9, adjust=False).mean()
        macd_val  = macd_line.iloc[-1]
        macd_sig  = sig_line.iloc[-1]
        sig_macd, col_macd = ("MUA", "#3fb950") if macd_val > macd_sig else ("BÁN", "#f85149")

        low14   = df_tech['Price Low'].rolling(14).min()
        high14  = df_tech['Price High'].rolling(14).max()
        k_pct   = 100 * ((df_tech['Price Close'] - low14) / (high14 - low14))
        d_pct   = k_pct.rolling(3).mean()
        stoch_k = k_pct.iloc[-1]
        stoch_d = d_pct.iloc[-1]

        if stoch_k > stoch_d and stoch_k < 20:   sig_stoch, col_stoch = "MUA",       "#3fb950"
        elif stoch_k < stoch_d and stoch_k > 80: sig_stoch, col_stoch = "BÁN",       "#f85149"
        else:                                     sig_stoch, col_stoch = "TRUNG TÍNH","#8b949e"

        prev_h = df_tech['Price High'].iloc[-2]
        prev_l = df_tech['Price Low'].iloc[-2]
        prev_c = df_tech['Price Close'].iloc[-2]
        pp = (prev_h + prev_l + prev_c) / 3
        r1 = 2 * pp - prev_l;  s1 = 2 * pp - prev_h
        r2 = pp + (prev_h - prev_l); s2 = pp - (prev_h - prev_l)
        r3 = prev_h + 2 * (pp - prev_l); s3 = prev_l - 2 * (prev_h - pp)

        buy_count = sell_count = 0
        signals   = [
            eval_ma(sma10)[0], eval_ma(sma20)[0], eval_ma(sma50)[0], eval_ma(sma200)[0],
            eval_ma(ema10)[0], eval_ma(ema20)[0], eval_ma(ema50)[0],
            "MUA" if rsi_val < 40 else ("BÁN" if rsi_val > 60 else "TRUNG TÍNH"),
            sig_macd,
            "MUA" if stoch_k > stoch_d else "BÁN"
        ]
        for s in signals:
            if s.startswith("MUA"):   buy_count  += 1
            elif s.startswith("BÁN"): sell_count += 1

        total_signals = buy_count + sell_count
        meter_score   = ((buy_count - sell_count) / max(total_signals, 1)) * 100

        if meter_score >= 50:    meter_text, meter_color = "MUA MẠNH", "#3fb950"
        elif meter_score >= 10:  meter_text, meter_color = "MUA",      "#2ea043"
        elif meter_score <= -50: meter_text, meter_color = "BÁN MẠNH", "#f85149"
        elif meter_score <= -10: meter_text, meter_color = "BÁN",      "#da3633"
        else:                    meter_text, meter_color = "TRUNG TÍNH","#8b949e"

        arc_color = T["negative"] if meter_score < -10 else (T["positive"] if meter_score > 10 else ("#b45309" if theme == "light" else "#ffb703"))
        gauge_tick_color = "rgba(15,23,42,0.45)" if theme == "light" else "rgba(255,255,255,0.3)"
        gauge_neutral_bg = "rgba(15,23,42,0.04)" if theme == "light" else "rgba(255,255,255,0.03)"
        neg_bg_strong = "rgba(220,38,38,0.18)" if theme == "light" else "rgba(255,77,109,0.18)"
        neg_bg_soft   = "rgba(220,38,38,0.08)" if theme == "light" else "rgba(255,77,109,0.08)"
        pos_bg_soft   = "rgba(21,128,61,0.08)" if theme == "light" else "rgba(0,255,200,0.08)"
        pos_bg_strong = "rgba(21,128,61,0.18)" if theme == "light" else "rgba(0,255,200,0.18)"
        fig_gauge = go.Figure(go.Indicator(
            mode="gauge", value=meter_score,
            gauge={'axis': {'range': [-100, 100], 'tickwidth': 1, 'tickcolor': gauge_tick_color,
                            'tickvals': [-100, -50, 0, 50, 100],
                            'tickfont': {'color': gauge_tick_color, 'size': 9}},
                   'bar': {'color': arc_color, 'thickness': 0.3},
                   'bgcolor': 'rgba(0,0,0,0)', 'borderwidth': 0,
                   'steps': [
                       {'range': [-100, -60], 'color': neg_bg_strong},
                       {'range': [-60,  -25], 'color': neg_bg_soft},
                       {'range': [-25,   25], 'color': gauge_neutral_bg},
                       {'range': [25,    60], 'color': pos_bg_soft},
                       {'range': [60,   100], 'color': pos_bg_strong},
                   ],
                   'threshold': {'line': {'color': arc_color, 'width': 3}, 'thickness': 0.85, 'value': meter_score}},
        ))
        # FIX: trước không set "width" nên Plotly mặc định render figure rộng hơn
        # nhiều so với cột chứa nó (300px) → gauge tràn ra ngoài viền khung. Giờ khoá
        # cứng width=230 để vừa khít cột (300px - 2*24px padding trái/phải ≈ 252px).
        fig_gauge.update_layout(height=200, width=230, margin=dict(l=10, r=10, t=20, b=5),
                                paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                                font={'color': gauge_tick_color})

        def ind_row(name, val, sig, col):
            val_str = f"{val:,.2f}" if not pd.isna(val) else "N/A"
            chip_bg = {"#3fb950":"rgba(63,185,80,0.15)" if theme=="light" else "rgba(0,230,118,0.15)",
                       "#2ea043":"rgba(63,185,80,0.10)" if theme=="light" else "rgba(0,230,118,0.1)",
                       "#f85149":"rgba(248,81,73,0.15)" if theme=="light" else "rgba(255,61,87,0.15)",
                       "#da3633":"rgba(248,81,73,0.10)" if theme=="light" else "rgba(255,61,87,0.1)",
                       "#8b949e":"rgba(139,148,158,0.12)" if theme=="light" else "rgba(139,148,158,0.1)"}.get(col, T["pastel"]["sky"]["icon_bg"])
            chip_bd = {"#3fb950":"rgba(63,185,80,0.4)","#2ea043":"rgba(63,185,80,0.3)",
                       "#f85149":"rgba(248,81,73,0.4)","#da3633":"rgba(248,81,73,0.3)",
                       "#8b949e":"rgba(139,148,158,0.3)"}.get(col, "rgba(56,189,248,0.3)")
            return html.Tr([
                html.Td(html.Span(name, style={"color":T["page_text"],"fontSize":"0.83rem","fontWeight":"500"})),
                html.Td(html.Span(val_str, style={"color":T["pastel"]["sky"]["fg"],"fontSize":"0.85rem","fontWeight":"700","float":"right"})),
                html.Td(html.Span(sig, style={"color":col,"fontSize":"0.73rem","fontWeight":"800","padding":"3px 10px",
                                              "borderRadius":"4px","backgroundColor":chip_bg,"border":f"1px solid {chip_bd}",
                                              "float":"right","whiteSpace":"nowrap"}))
            ], style={"borderBottom":f"1px solid {T['card_border']}"})

        def pivot_card(label, value, color):
            return html.Div([
                html.Div(label, style={"fontSize":"0.68rem","letterSpacing":"0.1em","textTransform":"uppercase",
                                       "color":color,"opacity":"0.85","marginBottom":"6px","fontWeight":"600"}),
                html.Div(f"{value:,.0f}", style={"fontSize":"1.1rem","fontWeight":"800","color":color,
                                                  "letterSpacing":"-0.02em"})
            ], style={"textAlign":"center","padding":"12px 8px","borderRadius":"8px",
                      "background":T["pastel"]["rose"]["bg"] if "f8" in color or "ff3" in color or "dc" in color or "f85" in color else (T["pastel"]["green"]["bg"]),
                      "border":f"1px solid {color}33","borderTop":f"2px solid {color}99"})

        neutral_count = len(signals) - buy_count - sell_count
        neutral_text_color = "rgba(15,23,42,0.55)" if theme == "light" else "rgba(255,255,255,0.55)"
        neutral_track_bg   = "rgba(15,23,42,0.08)" if theme == "light" else "rgba(255,255,255,0.05)"
        neutral_card_bg    = "rgba(15,23,42,0.04)" if theme == "light" else "rgba(255,255,255,0.04)"
        neutral_card_bd    = "rgba(15,23,42,0.12)" if theme == "light" else "rgba(255,255,255,0.1)"

        technical_content = html.Div([
            html.Div([
                # top accent line — dùng meter_color động (đổi màu theo tín hiệu MUA/BÁN/TRUNG TÍNH)
                # thay vì cyan cố định như bản tham khảo, để không vỡ theme sáng/tối
                html.Div(style={
                    "position": "absolute", "top": "0", "left": "8%", "right": "8%", "height": "2px",
                    "background": f"linear-gradient(90deg,transparent,{meter_color},transparent)",
                    "boxShadow": f"0 0 12px {meter_color}90",
                }),
                html.Div([
                    html.Div([
                        html.Div(style={"flex": "1", "height": "1px",
                                        "background": f"linear-gradient(90deg,{meter_color}4d,transparent)"}),
                        html.Span("SIGNAL METER", style={"fontSize": "8px", "letterSpacing": "0.3em",
                                   "color": T["page_text_dim"], "fontWeight": "700", "padding": "0 10px"}),
                        html.Div(style={"flex": "1", "height": "1px",
                                        "background": f"linear-gradient(90deg,transparent,{meter_color}4d)"}),
                    ], style={"display": "flex", "alignItems": "center", "marginBottom": "10px", "width": "100%"}),
                    dcc.Graph(figure=fig_gauge, config={"displayModeBar": False},
                              style={"height": "200px", "width": "230px", "marginBottom": "-20px"}),
                    html.Div(f"{meter_score:+.0f}", style={"fontSize": "5.5rem", "fontWeight": "800", "color": meter_color,
                                                            "letterSpacing": "-4px", "lineHeight": "1", "textAlign": "center",
                                                            "textShadow": f"0 0 18px {meter_color}80, 0 0 42px {meter_color}30"}),
                    html.Div([
                        html.Div(style={"width": "6px", "height": "6px", "borderRadius": "50%", "background": meter_color,
                                        "marginRight": "8px", "boxShadow": f"0 0 6px {meter_color}"}),
                        html.Span(f"{meter_text}", style={"fontSize": "10px", "fontWeight": "700", "letterSpacing": "0.18em", "color": meter_color}),
                    ], style={"display": "inline-flex", "alignItems": "center", "marginTop": "12px",
                              "background": f"{meter_color}1a", "border": f"1px solid {meter_color}66",
                              "borderRadius": "100px", "padding": "5px 14px"}),
                ], style={"display": "flex", "flexDirection": "column", "alignItems": "center", "width": "300px",
                          "flexShrink": "0", "paddingRight": "24px", "borderRight": f"1px solid {T['card_border']}"}),
                html.Div([
                    html.Div([
                        html.Div(style={"width": "7px", "height": "7px", "borderRadius": "50%", "background": meter_color,
                                        "boxShadow": f"0 0 8px {meter_color}, 0 0 16px {meter_color}80",
                                        "marginRight": "8px", "flexShrink": "0",
                                        "animation": "realtime-pulse 2s ease-in-out infinite"}),
                        html.Span("PHÂN TÍCH KỸ THUẬT · REAL-TIME", style={"fontSize": "8px", "letterSpacing": "0.22em",
                                   "color": T["page_text_dim"], "fontWeight": "700"}),
                    ], style={"display": "flex", "alignItems": "center", "marginBottom": "16px"}),
                    html.Div("10 CHỈ BÁO · MA + OSCILLATORS", style={"fontSize": "9px", "letterSpacing": "0.15em",
                                                                     "color": T["page_text_dim"], "marginBottom": "8px"}),
                    html.Div(meter_text, style={"fontSize": "3.2rem", "fontWeight": "800", "color": meter_color,
                                                "letterSpacing": "0.04em", "lineHeight": "1",
                                                "marginBottom": "22px",
                                                "textShadow": f"0 0 14px {meter_color}70, 0 0 34px {meter_color}25"}),
                    *[html.Div([
                        html.Span(lbl, style={"fontSize":"9px","color":neutral_text_color,"width":"70px","flexShrink":"0"}),
                        html.Div(html.Div(style={"height":"100%","borderRadius":"3px","width":f"{pct}%",
                                                 "background":bg,"transition":"width 1.4s cubic-bezier(.4,0,.2,1)"}),
                                 style={"flex":"1","height":"5px","background":neutral_track_bg,"borderRadius":"3px","overflow":"hidden"}),
                        html.Span(str(cnt), style={"fontSize":"10px","fontWeight":"700","color":col,
                                                    "width":"14px","textAlign":"right"}),
                    ], style={"display":"flex","alignItems":"center","gap":"10px","marginBottom":"8px"})
                      for lbl, pct, cnt, bg, col in [
                        ("Bán",       sell_count/len(signals)*100, sell_count,
                         f"linear-gradient(90deg,{T['negative']}4d,{T['negative']})", T['negative']),
                        ("Mua",       buy_count/len(signals)*100,  buy_count,
                         f"linear-gradient(90deg,{T['positive']}33,{T['positive']})", T['positive']),
                        ("Trung tính",neutral_count/len(signals)*100,
                         neutral_count, neutral_track_bg, neutral_text_color),
                      ]],
                    dbc.Row([
                        dbc.Col(html.Div([
                            html.Div(str(sell_count), style={"fontSize":"2.4rem","fontWeight":"800","color":T['negative']}),
                            html.Div("BÁN", style={"fontSize":"8px","fontWeight":"700","letterSpacing":"0.2em","color":f"{T['negative']}8c"}),
                        ], style={"borderRadius":"12px","padding":"16px 10px 14px","textAlign":"center",
                                  "background":T["pastel"]["rose"]["bg"],
                                  "border":f"1px solid {T['negative']}59"})),
                        dbc.Col(html.Div([
                            html.Div(str(buy_count), style={"fontSize":"2.4rem","fontWeight":"800","color":T['positive']}),
                            html.Div("MUA", style={"fontSize":"8px","fontWeight":"700","letterSpacing":"0.2em","color":f"{T['positive']}80"}),
                        ], style={"borderRadius":"12px","padding":"16px 10px 14px","textAlign":"center",
                                  "background":T["pastel"]["green"]["bg"],
                                  "border":f"1px solid {T['positive']}47"})),
                        dbc.Col(html.Div([
                            html.Div(str(neutral_count), style={"fontSize":"2.4rem","fontWeight":"800","color":neutral_text_color}),
                            html.Div("TRUNG TÍNH", style={"fontSize":"8px","fontWeight":"700","letterSpacing":"0.2em","color":neutral_text_color}),
                        ], style={"borderRadius":"12px","padding":"16px 10px 14px","textAlign":"center",
                                  "background":neutral_card_bg,
                                  "border":f"1px solid {neutral_card_bd}"})),
                    ], className="g-2"),
                ], style={"flex":"1","display":"flex","flexDirection":"column","justifyContent":"center"}),

            ], style={"display":"flex","alignItems":"center","gap":"28px", "background":T["pastel"]["sky"]["bg"],
                      "borderRadius":"16px","border":f"1px solid {T['card_border']}", "position": "relative", "overflow": "hidden",
                      "padding":"24px 28px","marginBottom":"16px","boxShadow":T["card_shadow"]}),

            dbc.Row([
                dbc.Col([
                    html.Div("TRUNG BÌNH ĐỘNG", style={"fontSize":"0.72rem","letterSpacing":"0.1em","fontWeight":"700",
                                                        "color":T["pastel"]["sky"]["fg"],"marginBottom":"12px"}),
                    html.Table([
                        html.Thead(html.Tr([
                            html.Th("Chỉ báo",  style={"color":T["page_text_dim"],"fontSize":"0.7rem","fontWeight":"700","paddingBottom":"8px","borderBottom":f"1px solid {T['card_border']}"}),
                            html.Th("Giá trị",  style={"color":T["page_text_dim"],"fontSize":"0.7rem","fontWeight":"700","textAlign":"right","paddingBottom":"8px","borderBottom":f"1px solid {T['card_border']}"}),
                            html.Th("Tín hiệu", style={"color":T["page_text_dim"],"fontSize":"0.7rem","fontWeight":"700","textAlign":"right","paddingBottom":"8px","borderBottom":f"1px solid {T['card_border']}"}),
                        ])),
                        html.Tbody([
                            ind_row("SMA 10",  sma10,  *eval_ma(sma10)),
                            ind_row("SMA 20",  sma20,  *eval_ma(sma20)),
                            ind_row("SMA 50",  sma50,  *eval_ma(sma50)),
                            ind_row("SMA 200", sma200, *eval_ma(sma200)),
                            ind_row("EMA 10",  ema10,  *eval_ma(ema10)),
                            ind_row("EMA 20",  ema20,  *eval_ma(ema20)),
                            ind_row("EMA 50",  ema50,  *eval_ma(ema50)),
                        ])
                    ], style={"width":"100%","borderCollapse":"collapse"})
                ], width=12, lg=6, style={"background":T["pastel"]["sky"]["bg"],
                                          "borderRadius":"10px","border":f"1px solid {T['card_border']}",
                                          "borderLeft":f"3px solid {T['pastel']['sky']['fg']}99","padding":"16px","marginBottom":"12px"}),
                dbc.Col([
                    html.Div("CHỈ BÁO ĐỘNG LƯỢNG", style={"fontSize":"0.72rem","letterSpacing":"0.1em","fontWeight":"700",
                                                            "color":T["pastel"]["amber"]["fg"],"marginBottom":"12px"}),
                    html.Table([
                        html.Thead(html.Tr([
                            html.Th("Chỉ báo",  style={"color":T["page_text_dim"],"fontSize":"0.7rem","fontWeight":"700","paddingBottom":"8px","borderBottom":f"1px solid {T['card_border']}"}),
                            html.Th("Giá trị",  style={"color":T["page_text_dim"],"fontSize":"0.7rem","fontWeight":"700","textAlign":"right","paddingBottom":"8px","borderBottom":f"1px solid {T['card_border']}"}),
                            html.Th("Tín hiệu", style={"color":T["page_text_dim"],"fontSize":"0.7rem","fontWeight":"700","textAlign":"right","paddingBottom":"8px","borderBottom":f"1px solid {T['card_border']}"}),
                        ])),
                        html.Tbody([
                            ind_row("RSI (14)",          rsi_val,  sig_rsi,   col_rsi),
                            ind_row("MACD (12,26)",      macd_val, sig_macd,  col_macd),
                            ind_row("Stochastic (14,3)", stoch_k,  sig_stoch, col_stoch),
                        ])
                    ], style={"width":"100%","borderCollapse":"collapse"})
                ], width=12, lg=6, style={"background":T["pastel"]["amber"]["bg"],
                                          "borderRadius":"10px","border":f"1px solid {T['card_border']}",
                                          "borderLeft":f"3px solid {T['pastel']['amber']['fg']}99","padding":"16px","marginBottom":"12px"}),
            ], className="g-3 mb-3"),

            html.Div([
                html.Div("HỖ TRỢ & KHÁNG CỰ", style={"fontSize":"0.72rem","letterSpacing":"0.1em","fontWeight":"700",
                                                        "color":T["pastel"]["rose"]["fg"],"marginBottom":"14px"}),
                dbc.Row([
                    dbc.Col(pivot_card("R3 — Kháng cự 3", r3, T["negative"]), width=4),
                    dbc.Col(pivot_card("R2 — Kháng cự 2", r2, T["negative"]), width=4),
                    dbc.Col(pivot_card("R1 — Kháng cự 1", r1, T["negative"]), width=4),
                ], className="g-2 mb-2"),
                html.Div([
                    html.Div("ĐIỂM XOAY — PIVOT", style={"fontSize":"0.68rem","letterSpacing":"0.12em","color":T["page_text_dim"],
                                                           "fontWeight":"600","marginBottom":"6px"}),
                    html.Div(f"{pp:,.0f}", style={"fontSize":"1.6rem","fontWeight":"900","color":T["page_text"]}),
                ], style={"textAlign":"center","padding":"14px","borderRadius":"8px","margin":"8px 0",
                          "background":T["pastel"]["sky"]["bg"],
                          "border":f"1px solid {T['card_border']}"}),
                dbc.Row([
                    dbc.Col(pivot_card("S1 — Hỗ trợ 1", s1, T["positive"]), width=4),
                    dbc.Col(pivot_card("S2 — Hỗ trợ 2", s2, T["positive"]), width=4),
                    dbc.Col(pivot_card("S3 — Hỗ trợ 3", s3, T["positive"]), width=4),
                ], className="g-2"),
            ], style={"background":T["pastel"]["rose"]["bg"],
                      "borderRadius":"12px","border":f"1px solid {T['card_border']}","padding":"16px 20px"}),
        ], style={"padding": "4px"})

        return technical_content

    except Exception as e:
        logger.error(f"Lỗi Tab Kỹ thuật: {e}")
        return html.Div(f"Lỗi tải dữ liệu kỹ thuật: {str(e)}", className="text-danger p-4")
    
    return technical_content


# ============================================================================
# CALLBACK 2D: LOAD TAB TÀI CHÍNH — CHỈ KHI CLICK VÀO TAB
# ============================================================================
@app.callback(
    Output("fin-table-is", "rowData"),   Output("fin-table-is", "columnDefs"),
    Output("fin-table-bs", "rowData"),   Output("fin-table-bs", "columnDefs"),
    Output("fin-table-cf", "rowData"),   Output("fin-table-cf", "columnDefs"),
    Output("tab-financial-kpi-strip", "children"),
    Input("detail-tabs",      "active_tab"),
    Input("fin-period-toggle", "value"),
    State("selected-stock-store", "data"),
    State("theme-store", "data"),
    prevent_initial_call=True,
)
def load_financial_tab(active_tab, period_toggle, stock, theme="dark"):
    if active_tab != "tab-financial" or not stock:
        return [], [], [], [], [], [], []
    theme = theme or "dark"
    from src.utils.kpi_theme import kpi_card as kpi_card_pastel
    ticker = stock.get('Ticker', 'N/A')

    try:
        df_fin   = load_financial_data(period_toggle)   # ← dùng bản có cache
        df_stock = df_fin[df_fin['Ticker'] == ticker].copy()
        if not df_stock.empty:
            df_stock['Date'] = pd.to_datetime(df_stock['Date'])
            df_stock = df_stock.sort_values("Date", ascending=False)
            if period_toggle == "yearly":
                df_stock['Period'] = df_stock['Date'].dt.year.astype(str)
            else:
                df_stock['Period'] = (
                    df_stock['Date'].dt.year.astype(str) + "-Q" +
                    df_stock['Date'].dt.quarter.astype(str)
                )
            raw_cols_to_keep = [col for col in FINANCIAL_UI_MAP.keys() if col in df_stock.columns]
            df_stock = df_stock[['Period'] + raw_cols_to_keep]
            df_stock.set_index('Period', inplace=True)
            df_t = df_stock.T.reset_index()
            df_t.rename(columns={'index': 'RawItem'}, inplace=True)
            df_t['Chỉ tiêu']  = df_t['RawItem'].apply(lambda x: FINANCIAL_UI_MAP[x]['name'] if x in FINANCIAL_UI_MAP else x)
            df_t['Nhóm BCTC'] = df_t['RawItem'].apply(lambda x: FINANCIAL_UI_MAP[x]['group'] if x in FINANCIAL_UI_MAP else "Khác")
            period_cols = [c for c in df_t.columns if c not in ['Chỉ tiêu', 'RawItem', 'Nhóm BCTC']]

            def create_col_defs(period_columns):
                col_defs = [{"field": "Chỉ tiêu", "pinned": "left", "width": 280,
                             "cellStyle": {"function": """
                                var __theme = document.documentElement.getAttribute('data-theme') || 'dark';
                                return __theme === 'light'
                                    ? {'fontWeight': 'bold', 'color': '#1e293b', 'backgroundColor': '#f8fafc'}
                                    : {'fontWeight': 'bold', 'color': '#e6edf3', 'backgroundColor': '#0d1b2a'};
                             """}}]
                for p in period_columns:
                    col_defs.append({"field": p, "headerName": p, "type": "rightAligned", "width": 120,
                        "valueFormatter": {"function": "params.value !== '' && params.value !== null ? d3.format(',.0f')(params.value) : '-'"}})
                return col_defs

            def process_sub_table(group_name):
                df_sub = df_t[df_t['Nhóm BCTC'] == group_name].copy()
                if df_sub.empty: return [], []
                df_sub = df_sub[['Chỉ tiêu'] + period_cols]
                for c in period_cols:
                    df_sub[c] = pd.to_numeric(df_sub[c], errors='coerce')
                    df_sub[c] = np.where(
                        df_sub['Chỉ tiêu'].isin(['EPS Cơ bản', 'Cổ tức mỗi cổ phiếu (DPS)']),
                        df_sub[c], df_sub[c] / 1_000_000)
                df_sub.replace({np.nan: None}, inplace=True)
                df_sub['Sort_Order'] = df_sub['Chỉ tiêu'].map({v['name']: i for i, v in enumerate(FINANCIAL_UI_MAP.values())})
                df_sub.sort_values('Sort_Order', inplace=True)
                df_sub.drop('Sort_Order', axis=1, inplace=True)
                return df_sub.to_dict('records'), create_col_defs(period_cols)

            is_row_data, is_col_defs = process_sub_table("1. Kết quả kinh doanh")
            bs_row_data, bs_col_defs = process_sub_table("2. Bảng cân đối kế toán")
            cf_row_data, cf_col_defs = process_sub_table("3. Lưu chuyển tiền tệ")
        else:
            err_msg = [{"field": "Lỗi", "headerName": "Không có dữ liệu BCTC"}]
            is_col_defs = bs_col_defs = cf_col_defs = err_msg
            is_row_data = bs_row_data = cf_row_data = []
    except Exception as e:
        logger.error(f"Lỗi Tab Tài Chính (2D): {e}")
        err_msg = [{"field": "Lỗi", "headerName": f"Lỗi hệ thống: {str(e)}"}]
        is_col_defs = bs_col_defs = cf_col_defs = err_msg
        is_row_data = bs_row_data = cf_row_data = []

    # ── KPI STRIP: 4 chỉ số nổi bật lấy từ kỳ gần nhất (cột đầu = mới nhất) ──
    kpi_strip = []
    try:
        if not df_stock.empty and locals().get('period_cols'):
            latest = period_cols[0]
            def _raw(raw_name):
                try:
                    row = df_t[df_t['RawItem'] == raw_name]
                    if row.empty:
                        return None
                    v = row[latest].iloc[0]
                    return float(v) if v is not None and pd.notna(v) else None
                except Exception:
                    return None
            revenue = _raw('Revenue from Business Activities - Total_x')
            net_income = _raw('Net Income after Minority Interest')
            total_assets = _raw('Total Assets')
            equity = _raw('Common Equity - Total')
            kpi_specs = [
                ("DOANH THU KỲ GẦN NHẤT", revenue, "sky", "fas fa-coins"),
                ("LỢI NHUẬN SAU THUẾ", net_income, "green" if (net_income or 0) >= 0 else "rose", "fas fa-dollar"),
                ("TỔNG TÀI SẢN", total_assets, "purple", "fas fa-building-columns"),
                ("VỐN CHỦ SỞ HỮU", equity, "amber", "fas fa-scale-balanced"),
            ]
            kpi_strip = [
                dbc.Col(kpi_card_pastel(
                    theme, label, f"{val/1_000_000:,.0f} tr" if val is not None else "N/A",
                    tone=tone, icon_class=icon,
                    sub_left=("KỲ", latest),
                ), width=6, lg=3) for label, val, tone, icon in kpi_specs
            ]
            kpi_strip = [dbc.Row(kpi_strip, className="g-2")]
    except Exception as e:
        logger.warning(f"KPI strip Tài chính lỗi: {e}")
        kpi_strip = []

    return is_row_data, is_col_defs, bs_row_data, bs_col_defs, cf_row_data, cf_col_defs, kpi_strip

@app.callback(
    [Output("metric-table-1", "rowData"), Output("metric-table-1", "columnDefs"),
     Output("metric-table-2", "rowData"), Output("metric-table-2", "columnDefs"),
     Output("metric-table-3", "rowData"), Output("metric-table-3", "columnDefs"),
     Output("metric-table-4", "rowData"), Output("metric-table-4", "columnDefs"),
     Output("metric-table-5", "rowData"), Output("metric-table-5", "columnDefs"),
     Output("metric-table-6", "rowData"), Output("metric-table-6", "columnDefs"),
     Output("tab-metrics-kpi-strip", "children")],
    [Input("screener-table", "selectedRows"),
     Input("metrics-period-toggle", "value"),
     Input("selected-stock-store", "data")],   # ← THÊM
    State("theme-store", "data"),
    prevent_initial_call=True
)
def update_metrics_tab(selected_rows, period, stock_store_data, theme="dark"):
    theme = theme or "dark"
    from src.utils.kpi_theme import kpi_card as kpi_card_pastel
    # ── Fallback: HF lag → selectedRows chưa set kịp ──
    if (not selected_rows or len(selected_rows) == 0) and stock_store_data:
        selected_rows = [stock_store_data]
    if not selected_rows:
        return ([], []) * 6 + ([],)
    ticker = selected_rows[0].get("Ticker")

    # Khởi tạo 6 cặp giá trị rỗng
    results = [([], []) for _ in range(6)]

    try:
        df = load_financial_data(period)
        df_stock = df[df['Ticker'] == ticker].copy()

        if df_stock.empty:
            return results[0][0], results[0][1], results[1][0], results[1][1], \
                   results[2][0], results[2][1], results[3][0], results[3][1], \
                   results[4][0], results[4][1], results[5][0], results[5][1], []

        df_stock['Date'] = pd.to_datetime(df_stock['Date'])
        df_stock = df_stock.sort_values("Date", ascending=False)  # Ngày mới nhất lên đầu

        # Tạo cột Thời gian (Header)
        if period == "yearly":
            df_stock['Period'] = df_stock['Date'].dt.year.astype(str)
        else:
            df_stock['Period'] = df_stock['Date'].dt.year.astype(str) + "-Q" + df_stock['Date'].dt.quarter.astype(str)

        # =================================================================
        # 🟢 TÍNH TOÁN CÁC CHỈ SỐ TÀI CHÍNH TỪ RAW DATA
        # =================================================================
        # 1. Per Share
        df_stock['EPS'] = df_stock['EPS - Basic - excl Extraordinary Items, Common - Total']
        df_stock['BVPS'] = df_stock['Common Equity - Total'] / df_stock['Common Shares - Outstanding - Total_x']

        # 2. Sinh lời
        df_stock['ROE'] = (df_stock['Net Income after Minority Interest'] / df_stock['Common Equity - Total']) * 100
        df_stock['ROA'] = (df_stock['Net Income after Minority Interest'] / df_stock['Total Assets']) * 100
        df_stock['Gross Margin'] = (df_stock['Gross Profit - Industrials/Property - Total'] / df_stock[
            'Revenue from Business Activities - Total_x']) * 100
        df_stock['Net Margin'] = (df_stock['Net Income after Minority Interest'] / df_stock[
            'Revenue from Business Activities - Total_x']) * 100
        df_stock['EBIT Margin'] = (df_stock['Earnings before Interest & Taxes (EBIT)'] / df_stock[
            'Revenue from Business Activities - Total_x']) * 100

        # 3. Thanh khoản
        df_stock['Current Ratio'] = df_stock['Total Current Assets'] / df_stock['Total Current Liabilities']

        cash_equiv = df_stock['Cash & Cash Equivalents - Total_x'].fillna(0)
        short_invest = df_stock['Short-Term Investments - Total'].fillna(0)
        receivables = df_stock['Trade Accounts & Trade Notes Receivable - Net'].fillna(0)

        df_stock['Quick Ratio'] = (cash_equiv + short_invest + receivables) / df_stock['Total Current Liabilities']
        df_stock['Cash Ratio'] = (cash_equiv + short_invest) / df_stock['Total Current Liabilities']

        # 4. Đòn bẩy
        total_debt = df_stock['Short-Term Debt & Current Portion of Long-Term Debt'].fillna(0) + df_stock[
            'Debt - Long-Term - Total'].fillna(0)
        df_stock['Debt to Equity'] = total_debt / df_stock['Common Equity - Total']
        df_stock['Debt to Assets'] = total_debt / df_stock['Total Assets']
        df_stock['Equity Multiplier'] = df_stock['Total Assets'] / df_stock['Common Equity - Total']

        # 5. Hiệu quả
        df_stock['Asset Turnover'] = df_stock['Revenue from Business Activities - Total_x'] / df_stock['Total Assets']
        # Cost of Revenues thường ghi âm, cần lấy Trị tuyệt đối
        df_stock['Inventory Turnover'] = abs(df_stock['Cost of Revenues - Total']) / df_stock[
            'Inventories - Total'].replace(0, np.nan)

        # 6. Tăng trưởng (Dùng shift(-1) vì data đang sort giảm dần theo Date)
        df_stock['Revenue Growth'] = (df_stock['Revenue from Business Activities - Total_x'] / df_stock[
            'Revenue from Business Activities - Total_x'].shift(-1) - 1) * 100
        df_stock['Net Income Growth'] = (df_stock['Net Income after Minority Interest'] / abs(
            df_stock['Net Income after Minority Interest'].shift(-1)) - 1) * 100

        # Thay thế vô cực bằng NaN
        df_stock.replace([float('inf'), float('-inf')], None, inplace=True)

        # ── KPI STRIP: 4 chỉ số nổi bật ở kỳ gần nhất (đã sort mới→cũ, dòng đầu = mới nhất) ──
        kpi_strip = []
        try:
            latest_row = df_stock.iloc[0]
            latest_period = latest_row['Period']
            def _safe(col):
                v = latest_row.get(col)
                return float(v) if v is not None and pd.notna(v) else None
            roe_v   = _safe('ROE')
            roa_v   = _safe('ROA')
            gm_v    = _safe('Gross Margin')
            de_v    = _safe('Debt to Equity')
            kpi_specs = [
                ("ROE", f"{roe_v:.1f}%" if roe_v is not None else "N/A", "green" if (roe_v or 0) >= 0 else "rose", "fas fa-arrow-trend-up"),
                ("ROA", f"{roa_v:.1f}%" if roa_v is not None else "N/A", "sky", "fas fa-percent"),
                ("BIÊN LỢI NHUẬN GỘP", f"{gm_v:.1f}%" if gm_v is not None else "N/A", "purple", "fas fa-layer-group"),
                ("NỢ VAY / VCSH", f"{de_v:.2f}x" if de_v is not None else "N/A", "amber", "fas fa-scale-balanced"),
            ]
            kpi_strip = [dbc.Row([
                dbc.Col(kpi_card_pastel(theme, label, value, tone=tone, icon_class=icon,
                                        sub_left=("KỲ", latest_period)), width=6, lg=3)
                for label, value, tone, icon in kpi_specs
            ], className="g-2")]
        except Exception as e:
            logger.warning(f"KPI strip Chỉ số lỗi: {e}")
            kpi_strip = []

        # =================================================================
        # 🟢 XOAY BẢNG (TRANSPOSE) VÀ CHIA GROUP
        # =================================================================
        cols_to_keep = [col for col in METRICS_UI_MAP.keys() if col in df_stock.columns]
        df_stock = df_stock[['Period'] + cols_to_keep].set_index('Period')

        df_t = df_stock.T.reset_index()
        df_t.rename(columns={'index': 'RawItem'}, inplace=True)

        df_t['Chỉ tiêu'] = df_t['RawItem'].apply(lambda x: METRICS_UI_MAP[x]['name'])
        df_t['Group'] = df_t['RawItem'].apply(lambda x: METRICS_UI_MAP[x]['group'])

        period_cols = [c for c in df_t.columns if c not in ['Chỉ tiêu', 'RawItem', 'Group']]

        # Hàm tạo Cấu trúc cột AG Grid (Format 2 chữ số thập phân, KHÔNG chia 1 triệu)
        def create_metric_col_defs(periods):
            defs = [{"field": "Chỉ tiêu", "pinned": "left", "width": 280,
                     "cellStyle": {"function": """
                        var __theme = document.documentElement.getAttribute('data-theme') || 'dark';
                        return __theme === 'light'
                            ? {'fontWeight': 'bold', 'color': '#1e293b', 'backgroundColor': '#f8fafc'}
                            : {'fontWeight': 'bold', 'color': '#e6edf3', 'backgroundColor': '#0d1b2a'};
                     """}}]
            for p in periods:
                defs.append({
                    "field": p, "headerName": p, "type": "rightAligned", "width": 120,
                    # Format: 12.34
                    "valueFormatter": {
                        "function": "params.value !== '' && params.value !== null ? d3.format(',.2f')(params.value) : '-'"}
                })
            return defs

        col_defs_template = create_metric_col_defs(period_cols)

        # Đẩy dữ liệu vào 6 bảng
        final_returns = []
        for i in range(1, 7):
            df_sub = df_t[df_t['Group'] == str(i)].copy()
            if not df_sub.empty:
                df_sub = df_sub[['Chỉ tiêu'] + period_cols]
                # Sort theo thứ tự trong từ điển
                df_sub['Sort'] = df_sub['Chỉ tiêu'].map(
                    {v['name']: idx for idx, v in enumerate(METRICS_UI_MAP.values())})
                df_sub.sort_values('Sort', inplace=True)
                df_sub.drop('Sort', axis=1, inplace=True)

                # Ép kiểu float và thay NaN bằng chuỗi rỗng để AG Grid hiện dấu "-"
                for c in period_cols:
                    df_sub[c] = pd.to_numeric(df_sub[c], errors='coerce')
                df_sub.replace({np.nan: None}, inplace=True)

                final_returns.extend([df_sub.to_dict('records'), col_defs_template])
            else:
                final_returns.extend([[], [{"field": "Chỉ tiêu", "headerName": "Không đủ dữ liệu"}]])

        return tuple(final_returns) + (kpi_strip,)

    except Exception as e:
        import traceback
        traceback.print_exc()
        err = [{"field": "Chỉ tiêu", "headerName": f"Lỗi: {e}"}]
        return ([], err) * 6 + ([],)


# ============================================================================
# COLLAPSIBLE SECTIONS CALLBACKS
# ============================================================================
# NOTE: Tất cả collapse group callbacks đã được chuyển sang
# filter_interaction_callbacks.py (callback toggle_all_collapses)
# để tránh duplicate output conflict.

# ============================================================================
# CALLBACK: ĐÓNG / MỞ BỘ LỌC OFFCANVAS
# ============================================================================
@app.callback(
    Output("filter-offcanvas", "is_open"),
    [Input("toggle-filter-btn", "n_clicks"),
     Input("btn-filter", "n_clicks"),
     Input("strategy-preset-dropdown", "value"),
     
     # Input("selected-filters-container", "children"),   # ← THÊM # COMMENTED OUT  to FIX Panel bộ lọc tự mở sau onboarding
    ],
    [State("filter-offcanvas", "is_open")],
    prevent_initial_call=True
)
def toggle_filter_offcanvas(n_clicks_open, n_clicks_apply, strategy_val,
                             # filter_children,   # Xóa tham số filter_children ở hàm def toggle_filter_offcanvas().
                             is_open):
    from dash import ctx as dash_ctx
    triggered_id = dash_ctx.triggered_id

    if triggered_id == "strategy-preset-dropdown":
        return True if strategy_val else is_open
    # elif triggered_id == "selected-filters-container":
    #   Xóa đoạn elif triggered_id == "selected-filters-container": ở bên dưới.
    elif triggered_id == "toggle-filter-btn":
        return not is_open
    elif triggered_id == "btn-filter":
        return False

    return is_open

# ============================================================================
# CALLBACK: HIỂN THỊ THÔNG TIN TRƯỜNG PHÁI ĐẦU TƯ (INFO OFFCANVAS)
# ============================================================================
@app.callback(
    [Output("strategy-info-offcanvas", "is_open"),
     Output("strategy-info-offcanvas", "title"),
     Output("strategy-info-offcanvas", "children")],
    [Input("btn-strategy-info", "n_clicks")],  # 🔴 XÓA CÁI INPUT SEARCH Ở ĐÂY ĐI NHÉ
    [State("strategy-preset-dropdown", "value"),
     State("strategy-info-offcanvas", "is_open")],
    prevent_initial_call=True
)
def toggle_strategy_info(n_clicks, current_strategy, is_open):
    if not n_clicks:
        return is_open, "", ""

    if not current_strategy:
        return True, "⚠️ Vui lòng chọn trường phái", html.P(
            "Bạn cần chọn một trường phái trong Dropdown trước khi xem thông tin chi tiết.", style={"color": "#ff7b72"})

    # ==========================================================
    # 🛠️ HELPER FORMAT SỐ (CẮT BỎ .0 NẾU LÀ SỐ NGUYÊN)
    # ==========================================================
    def fmt(val):
        """Nếu số là 20.0 -> trả về int(20). Nếu là 2.5 -> giữ nguyên float(2.5)"""
        try:
            if float(val).is_integer():
                return int(val)
            return val
        except Exception as _e:  # noqa: audit-fix bare-except
            logger.debug(f"Suppressed non-critical error in {__name__} near line 2617: {_e}")
            return val

    # ==========================================================
    # 📚 TỪ ĐIỂN TRIẾT LÝ & THÔNG SỐ
    # ==========================================================
    if current_strategy == "STRAT_VALUE":
        title = "📊 Đầu tư giá trị (Value Investing)"

        # Đọc trực tiếp biến số từ hệ thống Quant để đảm bảo luôn đồng bộ
        pe_max = VALUE_THRESHOLDS[VALUE_IDX_PE_MAX]
        pb_max = VALUE_THRESHOLDS[VALUE_IDX_PB_MAX]
        cr_min = VALUE_THRESHOLDS[VALUE_IDX_CURRENT_RATIO_MIN]
        debt_wc_max = VALUE_THRESHOLDS[VALUE_IDX_DEBT_TO_WC_MAX]

        content = html.Div([
            # Phần 1: Tác giả & Tham khảo
            html.Div([
                html.Span(dbc.Badge("Tác giả", color="primary", className="me-2")),
                "Benjamin Graham & Warren Buffett"
            ], className="mb-2"),
            html.Div([
                html.Span(dbc.Badge("Nguồn", color="info", className="me-2")),
                html.A("The Intelligent Investor & Website tham khảo", href="https://tranthinhlam.com/dau-tu-gia-tri/",
                       target="_blank", style={"color": "#58a6ff", "textDecoration": "none"})
            ], className="mb-4"),

            # Phần 2: Triết lý
            html.H5([html.I(className="fas fa-brain", style={"color": "#3fb950", "marginRight": "8px"}),
                     "Triết lý cốt lõi"], style={"color": "#e6edf3", "fontWeight": "bold"}),
            html.P(
                "Tập trung vào việc xác định giá trị nội tại của doanh nghiệp thông qua phân tích BCTC và mua cổ phiếu khi giá thị trường thấp hơn giá trị thực (margin of safety). Nhấn mạnh phân tích định lượng dài hạn thay vì đầu cơ ngắn hạn.",
                style={"fontSize": "0.95rem", "lineHeight": "1.6", "color": "#c9d1d9"}),
            html.Hr(style={"borderColor": "#30363d"}),

            # Phần 3: Logic Lọc (Lấy số từ hệ thống Quant)
            html.H5([html.I(className="fas fa-filter me-2", style={"color": "#f85149"}), "Logic Sàng Lọc (Hệ thống)"],
                    style={"color": "#e6edf3", "fontWeight": "bold"}),
            html.Ul([
                # 🟢 BAO BỌC BIẾN BẰNG HÀM fmt() 🟢
                html.Li([html.Strong("P/E (Price/Earnings): "), f"Nhỏ hơn hoặc bằng {fmt(pe_max)}"]),
                html.Li([html.Strong("P/B (Price/Book): "), f"Nhỏ hơn hoặc bằng {fmt(pb_max)}"]),
                html.Li([html.Strong("Chỉ số Graham (P/E * P/B): "), f"Không vượt quá {fmt(pe_max * pb_max)}"]),
                html.Li([html.Strong("Tài chính an toàn: "), f"Current Ratio >= {fmt(cr_min)}"]),
                html.Li([html.Strong("Tỷ lệ Nợ/Vốn lưu động: "), f"Nhỏ hơn {fmt(debt_wc_max)}"]),
                html.Li([html.Strong("Lợi nhuận dương: "), "Không lỗ trong 5 năm gần nhất"]),
            ], style={"fontSize": "0.95rem", "lineHeight": "1.8", "color": "#c9d1d9", "backgroundColor": "#0d1117",
                      "padding": "15px 15px 15px 35px", "borderRadius": "8px"}),
            html.Hr(style={"borderColor": "#30363d"}),

            # Phần 4: Mô tả UI
            html.H5([html.I(className="fas fa-desktop me-2", style={"color": "#1E88E5"}), "Tính năng mở rộng"],
                    style={"color": "#e6edf3", "fontWeight": "bold"}),
            html.P(
                "Cột Value Score trong bảng sẽ chấm điểm A, B, C, D dựa trên việc mã đó thỏa mãn được bao nhiêu tiêu chí của Graham.",
                style={"fontSize": "0.9rem", "fontStyle": "italic", "color": "#c9d1d9"})
        ])
    # ==========================================================
    # 🔄 CHIẾN LƯỢC PHỤC HỒI (TURNAROUND)
    # ==========================================================
    elif current_strategy == "STRAT_TURNAROUND":
        title = "🔄 Đầu tư phục hồi (Turnaround Investing)"

        # Đọc trực tiếp biến số từ hệ thống Quant
        pe_hist = TURNAROUND_THRESHOLDS[TURNAROUND_IDX_PE_HIST_NORM_MAX]
        op_margin = TURNAROUND_THRESHOLDS[TURNAROUND_IDX_OPERATING_MARGIN_MIN]
        peg_min = TURNAROUND_THRESHOLDS[TURNAROUND_IDX_PEG_MIN]
        peg_max = TURNAROUND_THRESHOLDS[TURNAROUND_IDX_PEG_MAX]

        content = html.Div([
            html.Div([
                html.Span(dbc.Badge("Tác giả", color="primary", className="me-2")),
                "Sir John Templeton"
            ], className="mb-2"),
            html.Div([
                html.Span(dbc.Badge("Nguồn", color="info", className="me-2")),
                html.A("Sách: Investing the Templeton Way",
                       href="https://www.templeton.org/articles/principles-turnaround-investing", target="_blank",
                       style={"color": "#58a6ff", "textDecoration": "none"})
            ], className="mb-4"),

            html.H5([html.I(className="fas fa-brain", style={"color": "#3fb950", "marginRight": "8px"}),
                     "Triết lý cốt lõi"], style={"color": "#e6edf3", "fontWeight": "bold"}),
            html.P(
                "Tìm kiếm các doanh nghiệp đang gặp khó khăn tạm thời (lợi nhuận giảm, hoạt động suy yếu) nhưng có khả năng phục hồi. Mua cổ phiếu ở mức giá cực rẻ khi thị trường bi quan cùng cực, trước khi sự phục hồi được phản ánh vào giá.",
                style={"fontSize": "0.95rem", "lineHeight": "1.6", "color": "#c9d1d9"}),
            html.Hr(style={"borderColor": "#30363d"}),

            html.H5([html.I(className="fas fa-filter", style={"color": "#f85149", "marginRight": "8px"}),
                     "Logic Sàng Lọc (Hệ thống)"], style={"color": "#e6edf3", "fontWeight": "bold"}),
            html.Ul([
                html.Li([html.Strong("Định giá hoảng loạn: "),
                         f"P/E hiện tại thấp hơn {fmt(pe_hist * 100)}% so với trung bình lịch sử"]),
                html.Li([html.Strong("Dấu hiệu hồi sinh: "),
                         f"Biên lợi nhuận HĐKD (Operating Margin) > {fmt(op_margin)}%"]),
                html.Li([html.Strong("Giá trên đà phục hồi: "),
                         f"Chỉ số PEG nằm trong khoảng an toàn ({fmt(peg_min)} - {fmt(peg_max)})"]),
            ], style={"fontSize": "0.95rem", "lineHeight": "1.8", "color": "#c9d1d9", "backgroundColor": "#0d1117",
                      "padding": "15px 15px 15px 35px", "borderRadius": "8px"}),
            html.Hr(style={"borderColor": "#30363d"}),

            html.H5([html.I(className="fas fa-desktop", style={"color": "#1E88E5", "marginRight": "8px"}),
                     "Tính năng mở rộng"], style={"color": "#e6edf3", "fontWeight": "bold"}),
            html.P(
                "Biểu đồ Performance ở Tab Kỹ thuật giúp bạn theo dõi đà giảm sâu trong 1 năm qua (Global Pessimism) và nhịp phục hồi (Recovery Tracker) trong 3 tháng gần nhất.",
                style={"fontSize": "0.9rem", "fontStyle": "italic", "color": "#c9d1d9"})
        ])

    # ==========================================================
    # 💎 CHIẾN LƯỢC CHẤT LƯỢNG (QUALITY)
    # ==========================================================
    elif current_strategy == "STRAT_QUALITY":
        title = "💎 Đầu tư chất lượng (Quality Investing)"

        # Đọc trực tiếp biến số từ hệ thống Quant
        roe_min = QUALITY_THRESHOLDS[QUALITY_IDX_ROE_MIN]
        gm_min = QUALITY_THRESHOLDS[QUALITY_IDX_GROSS_MARGIN_MIN]
        re_growth = QUALITY_THRESHOLDS[QUALITY_IDX_RE_GROWTH_MIN]
        fcf_margin = QUALITY_THRESHOLDS[QUALITY_IDX_FCF_MARGIN_MIN]

        content = html.Div([
            html.Div([
                html.Span(dbc.Badge("Tác giả", color="primary", className="me-2")),
                "Charlie Munger & Terry Smith"
            ], className="mb-2"),
            html.Div([
                html.Span(dbc.Badge("Nguồn", color="info", className="me-2")),
                html.A("Sách: Poor Charlie’s Almanack",
                       href="https://www.google.com/search?q=Poor+Charlie%E2%80%99s+Almanack", target="_blank",
                       style={"color": "#58a6ff", "textDecoration": "none"})
            ], className="mb-4"),

            html.H5([html.I(className="fas fa-brain", style={"color": "#3fb950", "marginRight": "8px"}),
                     "Triết lý cốt lõi"], style={"color": "#e6edf3", "fontWeight": "bold"}),
            html.P(
                "Sở hữu những doanh nghiệp vĩ đại có lợi thế cạnh tranh độc quyền (Economic Moat) và khả năng tạo ra lãi kép vượt trội. Chú trọng năng lực tái đầu tư lợi nhuận và dòng tiền mặt thực thu dồi dào, thay vì chỉ mua cổ phiếu giá rẻ.",
                style={"fontSize": "0.95rem", "lineHeight": "1.6", "color": "#c9d1d9"}),
            html.Hr(style={"borderColor": "#30363d"}),

            html.H5([html.I(className="fas fa-filter", style={"color": "#f85149", "marginRight": "8px"}),
                     "Logic Sàng Lọc (Hệ thống)"], style={"color": "#e6edf3", "fontWeight": "bold"}),
            html.Ul([
                html.Li([
                    html.Strong("Hiệu suất vốn vượt trội: "), f"ROE >= {fmt(roe_min)}%",
                    html.Div(f"Đảm bảo tạo ra tối thiểu {fmt(roe_min / 100)} đồng lợi nhuận trên mỗi đồng vốn.",
                             style={"fontSize": "0.85rem", "fontStyle": "italic", "color": "#8b949e",
                                    "marginTop": "2px"})
                ], style={"marginBottom": "8px"}),

                html.Li([
                    html.Strong("Con hào kinh tế: "), f"Biên lãi gộp >= {fmt(gm_min)}%",
                    html.Div("Xác lập ngưỡng lãi gộp tối thiểu để khẳng định lợi thế cạnh tranh bền vững.",
                             style={"fontSize": "0.85rem", "fontStyle": "italic", "color": "#8b949e",
                                    "marginTop": "2px"})
                ], style={"marginBottom": "8px"}),

                html.Li([
                    html.Strong("Sức mạnh lãi kép: "), f"Lợi nhuận giữ lại tăng >= {fmt(re_growth)}%",
                    html.Div("Yêu cầu lợi nhuận giữ lại (Retained Earnings) tăng trưởng dương để tiếp tục tạo lãi kép.",
                             style={"fontSize": "0.85rem", "fontStyle": "italic", "color": "#8b949e",
                                    "marginTop": "2px"})
                ], style={"marginBottom": "8px"}),

                html.Li([
                    html.Strong("Tiền mặt là vua: "), f"FCF Margin >= {fmt(fcf_margin)} Lần",
                    html.Div(
                        f"Xác thực {fmt(fcf_margin * 100)}% lợi nhuận ròng phải là tiền mặt thực thu (tiền chảy vào túi cổ đông).",
                        style={"fontSize": "0.85rem", "fontStyle": "italic", "color": "#8b949e", "marginTop": "2px"})
                ], style={"marginBottom": "8px"}),

                html.Li([
                    html.Strong("Quy mô an toàn: "), "Vốn hóa (Market Cap) > Mức trung bình",
                    html.Div("Loại bỏ các doanh nghiệp có quy mô vốn hóa dưới mức trung bình của thị trường (Indo).",
                             style={"fontSize": "0.85rem", "fontStyle": "italic", "color": "#8b949e",
                                    "marginTop": "2px"})
                ], style={"marginBottom": "8px"}),

                html.Li([
                    html.Strong("Hiệu suất tài sản: "), "Vòng quay PPE (Δ) > 0",
                    html.Div("Yêu cầu hiệu suất khai thác tài sản (PPE Turnover) năm sau phải cao hơn năm trước.",
                             style={"fontSize": "0.85rem", "fontStyle": "italic", "color": "#8b949e",
                                    "marginTop": "2px"})
                ], style={"marginBottom": "8px"}),

            ], style={"fontSize": "0.95rem", "lineHeight": "1.4", "color": "#c9d1d9", "backgroundColor": "#0d1117",
                      "padding": "15px 15px 15px 35px", "borderRadius": "8px"}),
        ])
    # ==========================================================
    # ⚖️ CHIẾN LƯỢC GARP (TĂNG TRƯỞNG VỚI GIÁ HỢP LÝ)
    # ==========================================================
    elif current_strategy == "STRAT_GARP":
        title = "⚖️ Tăng trưởng giá hợp lý (GARP)"

        # Đọc trực tiếp biến số từ hệ thống Quant
        eps_min = GARP_THRESHOLDS[GARP_IDX_EPS_GROWTH_MIN]
        eps_max = GARP_THRESHOLDS[GARP_IDX_EPS_GROWTH_MAX]
        pe_max = GARP_THRESHOLDS[GARP_IDX_PE_MAX]
        peg_min = GARP_THRESHOLDS[GARP_IDX_PEG_MIN]
        peg_max = GARP_THRESHOLDS[GARP_IDX_PEG_MAX]
        de_max = GARP_THRESHOLDS[GARP_IDX_D_E_MAX]
        sgr_min = GARP_THRESHOLDS[GARP_IDX_SGR_MIN_PCT]
        mc_quantile = GARP_THRESHOLDS[GARP_IDX_MC_QUANTILE]

        content = html.Div([
            html.Div([
                html.Span(dbc.Badge("Tác giả", color="primary", className="me-2")),
                "Peter Lynch"
            ], className="mb-2"),
            html.Div([
                html.Span(dbc.Badge("Nguồn", color="info", className="me-2")),
                html.A("Sách: One Up On Wall Street (FinanceStrategists)",
                       href="https://www.financestrategists.com/wealth-management/investment-management/growth-at-a-reasonable-price-garp/",
                       target="_blank", style={"color": "#58a6ff", "textDecoration": "none"})
            ], className="mb-4"),
            html.H5([html.I(className="fas fa-brain", style={"color": "#3fb950", "marginRight": "8px"}),
                     "Triết lý cốt lõi"], style={"color": "#e6edf3", "fontWeight": "bold"}),
            html.P(
                "Tìm kiếm sự giao thoa giữa tiềm năng tăng trưởng bền vững và mức định giá hợp lý (Growth At a Reasonable Price). Ưu tiên các doanh nghiệp có khả năng tự lớn mạnh từ nội lực tài chính mà không phải đánh đổi bằng rủi ro nợ nần quá mức.",
                style={"fontSize": "0.95rem", "lineHeight": "1.6", "color": "#c9d1d9"}),
            html.Hr(style={"borderColor": "#30363d"}),

            html.H5([html.I(className="fas fa-filter", style={"color": "#f85149", "marginRight": "8px"}),
                     "Logic Sàng Lọc (Hệ thống)"], style={"color": "#e6edf3", "fontWeight": "bold"}),
            html.Ul([
                html.Li([
                    html.Strong("Tăng trưởng bền vững: "), f"EPS 1Y tăng {fmt(eps_min)}% - {fmt(eps_max)}%",
                    html.Div("Lọc dải tăng trưởng ổn định, loại bỏ các mã tăng trưởng 'ảo'.",
                             style={"fontSize": "0.85rem", "fontStyle": "italic", "color": "#8b949e",
                                    "marginTop": "2px"})
                ], style={"marginBottom": "8px"}),

                html.Li([
                    html.Strong("Giá cả hợp lý (PEG): "), f"{fmt(peg_min)} < PEG <= {fmt(peg_max)}",
                    html.Div("Xác thực giá mua tương xứng với tốc độ tăng trưởng EPS.",
                             style={"fontSize": "0.85rem", "fontStyle": "italic", "color": "#8b949e",
                                    "marginTop": "2px"})
                ], style={"marginBottom": "8px"}),

                html.Li([
                    html.Strong("Giới hạn định giá: "), f"P/E <= {fmt(pe_max)}",
                    html.Div("Loại bỏ các mã bị thổi giá quá mức so với thu nhập.",
                             style={"fontSize": "0.85rem", "fontStyle": "italic", "color": "#8b949e",
                                    "marginTop": "2px"})
                ], style={"marginBottom": "8px"}),

                html.Li([
                    html.Strong("Nội lực tự lớn mạnh (SGR): "), f">= {fmt(sgr_min)}%",
                    html.Div("Yêu cầu tốc độ tự tăng trưởng từ nguồn vốn nội bộ.",
                             style={"fontSize": "0.85rem", "fontStyle": "italic", "color": "#8b949e",
                                    "marginTop": "2px"})
                ], style={"marginBottom": "8px"}),

                html.Li([
                    html.Strong("An toàn tài chính: "), f"D/E Ratio <= {fmt(de_max)}",
                    html.Div("Khống chế nợ không được lạm dụng quá vốn chủ sở hữu.",
                             style={"fontSize": "0.85rem", "fontStyle": "italic", "color": "#8b949e",
                                    "marginTop": "2px"})
                ], style={"marginBottom": "8px"}),

                html.Li([
                    html.Strong("Quy mô an toàn: "), f"Top {fmt((1 - mc_quantile) * 100)}% Vốn hóa lớn nhất sàn"
                ], style={"marginBottom": "8px"}),

            ], style={"fontSize": "0.95rem", "lineHeight": "1.4", "color": "#c9d1d9", "backgroundColor": "#0d1117",
                      "padding": "15px 15px 15px 35px", "borderRadius": "8px"}),
        ])

    # ==========================================================
    # 💰 CHIẾN LƯỢC CỔ TỨC & THU NHẬP (DIVIDEND)
    # ==========================================================
    elif current_strategy == "STRAT_DIVIDEND":
        title = "💰 Cổ tức & Thu nhập (Dividend)"

        # Đọc trực tiếp biến số từ hệ thống Quant
        mc_quantile = DIVIDEND_THRESHOLDS[DIV_IDX_MC_QUANTILE]
        yield_min = DIVIDEND_THRESHOLDS[DIV_IDX_YIELD_MIN]
        payout_max = DIVIDEND_THRESHOLDS[DIV_IDX_PAYOUT_MAX]

        content = html.Div([
            html.Div([
                html.Span(dbc.Badge("Tác giả", color="primary", className="me-2")),
                "John Neff"
            ], className="mb-2"),
            html.Div([
                html.Span(dbc.Badge("Nguồn", color="info", className="me-2")),
                html.A("Chiến lược Cổ tức (Vietcap)",
                       href="https://www.vietcap.com.vn/kien-thuc/chien-luoc-dau-tu-co-phieu-huong-co-tuc-hieu-qua",
                       target="_blank", style={"color": "#58a6ff", "textDecoration": "none"})
            ], className="mb-4"),

            html.H5([html.I(className="fas fa-brain", style={"color": "#3fb950", "marginRight": "8px"}),
                     "Triết lý cốt lõi"], style={"color": "#e6edf3", "fontWeight": "bold"}),
            html.P(
                "Tạo dựng dòng thu nhập thụ động ổn định thông qua cổ tức từ các doanh nghiệp tài chính vững mạnh, có lịch sử chi trả đều đặn. Thay vì theo đuổi sự tăng giá ngắn hạn đầy biến động, chiến lược này ưu tiên tổng lợi suất dài hạn giúp chống chịu tốt trước lạm phát.",
                style={"fontSize": "0.95rem", "lineHeight": "1.6", "color": "#c9d1d9"}),
            html.Hr(style={"borderColor": "#30363d"}),

            html.H5([html.I(className="fas fa-filter", style={"color": "#f85149", "marginRight": "8px"}),
                     "Logic Sàng Lọc (Hệ thống)"], style={"color": "#e6edf3", "fontWeight": "bold"}),
            html.Ul([
                html.Li([
                    html.Strong("Lợi suất hấp dẫn: "), f"Dividend Yield >= {fmt(yield_min * 100)}%",
                    html.Div("Đảm bảo lợi tức cao hơn lãi suất tiết kiệm ngân hàng (phi rủi ro).",
                             style={"fontSize": "0.85rem", "fontStyle": "italic", "color": "#8b949e",
                                    "marginTop": "2px"})
                ], style={"marginBottom": "8px"}),

                html.Li([
                    html.Strong("Sự an toàn của cổ tức: "), f"Payout Ratio <= {fmt(payout_max * 100)}%",
                    html.Div("Ngăn chặn các doanh nghiệp chia hết lợi nhuận hoặc vay nợ để trả cổ tức.",
                             style={"fontSize": "0.85rem", "fontStyle": "italic", "color": "#8b949e",
                                    "marginTop": "2px"})
                ], style={"marginBottom": "8px"}),

                html.Li([
                    html.Strong("Sự ổn định: "), "Duy trì trả cổ tức đều đặn trong 3 năm liên tiếp",
                ], style={"marginBottom": "8px"}),

                html.Li([
                    html.Strong("Sức khỏe tài chính: "), "Dòng tiền tự do (FCF) dương & Nợ vay thấp",
                    html.Div("Chỉ có doanh nghiệp tạo ra tiền mặt thực sự mới duy trì được cổ tức.",
                             style={"fontSize": "0.85rem", "fontStyle": "italic", "color": "#8b949e",
                                    "marginTop": "2px"})
                ], style={"marginBottom": "8px"}),

                html.Li([
                    html.Strong("Quy mô: "), f"Top {fmt((1 - mc_quantile) * 100)}% Vốn hóa thị trường",
                    html.Div("Ưu tiên các tập đoàn lớn, tránh rủi ro thanh khoản.",
                             style={"fontSize": "0.85rem", "fontStyle": "italic", "color": "#8b949e",
                                    "marginTop": "2px"})
                ], style={"marginBottom": "8px"}),

            ], style={"fontSize": "0.95rem", "lineHeight": "1.4", "color": "#c9d1d9", "backgroundColor": "#0d1117",
                      "padding": "15px 15px 15px 35px", "borderRadius": "8px"}),
        ])
        # ==========================================================
    # 📈 CHIẾN LƯỢC PIOTROSKI (SỨC KHỎE TÀI CHÍNH)
    # ==========================================================
    elif current_strategy == "STRAT_PIOTROSKI":
        title = "📈 Điểm sức khỏe Piotroski (F-Score)"

        # Đọc trực tiếp biến số từ hệ thống Quant
        f_min = PIOTROSKI_THRESHOLDS[PIOTROSKI_IDX_F_MIN]

        content = html.Div([
            html.Div([
                html.Span(dbc.Badge("Tác giả", color="primary", className="me-2")),
                "Giáo sư Joseph Piotroski (ĐH Stanford)"
            ], className="mb-2"),
            html.Div([
                html.Span(dbc.Badge("Nguồn", color="info", className="me-2")),
                html.A("Tikop",
                       href="https://tikop.vn/blog/piotroski-f-score-la-gi-cach-tinh-y-nghia-va-ung-dung-trong-dau-tu-10991",
                       target="_blank", style={"color": "#58a6ff", "textDecoration": "none"}),
                html.Span(" | ", style={"color": "#c9d1d9"}),
                html.A("GoValue", href="https://govalue.vn/piotroski-f-score/", target="_blank",
                       style={"color": "#58a6ff", "textDecoration": "none"})
            ], className="mb-4"),

            html.H5([html.I(className="fas fa-brain", style={"color": "#3fb950", "marginRight": "8px"}),
                     "Triết lý cốt lõi"], style={"color": "#e6edf3", "fontWeight": "bold"}),
            html.P(
                "Mô hình chấm điểm dựa trên 9 tiêu chí tài chính giúp phân biệt doanh nghiệp đang thực sự phục hồi với những 'cái bẫy giá trị' (doanh nghiệp yếu kém bị định giá thấp). Điểm 8-9 cho thấy sức khỏe tài chính rất tốt, trong khi 0-4 cảnh báo rủi ro cao.",
                style={"fontSize": "0.95rem", "lineHeight": "1.6", "color": "#c9d1d9"}),
            html.Hr(style={"borderColor": "#30363d"}),

            html.H5([html.I(className="fas fa-filter", style={"color": "#f85149", "marginRight": "8px"}),
                     "Logic Sàng Lọc (Hệ thống)"], style={"color": "#e6edf3", "fontWeight": "bold"}),
            html.Ul([
                html.Li([
                    html.Strong("Ngưỡng an toàn: "), f"Tổng điểm F-Score >= {fmt(f_min)}/9 điểm",
                    html.Div("Hệ thống chỉ lọc ra các doanh nghiệp có tình hình tài chính từ mức Khá đến Xuất sắc.",
                             style={"fontSize": "0.85rem", "fontStyle": "italic", "color": "#8b949e",
                                    "marginTop": "2px"})
                ], style={"marginBottom": "12px"}),

                html.Li([html.Strong(html.Span("1. Khả năng sinh lợi:", style={"color": "#58a6ff"}))]),
                html.Div(
                    "Lợi nhuận ròng > 0; ROA > 0 và cao hơn năm trước; CFO (Dòng tiền HĐKD) > 0; CFO > Lợi nhuận ròng.",
                    style={"fontSize": "0.85rem", "color": "#8b949e", "marginLeft": "20px", "marginBottom": "8px"}),

                html.Li([html.Strong(html.Span("2. Sức khỏe tài chính:", style={"color": "#58a6ff"}))]),
                html.Div(
                    "Tỷ lệ nợ dài hạn giảm; Current Ratio (Thanh toán ngắn hạn) tăng; Không phát hành thêm cổ phiếu (chống pha loãng).",
                    style={"fontSize": "0.85rem", "color": "#8b949e", "marginLeft": "20px", "marginBottom": "8px"}),

                html.Li([html.Strong(html.Span("3. Hiệu quả hoạt động:", style={"color": "#58a6ff"}))]),
                html.Div("Biên lợi nhuận gộp cải thiện so với năm trước; Vòng quay tài sản tăng.",
                         style={"fontSize": "0.85rem", "color": "#8b949e", "marginLeft": "20px",
                                "marginBottom": "8px"}),

            ], style={"fontSize": "0.95rem", "lineHeight": "1.4", "color": "#c9d1d9", "backgroundColor": "#0d1117",
                      "padding": "15px 15px 15px 35px", "borderRadius": "8px", "listStyleType": "none"}),
        ])

    # ==========================================================
    # 🚀 CHIẾN LƯỢC CANSLIM (ĐỘNG LƯỢNG TĂNG TRƯỞNG)
    # ==========================================================
    elif current_strategy == "STRAT_CANSLIM":
        title = "🚀 Siêu cổ phiếu (CANSLIM Proxy)"

        # Đọc trực tiếp biến số từ hệ thống Quant
        eps_q = CANSLIM_THRESHOLDS[CANSLIM_IDX_EPS_GROWTH_Q_MIN]
        rev_q = CANSLIM_THRESHOLDS[CANSLIM_IDX_REV_GROWTH_Q_MIN]
        eps_y = CANSLIM_THRESHOLDS[CANSLIM_IDX_EPS_GROWTH_Y_MIN]
        roe_min = CANSLIM_THRESHOLDS[CANSLIM_IDX_ROE_MIN]
        rs_min = CANSLIM_THRESHOLDS[CANSLIM_IDX_RS_MIN]
        vol_mult = CANSLIM_THRESHOLDS[CANSLIM_IDX_VOL_MULT]
        qr_min = CANSLIM_THRESHOLDS[CANSLIM_IDX_QUICK_RATIO_MIN]
        de_max = CANSLIM_THRESHOLDS[CANSLIM_IDX_DEBT_EQUITY_MAX]

        content = html.Div([
            html.Div([
                html.Span(dbc.Badge("Tác giả", color="primary", className="me-2")),
                "William J. O'Neil (Cập nhật: Richard Driehaus)"
            ], className="mb-2"),
            html.Div([
                html.Span(dbc.Badge("Nguồn", color="info", className="me-2")),
                html.A("Investopedia", href="https://www.investopedia.com/terms/c/canslim.asp", target="_blank",
                       style={"color": "#58a6ff", "textDecoration": "none"}),
                html.Span(" | ", style={"color": "#c9d1d9"}),
                html.A("Vietcap",
                       href="https://www.vietcap.com.vn/kien-thuc/huong-dan-thuc-hanh-canslim-phuong-phap-loc-co-phieu-hieu-qua",
                       target="_blank", style={"color": "#58a6ff", "textDecoration": "none"})
            ], className="mb-4"),

            html.Div([
                html.I(className="fas fa-triangle-exclamation", style={"color": "#e3b341", "marginRight": "8px"}),
                html.Strong("Minh bạch phương pháp: ", style={"color": "#e3b341"}),
                "Đây là bản CANSLIM PROXY (thích ứng), không phải CANSLIM chuẩn hóa đầy đủ. "
                "2 tiêu chí gốc của O'Neil được thay bằng proxy do thiếu dữ liệu: "
                "\"I - Institutional Sponsorship\" dùng Vốn hóa so với trung vị thị trường "
                "(không phải % sở hữu tổ chức thật), và \"M - Market Direction\" hiện chưa được "
                "mô hình hóa. Điểm số này là công cụ tham khảo, không phải phiên bản CANSLIM "
                "được kiểm định/backtest theo chuẩn gốc.",
            ], style={
                "fontSize": "0.85rem", "lineHeight": "1.6", "color": "#c9d1d9",
                "padding": "12px 14px", "background": "rgba(227,179,65,0.08)",
                "border": "1px solid rgba(227,179,65,0.3)", "borderRadius": "8px",
                "marginBottom": "16px",
            }),

            html.H5([html.I(className="fas fa-brain", style={"color": "#3fb950", "marginRight": "8px"}),
                     "Triết lý cốt lõi"], style={"color": "#e6edf3", "fontWeight": "bold"}),
            html.P(
                "Không dành cho người thích mua rẻ. CANSLIM kết hợp phân tích cơ bản và kỹ thuật để tìm kiếm cổ phiếu dẫn dắt (Leaders) đang bùng nổ lợi nhuận, được sự hậu thuẫn từ dòng tiền lớn. Chấp nhận mua cao để bán cao hơn tại các điểm Pivot.",
                style={"fontSize": "0.95rem", "lineHeight": "1.6", "color": "#c9d1d9"}),
            html.Hr(style={"borderColor": "#30363d"}),

            html.H5([html.I(className="fas fa-filter", style={"color": "#f85149", "marginRight": "8px"}),
                     "Logic Sàng Lọc (Hệ thống)"], style={"color": "#e6edf3", "fontWeight": "bold"}),
            html.Ul([
                html.Li([
                    html.Strong("C (Current): "), f"EPS Quý tăng >= {fmt(eps_q)}% & Doanh thu >= {fmt(rev_q)}%",
                    html.Div("Lợi nhuận quý hiện tại đột phá mạnh từ hoạt động cốt lõi.",
                             style={"fontSize": "0.85rem", "fontStyle": "italic", "color": "#8b949e",
                                    "marginTop": "2px"})
                ], style={"marginBottom": "8px"}),

                html.Li([
                    html.Strong("A (Annual): "), f"EPS Năm tăng >= {fmt(eps_y)}% & ROE >= {fmt(roe_min)}%",
                    html.Div("Tăng trưởng bền vững hàng năm và hiệu quả sử dụng vốn (ROE) xuất sắc.",
                             style={"fontSize": "0.85rem", "fontStyle": "italic", "color": "#8b949e",
                                    "marginTop": "2px"})
                ], style={"marginBottom": "8px"}),

                html.Li([
                    html.Strong("N (New): "), "Gần đỉnh 52 tuần",
                    html.Div("Sản phẩm mới, lãnh đạo mới, hoặc đang tích lũy chờ bứt phá đỉnh.",
                             style={"fontSize": "0.85rem", "fontStyle": "italic", "color": "#8b949e",
                                    "marginTop": "2px"})
                ], style={"marginBottom": "8px"}),

                html.Li([
                    html.Strong("S (Supply & Demand): "), f"Volume > {fmt(vol_mult)} lần trung bình 50 phiên",
                    html.Div("Cầu lớn hơn cung, dòng tiền đổ vào mạnh mẽ tại điểm Breakout.",
                             style={"fontSize": "0.85rem", "fontStyle": "italic", "color": "#8b949e",
                                    "marginTop": "2px"})
                ], style={"marginBottom": "8px"}),

                html.Li([
                    html.Strong("L (Leader): "), f"Sức mạnh giá (RS Rating) > {fmt(rs_min)}",
                    html.Div(f"Cổ phiếu thuộc Top {fmt(100 - rs_min)}% mạnh nhất thị trường.",
                             style={"fontSize": "0.85rem", "fontStyle": "italic", "color": "#8b949e",
                                    "marginTop": "2px"})
                ], style={"marginBottom": "8px"}),

                html.Li([
                    html.Strong("I (Institutional): "), "Thanh khoản cao",
                    html.Div("Loại bỏ các mã thanh khoản thấp để đảm bảo có dấu chân của Tổ chức lớn.",
                             style={"fontSize": "0.85rem", "fontStyle": "italic", "color": "#8b949e",
                                    "marginTop": "2px"})
                ], style={"marginBottom": "8px"}),

                html.Li([
                    html.Strong("M (Market): "), "Hệ thống tự động loại bỏ mã khi thị trường sập",
                    html.Div("Chỉ giao dịch khi thị trường chung đang trong xu hướng tăng (Uptrend).",
                             style={"fontSize": "0.85rem", "fontStyle": "italic", "color": "#8b949e",
                                    "marginTop": "2px"})
                ], style={"marginBottom": "8px"}),

            ], style={"fontSize": "0.95rem", "lineHeight": "1.4", "color": "#c9d1d9", "backgroundColor": "#0d1117",
                      "padding": "15px 15px 15px 35px", "borderRadius": "8px"}),
        ])
    # ==========================================================
    # 🌱 CHIẾN LƯỢC TĂNG TRƯỞNG BỀN VỮNG (GROWTH / FISHER)
    # ==========================================================
    elif current_strategy == "STRAT_GROWTH":
        title = "🌱 Tăng trưởng bền vững (Growth Investing)"

        # Đọc trực tiếp biến số từ hệ thống Quant
        rev_5y = FISHER_THRESHOLDS[FISHER_IDX_REV_GROWTH_5Y_MIN]
        dilution_max = FISHER_THRESHOLDS[FISHER_IDX_DILUTION_RATE_MAX]
        roe_min = FISHER_THRESHOLDS[FISHER_IDX_ROE_MIN]
        opex_max = FISHER_THRESHOLDS[FISHER_IDX_OPEX_EFF_MAX]
        turnover_min = FISHER_THRESHOLDS[FISHER_IDX_TURNOVER_MIN]
        reinvest_min = FISHER_THRESHOLDS[FISHER_IDX_REINVEST_MIN]

        content = html.Div([
            html.Div([
                html.Span(dbc.Badge("Tác giả", color="primary", className="me-2")),
                "Philip A. Fisher (Thomas Rowe Price Jr.)"
            ], className="mb-2"),
            html.Div([
                html.Span(dbc.Badge("Nguồn", color="info", className="me-2")),
                html.A("Common Stocks and Uncommon Profits (Vietcap)",
                       href="https://www.vietcap.com.vn/kien-thuc/dau-tu-tang-truong-la-gi-lam-sao-de-lua-chon-co-phieu-tang-truong",
                       target="_blank", style={"color": "#58a6ff", "textDecoration": "none"})
            ], className="mb-4"),

            html.H5([html.I(className="fas fa-brain", style={"color": "#3fb950", "marginRight": "8px"}),
                     "Triết lý cốt lõi"], style={"color": "#e6edf3", "fontWeight": "bold"}),
            html.P(
                "Tìm kiếm những 'gã khổng lồ tương lai' có khả năng tăng trưởng doanh thu và lợi nhuận vượt trội trong nhiều thập kỷ. Tập trung vào bộ máy quản lý, năng lực R&D và phương pháp 'Lời đồn đại' (Scuttlebutt). Phương châm: 'Thời điểm tốt nhất để bán cổ phiếu là gần như không bao giờ'.",
                style={"fontSize": "0.95rem", "lineHeight": "1.6", "color": "#c9d1d9"}),
            html.Hr(style={"borderColor": "#30363d"}),

            html.H5([html.I(className="fas fa-filter", style={"color": "#f85149", "marginRight": "8px"}),
                     "Logic Sàng Lọc (Hệ thống)"], style={"color": "#e6edf3", "fontWeight": "bold"}),
            html.Ul([
                html.Li([
                    html.Strong("Tăng trưởng bền vững: "), f"Doanh thu 5 năm tăng > {fmt(rev_5y)}%/năm",
                    html.Div("Lọc ra các công ty có đà tăng trưởng doanh thu dài hạn thay vì chỉ bùng nổ 1-2 năm.",
                             style={"fontSize": "0.85rem", "fontStyle": "italic", "color": "#8b949e",
                                    "marginTop": "2px"})
                ], style={"marginBottom": "8px"}),

                html.Li([
                    html.Strong("Hiệu quả quản trị: "), f"ROE > {fmt(roe_min)}%",
                    html.Div("Hiệu quả sử dụng vốn xuất sắc, bù đắp cho việc doanh nghiệp thường không trả cổ tức.",
                             style={"fontSize": "0.85rem", "fontStyle": "italic", "color": "#8b949e",
                                    "marginTop": "2px"})
                ], style={"marginBottom": "8px"}),

                html.Li([
                    html.Strong("Kỷ luật vốn (Chống pha loãng): "), f"Tỷ lệ pha loãng < {fmt(dilution_max)}%",
                    html.Div(
                        "Công ty dùng vốn nội tại để phát triển, tránh phát hành giấy liên tục làm loãng quyền lợi cổ đông.",
                        style={"fontSize": "0.85rem", "fontStyle": "italic", "color": "#8b949e", "marginTop": "2px"})
                ], style={"marginBottom": "8px"}),

                html.Li([
                    html.Strong("Quản trị chi phí: "), f"Opex Efficiency < {fmt(opex_max)} lần",
                    html.Div("Đảm bảo chi phí hoạt động không ăn mòn hết lợi nhuận biên.",
                             style={"fontSize": "0.85rem", "fontStyle": "italic", "color": "#8b949e",
                                    "marginTop": "2px"})
                ], style={"marginBottom": "8px"}),

                html.Li([
                    html.Strong("Năng lực tái đầu tư: "), f"Tỷ lệ giữ lại (Reinvest Rate) > {fmt(reinvest_min)}%",
                    html.Div("Doanh nghiệp ưu tiên giữ lại lợi nhuận để tái đầu tư R&D thay vì chia hết cho cổ đông.",
                             style={"fontSize": "0.85rem", "fontStyle": "italic", "color": "#8b949e",
                                    "marginTop": "2px"})
                ], style={"marginBottom": "8px"}),

                html.Li([
                    html.Strong("Thanh khoản an toàn: "), f"Giá trị giao dịch > {fmt(turnover_min)} VND",
                    html.Div("Loại bỏ các mã 'vô danh', rủi ro mất thanh khoản trên sàn IDX.",
                             style={"fontSize": "0.85rem", "fontStyle": "italic", "color": "#8b949e",
                                    "marginTop": "2px"})
                ], style={"marginBottom": "8px"}),

            ], style={"fontSize": "0.95rem", "lineHeight": "1.4", "color": "#c9d1d9", "backgroundColor": "#0d1117",
                      "padding": "15px 15px 15px 35px", "borderRadius": "8px"}),
        ])

    # ==========================================================
    # 🪄 CHIẾN LƯỢC CÔNG THỨC KỲ DIỆU (MAGIC FORMULA)
    # ==========================================================
    elif current_strategy == "STRAT_MAGIC":
        title = "🪄 Công Thức Kỳ Diệu (Magic Formula)"

        content = html.Div([
            html.Div([
                html.Span(dbc.Badge("Tác giả", color="primary", className="me-2")),
                "Joel Greenblatt"
            ], className="mb-2"),
            html.Div([
                html.Span(dbc.Badge("Nguồn", color="info", className="me-2")),
                html.A("Sách: The Little Book That Beats the Market",
                       href="https://www.quantifiedstrategies.com/the-magic-formula-strategy/", target="_blank",
                       style={"color": "#58a6ff", "textDecoration": "none"}),
                html.Span(" | ", style={"color": "#c9d1d9"}),
                html.A("Investopedia", href="https://www.investopedia.com/terms/m/magic-formula-investing.asp",
                       target="_blank", style={"color": "#58a6ff", "textDecoration": "none"})
            ], className="mb-4"),

            html.H5([html.I(className="fas fa-brain", style={"color": "#3fb950", "marginRight": "8px"}),
                     "Triết lý cốt lõi"], style={"color": "#e6edf3", "fontWeight": "bold"}),
            html.P(
                "Chiến lược định lượng lai có mục tiêu vô cùng đơn giản: Mua những doanh nghiệp 'Tốt' ở một mức giá 'Rẻ'. Triết lý này loại bỏ hoàn toàn cảm xúc bằng cách lượng hóa và chấm điểm dựa trên Tỷ suất sinh lời trên vốn (ROC - hoạt động hiệu quả) và Tỷ suất lợi tức (Earnings Yield - định giá rẻ). Đặc biệt, hệ thống kiên quyết loại bỏ ngành Tài chính (nợ là nguyên liệu kinh doanh làm sai lệch EV) và ngành Tiện ích (biên lợi nhuận bị kiểm soát).",
                style={"fontSize": "0.95rem", "lineHeight": "1.6", "color": "#c9d1d9"}),
            html.Hr(style={"borderColor": "#30363d"}),

            html.H5([html.I(className="fas fa-filter", style={"color": "#f85149", "marginRight": "8px"}),
                     "Logic Sàng Lọc (Hệ thống)"], style={"color": "#e6edf3", "fontWeight": "bold"}),
            html.Ul([
                html.Li([
                    html.Strong("Bước 1: Loại bỏ đặc thù: "), "Loại ngành Tài chính, Tiện ích & Vốn hóa nhỏ",
                    html.Div("Hệ thống tự động loại bỏ các mã có cấu trúc vốn sai lệch hoặc quy mô dưới 1.000 Tỷ VND.",
                             style={"fontSize": "0.85rem", "fontStyle": "italic", "color": "#8b949e",
                                    "marginTop": "2px"})
                ], style={"marginBottom": "8px"}),

                html.Li([
                    html.Strong("Bước 2: Doanh nghiệp 'TỐT': "), "Xếp hạng ROC từ cao xuống thấp",
                    html.Div("Công ty có tỷ suất sinh lời trên vốn càng cao thì hạng càng gần Top 1.",
                             style={"fontSize": "0.85rem", "fontStyle": "italic", "color": "#8b949e",
                                    "marginTop": "2px"})
                ], style={"marginBottom": "8px"}),

                html.Li([
                    html.Strong("Bước 3: Mức giá 'RẺ': "), "Xếp hạng Earnings Yield từ cao xuống thấp",
                    html.Div("Đại diện cho việc cổ phiếu đang được giao dịch ở mức định giá hấp dẫn nhất.",
                             style={"fontSize": "0.85rem", "fontStyle": "italic", "color": "#8b949e",
                                    "marginTop": "2px"})
                ], style={"marginBottom": "8px"}),

                html.Li([
                    html.Strong("Bước 4: Chốt hạ danh mục: "), "Top 20-30 mã có Tổng điểm thấp nhất",
                    html.Div(
                        "Tổng điểm = (Hạng ROC + Hạng EY). Điểm càng thấp chứng tỏ sự giao thoa càng hoàn hảo giữa 'Chất lượng' và 'Giá cả'.",
                        style={"fontSize": "0.85rem", "fontStyle": "italic", "color": "#8b949e", "marginTop": "2px"})
                ], style={"marginBottom": "8px"}),

            ], style={"fontSize": "0.95rem", "lineHeight": "1.4", "color": "#c9d1d9", "backgroundColor": "#0d1117",
                      "padding": "15px 15px 15px 35px", "borderRadius": "8px"}),
        ])
    elif current_strategy == "STRAT_NCN":
        title = "🛡️ Khẩu Vị Phòng Thủ "
        content = html.Div([
            # ── Thông tin tác giả ──
            html.Div([
                html.Span(dbc.Badge("Chuyên viên", color="warning", className="me-2")),
                html.Strong("Ngô Cao Nguyên", style={"color": "#e6edf3"}),
                html.Span(" · K16 · Chứng khoán",
                          style={"color": "#8b949e", "marginLeft": "8px", "fontSize": "0.85rem"}),
            ], className="mb-1"),
            html.Div([
                html.Span(dbc.Badge("Chuyên viên", color="warning", className="me-2")),
                html.Strong("Phan Đặng Anh Kiệt", style={"color": "#e6edf3"}),
                html.Span(" · K16 · Chứng khoán",
                          style={"color": "#8b949e", "marginLeft": "8px", "fontSize": "0.85rem"}),
            ], className="mb-1"),
            html.Div([
                html.Span(dbc.Badge("Chuyên viên", color="warning", className="me-2")),
                html.Strong("Cao Huỳnh Tuyết Trân", style={"color": "#e6edf3"}),
                html.Span(" · K16 · Chứng khoán",
                          style={"color": "#8b949e", "marginLeft": "8px", "fontSize": "0.85rem"}),
            ], className="mb-1"),
            html.Div([
                html.Span(dbc.Badge("Liên hệ tư vấn", color="success", className="me-2")),
                html.Span("0946 700 605 - Zalo / SMS - Để nhận báo cáo phân tích chuyên sâu & danh mục cá nhân hóa",
                          style={"color": "#8b949e", "fontSize": "0.85rem", "fontStyle": "italic"}),
            ], className="mb-4"),

            # ── Triết lý ──
            html.H5([html.I(className="fas fa-shield-alt",
                            style={"color": "#3fb950", "marginRight": "8px"}),
                     "Triết lý Đầu tư Phòng thủ"],
                    style={"color": "#e6edf3", "fontWeight": "bold"}),
            html.P([
                "Trong một thị trường đầy biến động, ", html.Strong("bảo toàn vốn"),
                " luôn là ưu tiên số 1. Khẩu vị này tìm kiếm các doanh nghiệp ",
                html.Strong("có lợi thế cạnh tranh bền vững"), " (Moat), ",
                html.Strong("dòng tiền thật"), " (không phải lợi nhuận kế toán), và ",
                html.Strong("ban lãnh đạo liêm chính"), " — những yếu tố mà nhiều nhà đầu tư ",
                "ngắn hạn bỏ qua nhưng lại ", html.Strong("tạo ra lợi nhuận vượt trội"),
                " trong chu kỳ 3–5 năm.",
            ], style={"fontSize": "0.95rem", "lineHeight": "1.7", "color": "#c9d1d9"}),
            html.Hr(style={"borderColor": "#30363d"}),

            # ── 3 trụ cột ──
            html.H5([html.I(className="fas fa-layer-group",
                            style={"color": "#f0883e", "marginRight": "8px"}),
                     "3 Trụ Cột Sàng Lọc"],
                    style={"color": "#e6edf3", "fontWeight": "bold"}),

            # Trụ 1
            html.Div([
                html.Div([
                    html.I(className="fas fa-ban", style={"color": "#f85149", "marginRight": "8px"}),
                    html.Strong("Tầng 1 · Loại bỏ ngay các Red Flag (Zero-tolerance)",
                                style={"color": "#f85149"}),
                ], style={"marginBottom": "6px"}),
                html.Ul([
                    html.Li([html.Code("CFO / Net Income ≥ 0.8", style={"color": "#79c0ff"}),
                             " (trung bình 3 năm) — lợi nhuận phải đi kèm dòng tiền thật"],
                            style={"marginBottom": "4px"}),
                    html.Li([html.Code("Dilution Rate ≤ 8%/năm", style={"color": "#79c0ff"}),
                             " — phát hành cổ phiếu liên tục là dấu hiệu rút ruột"],
                            style={"marginBottom": "4px"}),
                ], style={"fontSize": "0.9rem", "color": "#c9d1d9",
                          "paddingLeft": "20px", "marginBottom": "0"}),
            ], style={"backgroundColor": "#1c1c2e", "borderLeft": "3px solid #f85149",
                      "padding": "12px 16px", "borderRadius": "6px", "marginBottom": "12px"}),

            # Trụ 2
            html.Div([
                html.Div([
                    html.I(className="fas fa-coins", style={"color": "#3fb950", "marginRight": "8px"}),
                    html.Strong("Tầng 2 · Chất lượng Tài chính",
                                style={"color": "#3fb950"}),
                ], style={"marginBottom": "6px"}),
                html.Ul([
                    html.Li([html.Code("FCF / Tổng Nợ ≥ 0", style={"color": "#79c0ff"}),
                             " — dòng tiền tự do đủ để trả nợ (không đốt tiền)"],
                            style={"marginBottom": "4px"}),
                    html.Li([html.Code("ROIC ≥ 12%", style={"color": "#79c0ff"}),
                             " — tỷ suất sinh lời trên vốn đầu tư, thước đo Moat thực sự"],
                            style={"marginBottom": "4px"}),
                    html.Li([html.Code("Gross Margin ≥ 15%", style={"color": "#79c0ff"}),
                             " — quyền lực định giá, biên gộp cao hơn ngành = rào cản gia nhập"],
                            style={"marginBottom": "4px"}),
                    html.Li([html.Code("ROE ≥ 15%  |  Net Margin ≥ 5%  |  D/E ≤ 1.5",
                                       style={"color": "#79c0ff"})],
                            style={"marginBottom": "4px"}),
                ], style={"fontSize": "0.9rem", "color": "#c9d1d9",
                          "paddingLeft": "20px", "marginBottom": "0"}),
            ], style={"backgroundColor": "#0d1f15", "borderLeft": "3px solid #3fb950",
                      "padding": "12px 16px", "borderRadius": "6px", "marginBottom": "12px"}),

            # Trụ 3
            html.Div([
                html.Div([
                    html.I(className="fas fa-trophy", style={"color": "#f0883e", "marginRight": "8px"}),
                    html.Strong("Tầng 3 · Xếp hạng tổng hợp — Top 40 mã tốt nhất",
                                style={"color": "#f0883e"}),
                ], style={"marginBottom": "6px"}),
                html.P("Điểm NCN Score = tổng hợp có trọng số của ROIC (30%), ROE (20%), "
                       "Gross Margin (20%), Net Margin (15%), Chất lượng CFO (15%). "
                       "Chỉ những mã đã vượt Tầng 1 & 2 mới được xét xếp hạng.",
                       style={"fontSize": "0.9rem", "color": "#c9d1d9", "marginBottom": "0"}),
            ], style={"backgroundColor": "#1f1200", "borderLeft": "3px solid #f0883e",
                      "padding": "12px 16px", "borderRadius": "6px", "marginBottom": "16px"}),

            html.Hr(style={"borderColor": "#30363d"}),

            # ── Lưu ý & CTA ──
            html.Div([
                html.I(className="fas fa-info-circle",
                       style={"color": "#58a6ff", "marginRight": "8px"}),
                html.Span("Bộ lọc này phản ánh ",
                          style={"color": "#8b949e", "fontSize": "0.85rem"}),
                html.Strong("quan điểm đầu tư cá nhân",
                            style={"color": "#c9d1d9", "fontSize": "0.85rem"}),
                html.Span(" của chuyên viên và không phải khuyến nghị mua/bán chính thức. "
                          "Kết quả sàng lọc cần kết hợp với phân tích định tính trước khi ra quyết định.",
                          style={"color": "#8b949e", "fontSize": "0.85rem"}),
            ], style={"backgroundColor": "#0d1624", "borderLeft": "3px solid #58a6ff",
                      "padding": "10px 14px", "borderRadius": "6px"}),
        ])
    # ==========================================================
    # 🔥 CHIẾN LƯỢC ADX MOMENTUM (Wilder — bản cải tiến v3)
    # ==========================================================
    elif current_strategy == "STRAT_ADX_MOMENTUM":
        title = "🔥 ADX Momentum — Xu hướng & Siêu Cổ Phiếu"

        content = html.Div([
            html.Div([
                html.Span(dbc.Badge("Lý thuyết gốc", color="primary", className="me-2")),
                "J. Welles Wilder — Average Directional Index (ADX/+DI/-DI)"
            ], className="mb-4"),

            html.H5([html.I(className="fas fa-brain", style={"color": "#3fb950", "marginRight": "8px"}),
                     "Triết lý cốt lõi"], style={"color": "#e6edf3", "fontWeight": "bold"}),
            html.P(
                "ADX chỉ đo ĐỘ MẠNH của xu hướng, không đo HƯỚNG — phải kết hợp với +DI/-DI để biết "
                "xu hướng đang nghiêng về phía tăng hay giảm. Chiến lược này tìm các mã đang trong xu "
                "hướng tăng rõ rệt theo đúng chuẩn Wilder gốc, đồng thời tự loại trước các mã thanh "
                "khoản thấp/UPCoM dễ tạo tín hiệu ADX giả tạo do biến động giá không phản ánh dòng "
                "tiền thật.",
                style={"fontSize": "0.95rem", "lineHeight": "1.6", "color": "#c9d1d9"}),
            html.Hr(style={"borderColor": "#30363d"}),

            html.H5([html.I(className="fas fa-shield-alt", style={"color": "#f59e0b", "marginRight": "8px"}),
                     "Bước 1 — Gatekeeper chất lượng & thanh khoản (đặc thù thị trường VN)"],
                    style={"color": "#e6edf3", "fontWeight": "bold"}),
            html.P(
                "Áp dụng TRƯỚC khi xét ADX, không phụ thuộc chế độ giao dịch đang chọn: loại hoàn "
                "toàn sàn UPCoM · khối lượng TB 20 phiên ≥ 30.000 CP/ngày · giá đóng cửa ≥ 3.000 VNĐ "
                "· vốn hóa ≥ 200 tỷ VNĐ.",
                style={"fontSize": "0.85rem", "color": "#8b949e", "marginBottom": "12px"}),

            html.H5([html.I(className="fas fa-filter", style={"color": "#f85149", "marginRight": "8px"}),
                     "Bước 2 — Logic Sàng Lọc Chính"],
                    style={"color": "#e6edf3", "fontWeight": "bold"}),
            html.Ul([
                html.Li([
                    html.Strong(html.Span("Xu hướng Tăng vững vàng (Is_Steady_Uptrend):",
                                          style={"color": "#58a6ff"})),
                    html.Div(
                        "+DI(14) > -DI(14)  VÀ  ADX(14) ≥ 25 — đúng chuẩn Wilder gốc, không thêm "
                        "ràng buộc về động lượng của ADX. Mã có xu hướng tăng đã trưởng thành mà ADX "
                        "đi ngang ở vùng cao (không còn dốc lên) vẫn được tính là hợp lệ — đây mới là "
                        "điều kiện duy nhất quyết định mã có vào danh sách hay không.",
                        style={"fontSize": "0.85rem", "color": "#8b949e", "marginLeft": "20px",
                               "marginTop": "2px", "marginBottom": "10px"}),
                ]),
            ], style={"fontSize": "0.95rem", "lineHeight": "1.4", "color": "#c9d1d9",
                      "backgroundColor": "#0d1117", "padding": "15px 15px 15px 35px",
                      "borderRadius": "8px", "listStyleType": "none"}),

            html.H5([html.I(className="fas fa-sort-amount-down", style={"color": "#8b949e", "marginRight": "8px"}),
                     "Bước 3 — Sắp xếp kết quả"],
                    style={"color": "#e6edf3", "fontWeight": "bold"}),
            html.P([
                "Mã thỏa thêm tiêu chí ",
                html.Strong("Siêu Cổ Phiếu (Is_Super_Stock_ADX)", style={"color": "#58a6ff"}),
                " — (+DI>-DI VÀ ADX≥50) đúng ≥50% trong 20 phiên gần nhất — được xếp lên ĐẦU danh "
                "sách, sau đó sắp theo ADX(14) giảm dần. Đây chỉ là tiêu chí XẾP HẠNG tham khảo, "
                "không phải điều kiện loại trừ.",
            ], style={"fontSize": "0.85rem", "color": "#8b949e"}),

            html.Div([
                html.I(className="fas fa-info-circle", style={"color": "#58a6ff", "marginRight": "8px"}),
                html.Span(
                    "Số mã trả về phụ thuộc vào trạng thái chung của thị trường tại thời điểm lọc — "
                    "thị trường càng nhiều mã có dòng tiền dẫn dắt rõ rệt, danh sách càng dài.",
                    style={"color": "#8b949e", "fontSize": "0.85rem"}),
            ], style={"backgroundColor": "#0d1624", "borderLeft": "3px solid #58a6ff",
                      "padding": "10px 14px", "borderRadius": "6px", "marginTop": "10px"}),
        ])
    # (Bạn có thể sao chép block if elif trên cho 7 trường phái còn lại sau này)
    else:
        title = "Đang cập nhật..."
        content = html.P("Nội dung cho trường phái này đang được tổng hợp.", style={"color": "#c9d1d9"})

    return not is_open, title, content


# ============================================================================
# CALLBACK: EXPORT CSV
# ============================================================================
@app.callback(
    Output("download-csv", "data"),
    Input("btn-export-csv", "n_clicks"),
    State("screener-table", "rowData"),
    State("auth-store", "data"),          # ← thêm dòng này
    prevent_initial_call=True
)
def export_csv(n_clicks, row_data, auth_data):   # ← thêm auth_data
    if not row_data:
        return no_update

    # CHẶN SERVER-SIDE: chỉ B2B được export
    from src.callbacks.auth_callbacks import require_entitlement
    if not require_entitlement(auth_data, allowed_tiers=["b2b"]):
        logger.warning(f"[export_csv] Bị chặn — user '{(auth_data or {}).get('username')}' không phải B2B")
        return no_update

    df = pd.DataFrame(row_data)
    export_cols = [c for c in [
        'Ticker', 'Company Common Name', 'Sector', 'Price Close',
        'Perf_1W', 'Perf_1M', 'Perf_3M', 'Perf_1Y', 'Perf_YTD',
        'Market Cap', 'Volume', 'P/E', 'P/B', 'P/S', 'EV/EBITDA',
        'ROE (%)', 'ROA (%)', 'Net Margin (%)', 'Gross Margin (%)',
        'Revenue Growth YoY (%)', 'EPS Growth YoY (%)',
        'D/E', 'Current Ratio', 'Dividend Yield (%)',
        'RSI_14', 'MACD_Histogram', 'Beta', 'Alpha',
        'RS_1M', 'RS_3M', 'RS_1Y'
    ] if c in df.columns]
    from datetime import datetime
    filename = f"FSS_Screener_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    return dcc.send_data_frame(df[export_cols].to_csv, filename, index=False, encoding='utf-8-sig')


# ============================================================================
# CALLBACK: EXPORT EXCEL
# ============================================================================
@app.callback(
    Output("download-excel", "data"),
    Input("btn-export-excel", "n_clicks"),
    State("screener-table", "rowData"),
    State("auth-store", "data"),      # [FIX] thêm — trước đây không hề kiểm tra quyền
    prevent_initial_call=True
)
def export_excel(n_clicks, row_data, auth_data):
    # [FIX] BẢO VỆ SERVER-SIDE — B2B-ONLY theo đề xuất CFO: Export Excel là
    # tính năng cho Broker/Môi giới tải data thô, không mở cho Premium 199k
    # (khác export_excel bản trước dùng chung ["pro","b2b"]).
    from src.callbacks.auth_callbacks import require_entitlement
    if not require_entitlement(auth_data, allowed_tiers=["b2b"]):
        return no_update

    if not row_data:
        return no_update
    try:
        import io
        from datetime import datetime
        df = pd.DataFrame(row_data)

        export_cols = [c for c in [
            'Ticker', 'Company Common Name', 'Sector', 'Price Close',
            'Perf_1W', 'Perf_1M', 'Perf_3M', 'Perf_1Y', 'Perf_YTD',
            'Market Cap', 'Volume', 'P/E', 'P/B', 'P/S', 'EV/EBITDA',
            'ROE (%)', 'ROA (%)', 'Net Margin (%)', 'Gross Margin (%)',
            'Revenue Growth YoY (%)', 'EPS Growth YoY (%)',
            'D/E', 'Current Ratio', 'Dividend Yield (%)',
            'RSI_14', 'MACD_Histogram', 'Beta', 'Alpha',
            'RS_1M', 'RS_3M', 'RS_1Y'
        ] if c in df.columns]

        df_export = df[export_cols].copy()

        # Viết Excel vào buffer
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine='openpyxl') as writer:
            df_export.to_excel(writer, index=False, sheet_name='VSS Screener')

            # Style cơ bản
            wb = writer.book
            ws = writer.sheets['VSS Screener']

            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            # Header style — nền xanh navy, chữ trắng
            header_fill = PatternFill("solid", fgColor="0A1628")
            header_font = Font(bold=True, color="00D4FF", size=10, name="Calibri")
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin = Side(style='thin', color="1D4D80")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for cell in ws[1]:
                cell.fill   = header_fill
                cell.font   = header_font
                cell.alignment = header_align
                cell.border = border

            ws.row_dimensions[1].height = 28

            # Auto-fit column width
            for col_idx, col in enumerate(df_export.columns, 1):
                col_letter = get_column_letter(col_idx)
                max_len = max(len(str(col)), df_export[col].astype(str).str.len().max() if not df_export.empty else 0)
                ws.column_dimensions[col_letter].width = min(max_len + 4, 25)

            # Zebra striping rows
            fill_even = PatternFill("solid", fgColor="0D1B2A")
            fill_odd  = PatternFill("solid", fgColor="091526")
            font_row  = Font(size=9, name="Calibri", color="C9D1D9")

            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                fill = fill_even if row_idx % 2 == 0 else fill_odd
                for cell in row:
                    cell.fill      = fill
                    cell.font      = font_row
                    cell.alignment = Alignment(horizontal="right" if cell.column > 1 else "left", vertical="center")
                    cell.border    = border

            # Freeze header row
            ws.freeze_panes = "A2"

            # Sheet tab color
            ws.sheet_properties.tabColor = "00D4FF"

        buf.seek(0)
        filename = f"VSS_Screener_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return dcc.send_bytes(buf.read(), filename)

    except Exception as e:
        logger.error(f"Excel export error: {e}")
        return no_update


# ============================================================================
# CALLBACK: SUB-INDUSTRY FILTER — load options theo Sector đã chọn
# ============================================================================
@app.callback(
    Output("filter-sub-industry", "options"),
    Input("filter-all-industry", "value"),
    prevent_initial_call=False
)
def update_sub_industry_options(selected_sectors):
    try:
        from src.constants.gics_translation import GICS_INDUSTRY_TRANSLATION
        # ✅ Mới — nhanh hơn ~3x
        from src.backend.data_loader import get_snapshot_df
        df = get_snapshot_df().copy()

        # Tìm cột ngành con
        sub_col = next((c for c in ['GICS Industry Name', 'GICS Sub-Industry Name'] if c in df.columns), None)
        if not sub_col:
            return [{"label": "Tất cả ngành con", "value": "all"}]

        # Lọc theo sector đã chọn
        if selected_sectors and selected_sectors != ["all"] and "all" not in selected_sectors:
            sec_col = next((c for c in ['Sector', 'GICS Sector Name'] if c in df.columns), None)
            if sec_col:
                df = df[df[sec_col].isin(selected_sectors)]

        subs = sorted(df[sub_col].dropna().unique().tolist())
        options = [{"label": "Tất cả ngành con", "value": "all"}]
        options += [
            {"label": GICS_INDUSTRY_TRANSLATION.get(s, s), "value": s}
            for s in subs if s not in ('nan', '0', 'Chưa phân loại')
        ]
        return options
    except Exception:
        return [{"label": "Tất cả ngành con", "value": "all"}]


# ============================================================================
# CALLBACK: ÁP DỤNG SUB-INDUSTRY FILTER
# ĐÃ ĐƯỢC GỘP VÀO update_screener_table (Bug #2 fix).
# Callback riêng này bị vô hiệu hóa để tránh overwrite rowData của callback chính.
# ============================================================================
# @app.callback(
#     Output("screener-table", "rowData", allow_duplicate=True),
#     Input("filter-sub-industry", "value"),
#     State("screener-table", "rowData"),
#     prevent_initial_call=True
# )
# def apply_sub_industry_filter(selected_subs, row_data):
#     ... (đã gộp vào update_screener_table)


# ============================================================================
# CALLBACK: WATCHLIST — Thêm/xóa mã + hiển thị modal
# ============================================================================
@app.callback(
    Output("watchlist-modal", "is_open"),
    Output("watchlist-content", "children"),
    Output("watchlist-store", "data", allow_duplicate=True),
    Input("btn-watchlist", "n_clicks"),
    Input("btn-close-watchlist", "n_clicks"),
    Input("screener-table", "selectedRows"),
    State("watchlist-store", "data"),
    State("watchlist-modal", "is_open"),
    prevent_initial_call=True
)
def manage_watchlist(n_open, n_close, selected_rows, watchlist, is_open):
    ctx = callback_context
    if not ctx.triggered:
        return no_update, no_update, no_update

    trigger = ctx.triggered[0]['prop_id']
    watchlist = watchlist or []

    # Đóng modal
    if 'btn-close-watchlist' in trigger:
        return False, no_update, no_update

    # Thêm mã từ bảng vào watchlist
    if 'selectedRows' in trigger and selected_rows:
        ticker = selected_rows[0].get('Ticker', '')
        if ticker and ticker not in watchlist:
            watchlist = watchlist + [ticker]
        return no_update, no_update, watchlist

    # Mở modal + hiển thị watchlist
    if 'btn-watchlist' in trigger:
        if not watchlist:
            content = html.Div([
                html.I(className="fas fa-star", style={"fontSize": "32px", "color": "#484f58", "marginBottom": "12px"}),
                html.P("Chưa có mã nào trong danh sách theo dõi.", style={"color": "#6e7681"}),
                html.P("Click vào một mã trong bảng để thêm vào watchlist.",
                       style={"color": "#484f58", "fontSize": "12px"}),
            ], style={"textAlign": "center", "padding": "40px 0"})
        else:
            rows = []
            from src.backend.data_loader import get_snapshot_df
            try:
                df_snap = get_snapshot_df().copy()
            except Exception:
                df_snap = pd.DataFrame()

            for ticker in watchlist:
                row_data_snap = {}
                if not df_snap.empty and 'Ticker' in df_snap.columns:
                    match = df_snap[df_snap['Ticker'] == ticker]
                    if not match.empty:
                        row_data_snap = match.iloc[0].to_dict()

                price = row_data_snap.get('Price Close', '–')
                perf1w = row_data_snap.get('Perf_1W', None)
                sector = row_data_snap.get('Sector', '–')

                perf_color = '#10b981' if isinstance(perf1w, (int, float)) and perf1w > 0 else '#ef4444'
                perf_str = f"+{perf1w:.1f}%" if isinstance(perf1w, (int, float)) and perf1w > 0 else (
                    f"{perf1w:.1f}%" if isinstance(perf1w, (int, float)) else '–')

                rows.append(html.Div([
                    html.Span(ticker, style={"fontFamily": "'JetBrains Mono', monospace", "fontWeight": "700",
                                             "color": "#1E88E5", "width": "80px", "display": "inline-block"}),
                    html.Span(sector, style={"color": "#7fa8cc", "fontSize": "12px", "width": "160px",
                                             "display": "inline-block"}),
                    html.Span(f"{price:,.0f}" if isinstance(price, (int, float)) else str(price),
                              style={"fontFamily": "'JetBrains Mono', monospace", "color": "#d6eaf8", "width": "100px",
                                     "display": "inline-block", "textAlign": "right"}),
                    html.Span(perf_str,
                              style={"fontFamily": "'JetBrains Mono', monospace", "color": perf_color, "width": "80px",
                                     "display": "inline-block", "textAlign": "right"}),
                    html.I(className="fas fa-times", id={"type": "watchlist-remove", "ticker": ticker},
                           style={"color": "#ef4444", "cursor": "pointer", "marginLeft": "16px", "fontSize": "12px"},
                           n_clicks=0),
                ], style={"display": "flex", "alignItems": "center", "padding": "10px 16px",
                          "borderBottom": "1px solid #0e2540", "gap": "8px"}))

            content = html.Div([
                html.Div([
                    html.Span("MÃ CK",
                              style={"width": "80px", "display": "inline-block", "fontSize": "10px", "color": "#5a8ab0",
                                     "fontWeight": "700", "letterSpacing": "1px"}),
                    html.Span("NGÀNH", style={"width": "160px", "display": "inline-block", "fontSize": "10px",
                                              "color": "#5a8ab0", "fontWeight": "700", "letterSpacing": "1px"}),
                    html.Span("GIÁ", style={"width": "100px", "display": "inline-block", "fontSize": "10px",
                                            "color": "#5a8ab0", "fontWeight": "700", "letterSpacing": "1px",
                                            "textAlign": "right"}),
                    html.Span("%1T",
                              style={"width": "80px", "display": "inline-block", "fontSize": "10px", "color": "#5a8ab0",
                                     "fontWeight": "700", "letterSpacing": "1px", "textAlign": "right"}),
                ], style={"display": "flex", "padding": "8px 16px", "backgroundColor": "#040d18",
                          "borderRadius": "6px 6px 0 0"}),
                *rows,
                html.Div([
                    html.Small(f"Tổng cộng {len(watchlist)} mã đang theo dõi", style={"color": "#484f58"}),
                ], style={"padding": "10px 16px", "borderTop": "1px solid #0e2540"}),
            ])

        return True, content, watchlist

    return no_update, no_update, no_update


@app.callback(
    Output("watchlist-store",   "data",     allow_duplicate=True),
    Output("watchlist-content", "children", allow_duplicate=True),
    Input("btn-clear-watchlist", "n_clicks"),
    prevent_initial_call=True
)
def clear_watchlist(n_clicks):
    if not n_clicks:
        return no_update, no_update
    empty = html.Div([
        html.I(className="fas fa-star",
               style={"fontSize": "32px", "color": "#484f58", "marginBottom": "12px"}),
        html.P("Chưa có mã nào trong danh sách theo dõi.", style={"color": "#6e7681"}),
        html.P("Click vào một mã trong bảng để thêm vào watchlist.",
               style={"color": "#484f58", "fontSize": "12px"}),
    ], style={"textAlign": "center", "padding": "40px 0"})
    return [], empty


# add_forward_pe đã được gộp vào update_screener_table dưới dạng hàm _add_forward_pe()
# để tránh double-render khi rowData thay đổi trigger callback riêng lẻ.

# ============================================================================
# CALLBACK: TOGGLE HEALTH METHODOLOGY MODAL
# ============================================================================
@app.callback(
    Output("health-methodology-modal", "is_open"),
    Input("btn-health-methodology", "n_clicks"),
    State("health-methodology-modal", "is_open"),
    prevent_initial_call=True,
)
def toggle_health_methodology_modal(n, is_open):
    if n:
        return not is_open
    return is_open

# ============================================================================
# CALLBACK: TOGGLE FSS SMART RANK METHODOLOGY MODAL
# ============================================================================
@app.callback(
    Output("modal-fss-methodology", "is_open"),
    Input("btn-fss-methodology", "n_clicks"),
    State("modal-fss-methodology", "is_open"),
    prevent_initial_call=True,
)
def toggle_fss_methodology_modal(n, is_open):
    if n:
        return not is_open
    return is_open

# ============================================================================
# CALLBACK: XÓA TỪNG MÃ KHỎI WATCHLIST
# ============================================================================
@app.callback(
    Output("watchlist-store",   "data",     allow_duplicate=True),
    Output("watchlist-content", "children", allow_duplicate=True),
    Input({"type": "watchlist-remove", "ticker": ALL}, "n_clicks"),
    State("watchlist-store", "data"),
    prevent_initial_call=True,
)
def remove_watchlist_ticker(n_clicks_list, watchlist):
    from dash import callback_context
    ctx = callback_context
    if not ctx.triggered or not any(n for n in (n_clicks_list or []) if n):
        return no_update, no_update

    # Tìm ticker nào được click
    triggered_id = ctx.triggered[0]["prop_id"].split(".")[0]
    try:
        import json
        ticker_to_remove = json.loads(triggered_id)["ticker"]
    except Exception:
        return no_update, no_update

    watchlist = [t for t in (watchlist or []) if t != ticker_to_remove]

    if not watchlist:
        empty = html.Div([
            html.I(className="fas fa-star",
                   style={"fontSize": "32px", "color": "#484f58", "marginBottom": "12px"}),
            html.P("Chưa có mã nào trong danh sách theo dõi.", style={"color": "#6e7681"}),
            html.P("Click vào một mã trong bảng để thêm vào watchlist.",
                   style={"color": "#484f58", "fontSize": "12px"}),
        ], style={"textAlign": "center", "padding": "40px 0"})
        return watchlist, empty
    from src.backend.data_loader import get_snapshot_df
    # Re-render danh sách còn lại
    try:
        df_snap = get_snapshot_df().copy()
    except Exception:
        df_snap = pd.DataFrame()

    rows = []
    for ticker in watchlist:
        row_data_snap = {}
        if not df_snap.empty and "Ticker" in df_snap.columns:
            match = df_snap[df_snap["Ticker"] == ticker]
            if not match.empty:
                row_data_snap = match.iloc[0].to_dict()

        price   = row_data_snap.get("Price Close", "–")
        perf1w  = row_data_snap.get("Perf_1W", None)
        sector  = row_data_snap.get("Sector", "–")
        p_color = "#10b981" if isinstance(perf1w, (int, float)) and perf1w > 0 else "#ef4444"
        p_str   = (f"+{perf1w:.1f}%" if isinstance(perf1w, (int, float)) and perf1w > 0
                   else (f"{perf1w:.1f}%" if isinstance(perf1w, (int, float)) else "–"))

        rows.append(html.Div([
            html.Span(ticker, style={"fontFamily": "'JetBrains Mono', monospace", "fontWeight": "700",
                                     "color": "#1E88E5", "width": "80px", "display": "inline-block"}),
            html.Span(sector, style={"color": "#7fa8cc", "fontSize": "12px", "width": "160px",
                                     "display": "inline-block"}),
            html.Span(f"{price:,.0f}" if isinstance(price, (int, float)) else str(price),
                      style={"fontFamily": "'JetBrains Mono', monospace", "color": "#d6eaf8",
                             "width": "100px", "display": "inline-block", "textAlign": "right"}),
            html.Span(p_str, style={"fontFamily": "'JetBrains Mono', monospace", "color": p_color,
                                    "width": "80px", "display": "inline-block", "textAlign": "right"}),
            html.I(className="fas fa-times",
                   id={"type": "watchlist-remove", "ticker": ticker},
                   n_clicks=0,
                   style={"color": "#ef4444", "cursor": "pointer",
                          "marginLeft": "16px", "fontSize": "12px"}),
        ], style={"display": "flex", "alignItems": "center", "padding": "10px 16px",
                  "borderBottom": "1px solid #0e2540", "gap": "8px"}))

    content = html.Div([
        html.Div([
            html.Span("MÃ CK",  style={"width":"80px","display":"inline-block","fontSize":"10px","color":"#5a8ab0","fontWeight":"700","letterSpacing":"1px"}),
            html.Span("NGÀNH",  style={"width":"160px","display":"inline-block","fontSize":"10px","color":"#5a8ab0","fontWeight":"700","letterSpacing":"1px"}),
            html.Span("GIÁ",    style={"width":"100px","display":"inline-block","fontSize":"10px","color":"#5a8ab0","fontWeight":"700","letterSpacing":"1px","textAlign":"right"}),
            html.Span("%1T",    style={"width":"80px","display":"inline-block","fontSize":"10px","color":"#5a8ab0","fontWeight":"700","letterSpacing":"1px","textAlign":"right"}),
        ], style={"display":"flex","padding":"8px 16px","backgroundColor":"#040d18","borderRadius":"6px 6px 0 0"}),
        *rows,
        html.Div([
            html.Small(f"Tổng cộng {len(watchlist)} mã đang theo dõi", style={"color": "#484f58"}),
        ], style={"padding": "10px 16px", "borderTop": "1px solid #0e2540"}),
    ])
    return watchlist, content

# ============================================================================
# FIX YC2: Reset selectedRows khi đóng detail-modal
# → Cho phép click đúp lại vào cùng 1 mã để mở modal lần 2
# Nguyên nhân bug: AG Grid chỉ trigger callback khi selectedRows thay đổi.
# Khi đóng modal, selectedRows vẫn còn giữ mã cũ → click lại không fire.
# Fix: khi modal đóng (is_open = False) → reset selectedRows về [] ngay lập tức.
# ============================================================================
@app.callback(
    Output("screener-table", "selectedRows", allow_duplicate=True),
    Input("detail-modal", "is_open"),
    prevent_initial_call=True
)
def reset_selected_rows_on_modal_close(is_open):
    if not is_open:
        return []
    raise PreventUpdate

# Thêm callback sync Store + Sửa logic lọc null
@app.callback(
    Output("include-null-data-store", "data"),
    Input("include-null-data-toggle", "value"),
    prevent_initial_call=True,
)
def sync_include_null_store(toggle_val):
    """Đồng bộ trạng thái toggle vào Store để các callback khác đọc được."""
    return bool(toggle_val)

# ============================================================================
# TỰ ĐỘNG CẬP NHẬT OPTIONS CHO DROPDOWN NGÀNH (GIỮ NGUYÊN Ô SÀN CỦA SIDEBAR)
# ============================================================================
@app.callback(
    Output("filter-all-industry", "options"), # Chỉ update ngành, bỏ qua Sàn
    Input("search-ticker-input", "id"), 
    prevent_initial_call=False
)
def auto_update_dropdowns(_):
    try:
        df = get_snapshot_df()
        sec_options = [{"label": "Tất cả ngành", "value": "all"}]
        
        if df is not None and not df.empty:
            # Quét danh sách Ngành thực tế
            sec_col = next((c for c in ['Sector', 'GICS Sector Name'] if c in df.columns), None)
            if sec_col:
                sectors = sorted(df[sec_col].dropna().unique().tolist())
                sec_options += [
                    {"label": translate_gics_sector(s), "value": s}
                    for s in sectors if str(s).strip() not in ('nan', '0', 'Chưa phân loại', '')
                ]
        return sec_options
    except Exception as e:
        logger.error(f"Lỗi auto load dropdowns: {e}")
        return [{"label": "Tất cả ngành", "value": "all"}]

# SAU:
@app.callback(
    Output("data-cutoff-label", "children"),
    Input("screener-table",          "rowData"),
    Input("realtime-fetch-ts",       "data"),       # ← trigger thêm khi fetch xong
    prevent_initial_call=False,
)
def update_cutoff_label(row_data, fetch_ts):
    import os, pandas as pd
    from datetime import datetime
    import pytz

    _TZ_VN = pytz.timezone("Asia/Ho_Chi_Minh")

    try:
        # Ưu tiên 1: Nếu vừa có fetch realtime thì hiện giờ VN hiện tại
        if fetch_ts and float(fetch_ts) > 0:
            try:
                from src.backend.wifeed_updater import get_snapshot_timestamp
                ts = get_snapshot_timestamp()
                if ts > 0:
                    dt_vn  = datetime.fromtimestamp(ts, tz=_TZ_VN)
                    return [
                        html.Span("● LIVE", style={
                            "color":       "#00e676",
                            "fontWeight":  "700",
                            "fontSize":    "9px",
                            "marginRight": "5px",
                            "letterSpacing": "0.1em",
                            "animation":   "realtime-pulse 2s ease-in-out infinite",
                        }),
                        html.Span(
                            f"{dt_vn.strftime('%H:%M')} · {dt_vn.strftime('%d/%m/%Y')}",
                            style={"color": "#5a8ab0", "fontSize": "11px"},
                        ),
                    ]
            except Exception:
                pass

        # Ưu tiên 2: Ngày từ global var data_loader
        from src.backend.data_loader import get_data_cutoff_date
        d = get_data_cutoff_date()
        if d:
            # d có dạng "dd/mm/yyyy" — thêm giờ load nếu có
            now_vn = datetime.now(_TZ_VN)
            return f"EOD {d} · tải {now_vn.strftime('%H:%M')}"

        # Fallback 3: Đọc từ parquet
        parquet = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
            "data", "processed", "market_prices.parquet"
        )
        if os.path.exists(parquet):
            df_tmp   = pd.read_parquet(parquet, columns=["Date"])
            max_date = pd.to_datetime(df_tmp["Date"]).max()
            if pd.notna(max_date):
                now_vn = datetime.now(_TZ_VN)
                return f"EOD {max_date.strftime('%d/%m/%Y')} · tải {now_vn.strftime('%H:%M')}"

        return ""

    except Exception:
        return ""

@app.callback(
    [Output("action-buttons-container", "style"),
     Output("toggle-actions-icon", "className")],
    [Input("btn-toggle-actions", "n_clicks")],
    [State("action-buttons-container", "style")],
    prevent_initial_call=True
)
def toggle_action_buttons(n_clicks, current_style):
    base_style = {"gap": "8px", "alignItems": "center"}
    
    if current_style and current_style.get("display") == "none":
        # Mở rộng (hiện dải nút)
        base_style["display"] = "flex"
        return base_style, "fas fa-angles-right"
    else:
        # Thu gọn lại
        base_style["display"] = "none"
        return base_style, "fas fa-angles-left"

from dash import clientside_callback

# Tự động format dấu phẩy khi người dùng nhập số tiền
clientside_callback(
    """
    function(val) {
        if (!val) return val;
        // Xóa hết các ký tự không phải là số (chỉ giữ lại số)
        let numStr = val.toString().replace(/\\D/g, "");
        // Thêm dấu phẩy phân cách hàng nghìn
        return numStr.replace(/\\B(?=(\\d{3})+(?!\\d))/g, ",");
    }
    """,
    Output("nav-input", "value"),
    Input("nav-input", "value"),
    prevent_initial_call=True
)

@app.callback(
    Output("about-fss-modal", "is_open"),
    Input("btn-about-fss", "n_clicks"),
    prevent_initial_call=True,
)
def open_about_fss_modal(n):
    return True if n else no_update

from dash import Input, Output, callback, no_update

# ============================================================
# CALLBACK: QUY ĐỔI SỐ NAV THÀNH TỶ ĐỒNG TRÊN MODAL
# ============================================================
@callback(
    Output("modal-nav-formatted-display", "children"),
    Input("modal-nav-input", "value")
)
def update_nav_display(nav_val):
    if nav_val is None or nav_val == "":
        return "(= 0.00 Tỷ VNĐ)"
    
    try:
        val = float(nav_val)
        return f"(= {val / 1e9:,.2f} Tỷ VNĐ)"
    except Exception:
        return "(Lỗi định dạng)"